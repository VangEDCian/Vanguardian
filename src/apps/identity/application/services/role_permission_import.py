import re
from dataclasses import dataclass, field

from django.db import transaction
from django.utils.translation import gettext_lazy as _
from openpyxl import load_workbook

from apps.identity.models import IdentityPermission, Role, RoleScopeLevel  # noqa: DDD022

_REQUIRED_COLUMNS = {
    "role_name",
    "permission",
}


@dataclass
class RolePermissionImportResult:
    total_rows: int = 0
    imported_rows: int = 0
    skipped_rows: int = 0
    created_roles: int = 0
    updated_roles: int = 0
    role_permission_links: int = 0
    issues: list[str] = field(default_factory=list)

    def as_dict(self):
        return {
            "total_rows": self.total_rows,
            "imported_rows": self.imported_rows,
            "skipped_rows": self.skipped_rows,
            "created_roles": self.created_roles,
            "updated_roles": self.updated_roles,
            "role_permission_links": self.role_permission_links,
            "issues": self.issues,
        }


class IdentityRolePermissionImportService:
    def build_role_create_options(self):
        return {
            "scope_options": [
                {"value": RoleScopeLevel.STUDY, "label": _("Study")},
                {"value": RoleScopeLevel.STUDY_SITE, "label": _("Study site")},
            ],
            "permission_options": [
                {
                    "value": str(permission.pk),
                    "label": f"{self.permission_code_for(permission)} - {permission.name}",
                }
                for permission in IdentityPermission.objects.order_by("app_label", "codename")
            ],
        }

    @transaction.atomic
    def create_role(
        self,
        *,
        study_id: int,
        name: str,
        code: str = "",
        description: str = "",
        scope_level: str = RoleScopeLevel.STUDY_SITE,
        permission_ids=(),
    ):
        normalized_name = str(name or "").strip()
        if not normalized_name:
            raise ValueError(_("Role name is required."))

        normalized_code = str(code or "").strip() or self._role_code_from_name(normalized_name)
        normalized_scope = RoleScopeLevel(scope_level or RoleScopeLevel.STUDY_SITE)
        if Role.objects.filter(study_id=study_id, name=normalized_name).exists():
            raise ValueError(_("This role name already exists for the study."))
        if Role.objects.filter(
            study_id=study_id,
            code=normalized_code,
            scope_level=normalized_scope,
            version_no=1,
        ).exists():
            raise ValueError(_("This role code already exists for the selected scope."))

        permissions = self._objects_by_ids(
            IdentityPermission.objects.order_by("app_label", "codename"),
            permission_ids,
        )
        role = Role.objects.create(
            study_id=study_id,
            name=normalized_name,
            code=normalized_code,
            description=str(description or "").strip()[:255],
            scope_level=normalized_scope,
        )
        role.permissions.set(permissions)
        return {
            "id": role.pk,
            "name": role.name,
            "permission_count": len(permissions),
        }

    @transaction.atomic
    def import_workbook(self, *, study_id: int, import_file):
        rows = self._read_rows(import_file)
        result = RolePermissionImportResult(total_rows=len(rows))

        for row_number, row in rows:
            role_name = row.get("role_name", "").strip()
            permission_key = row.get("permission", "").strip()
            description = row.get("description", "").strip()
            scope_level = self._normalize_scope_level(row.get("scope_level", ""))

            if not role_name or not permission_key:
                result.skipped_rows += 1
                result.issues.append(f"Row {row_number}: missing role_name or permission.")
                continue
            if scope_level is None:
                result.skipped_rows += 1
                result.issues.append(f"Row {row_number}: invalid scope_level '{row.get('scope_level', '')}'.")
                continue
            permission = self._permission_by_code(permission_key)
            if permission is None:
                result.skipped_rows += 1
                result.issues.append(f"Row {row_number}: permission '{permission_key}' does not exist.")
                continue

            role, created = Role.objects.get_or_create(
                study_id=study_id,
                name=role_name,
                defaults={
                    "code": self._role_code_from_name(role_name),
                    "description": description[:255],
                    "scope_level": scope_level,
                },
            )
            if created:
                result.created_roles += 1
            else:
                result.updated_roles += self._update_role_metadata(
                    role,
                    description=description,
                    scope_level=scope_level,
                )

            if not role.permissions.filter(pk=permission.pk).exists():
                role.permissions.add(permission)
                result.role_permission_links += 1
            result.imported_rows += 1

        return result.as_dict()

    def build_summary(self, *, study_id: int):
        roles = [
            {
                "name": role.name,
                "scope_level": role.scope_level,
                "description": role.description,
                "permission_count": role.permissions.count(),
                "permission_codes": [
                    self.permission_code_for(permission)
                    for permission in role.permissions.all()
                ],
            }
            for role in Role.objects.filter(study_id=study_id)
            .prefetch_related("permissions")
            .order_by("name")
        ]
        return {"roles": roles, "role_count": len(roles)}

    def _objects_by_ids(self, queryset, raw_ids):
        normalized_ids = self._normalize_ids(raw_ids)
        if not normalized_ids:
            return []
        objects = list(queryset.filter(pk__in=normalized_ids))
        if len(objects) != len(normalized_ids):
            raise ValueError(_("One or more selected permissions no longer exist."))
        return objects

    def _read_rows(self, import_file):
        filename = str(getattr(import_file, "name", "") or "").lower()
        if not filename.endswith((".xlsx", ".xlsm")):
            raise ValueError(_("Only .xlsx and .xlsm files are supported."))
        return self._read_excel_rows(import_file)

    def _read_excel_rows(self, import_file):
        workbook = load_workbook(import_file, read_only=True, data_only=True)
        sheet = workbook.active
        rows = list(sheet.iter_rows(values_only=True))
        if not rows:
            return []
        headers = [self._normalize_header(value) for value in rows[0]]
        dict_rows = []
        for raw_row in rows[1:]:
            row = {headers[index]: "" if value is None else str(value).strip() for index, value in enumerate(raw_row) if index < len(headers) and headers[index]}
            dict_rows.append(row)
        return self._normalize_dict_rows(dict_rows)

    def _normalize_dict_rows(self, rows):
        normalized_rows = []
        for row_number, raw_row in enumerate(rows, start=2):
            row = {self._normalize_header(key): str(value or "").strip() for key, value in raw_row.items()}
            missing = _REQUIRED_COLUMNS.difference(row)
            if missing:
                missing_text = ", ".join(sorted(missing))
                raise ValueError(f"Missing required import columns: {missing_text}")
            normalized_rows.append((row_number, row))
        return normalized_rows

    @staticmethod
    def _normalize_header(value):
        normalized = re.sub(r"[^a-z0-9]+", "_", str(value or "").strip().lower()).strip("_")
        return normalized

    @staticmethod
    def _role_code_from_name(value):
        return re.sub(r"[^A-Z0-9]+", "_", str(value or "").strip().upper()).strip("_")

    @staticmethod
    def _normalize_scope_level(value):
        normalized = str(value or RoleScopeLevel.STUDY_SITE).strip().upper()
        if normalized in {"STUDY SITE", "STUDY-SITE"}:
            normalized = RoleScopeLevel.STUDY_SITE
        if normalized in RoleScopeLevel.values:
            return RoleScopeLevel(normalized)
        return None

    @staticmethod
    def permission_code_for(permission: IdentityPermission) -> str:
        return permission.permission_code

    @staticmethod
    def _permission_by_code(permission_code: str):
        normalized_code = str(permission_code or "").strip()
        permission = IdentityPermission.objects.filter(codename=normalized_code).first()
        if permission or "." not in normalized_code:
            return permission
        app_label, codename = normalized_code.split(".", 1)
        return IdentityPermission.objects.filter(
            app_label=app_label.strip(),
            codename=codename.strip(),
        ).first()

    @staticmethod
    def _normalize_ids(raw_ids):
        normalized_ids = []
        for raw_id in raw_ids or ():
            try:
                normalized_id = int(raw_id)
            except (TypeError, ValueError):
                raise ValueError(_("Invalid permission selection.")) from None
            if normalized_id not in normalized_ids:
                normalized_ids.append(normalized_id)
        return normalized_ids

    def _update_role_metadata(self, role, *, description, scope_level):
        normalized_description = str(description or "").strip()[:255]
        update_fields = []
        if normalized_description and role.description != normalized_description:
            role.description = normalized_description
            update_fields.append("description")
        if scope_level and role.scope_level != scope_level:
            role.scope_level = scope_level
            update_fields.append("scope_level")
        if not update_fields:
            return 0
        role.save(update_fields=update_fields)
        return 1
