from dataclasses import dataclass
from io import BytesIO

from django.db import transaction
from django.utils import timezone

from apps.crf.domain.exceptions import StudyScopeViolationError
from apps.crf.public import (
    CrfContextAdapter,
    CrfTemplateAmbiguousError,
    CrfTemplateNotFoundError,
)
from apps.shared.constants import EventFormEntryModeChoices
from apps.study.application.exceptions import (
    EventFormBindingImportDependencyError,
    EventFormBindingImportFormatError,
)
from apps.study.application.services.event_form_display_label import (
    EventFormDisplayLabelService,
    EventFormDisplayLabelValidationError,
)
from apps.study.infrastructure.repositories import DjangoStudyEventRepository


@dataclass(frozen=True)
class ImportStudyEventFormBindingsTemplateCommand:
    actor_user_id: int
    selected_study_id: int
    study_id: int
    file_name: str
    file_content: bytes


@dataclass(frozen=True)
class EventFormBindingImportIssue:
    row_number: int
    event_code: str
    form_code: str
    reason: str


@dataclass(frozen=True)
class ImportStudyEventFormBindingsTemplateResult:
    total_rows: int
    created_count: int
    updated_count: int
    skipped_count: int
    issues: tuple[EventFormBindingImportIssue, ...] = ()
    warnings: tuple[str, ...] = ()


@dataclass(frozen=True)
class _PreparedEventFormBindingImportRow:
    row_number: int
    event_code: str
    form_code: str
    event_definition: object
    form_definition: object
    display_order: int
    is_repeatable_within_event: bool
    role_scope: str | None
    entry_mode: str | None
    display_label_config: dict | None


class ImportStudyEventFormBindingsTemplateService:
    crf_context_adapter_class = CrfContextAdapter
    repository_class = DjangoStudyEventRepository
    display_label_service_class = EventFormDisplayLabelService
    expected_columns = (
        "Event Code",
        "Form Code",
        "Order",
        "Repeatable",
        "Role Scope",
        "Entry Mode",
        "Display Label Enabled",
        "Display Label Max Length",
        "Display Label Use Choice Label",
        "Display Label Empty Value Policy",
        "Display Label Template VI",
        "Display Label Fallback Template VI",
        "Display Label Empty Text VI",
        "Display Label Template EN",
        "Display Label Fallback Template EN",
        "Display Label Empty Text EN",
    )
    expected_header_map = {
        "event code": "event_code",
        "form code": "form_code",
        "order": "display_order",
        "repeatable": "repeatable",
        "role scope": "role_scope",
        "entry mode": "entry_mode",
        "display label enabled": "display_label_enabled",
        "display label max length": "display_label_max_length",
        "display label use choice label": "display_label_use_choice_label",
        "display label empty value policy": "display_label_empty_value_policy",
        "display label template vi": "display_label_template_vi",
        "display label fallback template vi": "display_label_fallback_template_vi",
        "display label empty text vi": "display_label_empty_text_vi",
        "display label template en": "display_label_template_en",
        "display label fallback template en": "display_label_fallback_template_en",
        "display label empty text en": "display_label_empty_text_en",
    }
    true_values = {"true", "1", "yes", "y"}
    false_values = {"", "false", "0", "no", "n"}
    empty_value_policy_aliases = {
        "fallback": "FALLBACK",
        "empty text": "EMPTY_TEXT",
        "empty_text": "EMPTY_TEXT",
        "omit token": "OMIT_TOKEN",
        "omit_token": "OMIT_TOKEN",
    }
    entry_mode_aliases = {
        "single": EventFormEntryModeChoices.SINGLE,
        "double entry": EventFormEntryModeChoices.DOUBLE_ENTRY,
        "double_entry": EventFormEntryModeChoices.DOUBLE_ENTRY,
        # "review only": EventFormEntryModeChoices.REVIEW_ONLY,
        # "review_only": EventFormEntryModeChoices.REVIEW_ONLY,
        # "verification": EventFormEntryModeChoices.VERIFICATION,
        # "query response": EventFormEntryModeChoices.QUERY_RESPONSE,
        # "query_response": EventFormEntryModeChoices.QUERY_RESPONSE,
    }

    def __init__(self, crf_context_adapter=None, repository=None, display_label_service=None):
        self.crf_context_adapter = crf_context_adapter or self.crf_context_adapter_class()
        self.repository = repository or self.repository_class()
        self.display_label_service = display_label_service or self.display_label_service_class()

    @staticmethod
    def _normalize_selected_study_id(selected_study_id):
        if selected_study_id is None:
            raise StudyScopeViolationError("No study is selected in the current context.")
        return int(selected_study_id)

    @classmethod
    def _ensure_current_study_scope(cls, *, selected_study_id, study_id):
        selected_study_id = cls._normalize_selected_study_id(selected_study_id)
        if int(study_id) != selected_study_id:
            raise StudyScopeViolationError("Command study scope does not match the selected study.")
        return selected_study_id

    def execute(self, command: ImportStudyEventFormBindingsTemplateCommand) -> ImportStudyEventFormBindingsTemplateResult:
        self._ensure_current_study_scope(
            selected_study_id=command.selected_study_id,
            study_id=command.study_id,
        )
        workbook_rows = self._load_rows_from_workbook(
            file_name=command.file_name,
            file_content=command.file_content,
        )

        created_count = 0
        updated_count = 0
        issues = []
        prepared_rows = []

        for row_number, row_data in workbook_rows:
            try:
                prepared_rows.append(
                    self._prepare_row(
                        study_id=command.study_id,
                        row_data=row_data,
                        row_number=row_number,
                    )
                )
            except EventFormBindingImportFormatError as exc:
                issues.append(
                    EventFormBindingImportIssue(
                        row_number=row_number,
                        event_code=str(row_data.get("event_code", "") or ""),
                        form_code=str(row_data.get("form_code", "") or ""),
                        reason=str(exc),
                    )
                )

        self._reset_bindings_for_import(
            prepared_rows=prepared_rows,
            actor_user_id=command.actor_user_id,
        )

        for prepared_row in prepared_rows:
            try:
                import_outcome = self._import_prepared_row(
                    study_id=command.study_id,
                    prepared_row=prepared_row,
                    actor_user_id=command.actor_user_id,
                )
            except EventFormBindingImportFormatError as exc:
                issues.append(
                    EventFormBindingImportIssue(
                        row_number=prepared_row.row_number,
                        event_code=prepared_row.event_code,
                        form_code=prepared_row.form_code,
                        reason=str(exc),
                    )
                )
                continue

            if import_outcome is None:
                continue

            if import_outcome == "created":
                created_count += 1
            else:
                updated_count += 1

        return ImportStudyEventFormBindingsTemplateResult(
            total_rows=len(workbook_rows),
            created_count=created_count,
            updated_count=updated_count,
            skipped_count=len(issues),
            issues=tuple(issues),
            warnings=(),
        )

    def _prepare_row(self, *, study_id, row_data, row_number):
        event_code = self._as_text(row_data.get("event_code"))
        if not event_code:
            raise EventFormBindingImportFormatError("Event Code is required.")

        form_code = self._as_text(row_data.get("form_code"))
        if not form_code:
            raise EventFormBindingImportFormatError("Form Code is required.")

        event_definition = self._resolve_event_definition(study_id=study_id, event_code=event_code)
        form_definition = self._resolve_form_definition(study_id=study_id, form_code=form_code)

        display_order = self._coerce_int(
            row_data.get("display_order"),
            field_label="Order",
            default=1,
        )
        is_repeatable_within_event = self._coerce_bool(row_data.get("repeatable"))
        role_scope = self._nullable_text(row_data.get("role_scope"))
        entry_mode = self._normalize_choice(
            raw_value=row_data.get("entry_mode"),
            aliases=self.entry_mode_aliases,
            field_label="Entry Mode",
            allow_blank=True,
        )
        return _PreparedEventFormBindingImportRow(
            row_number=row_number,
            event_code=event_code,
            form_code=form_code,
            event_definition=event_definition,
            form_definition=form_definition,
            display_order=display_order,
            is_repeatable_within_event=is_repeatable_within_event,
            role_scope=role_scope,
            entry_mode=entry_mode,
            display_label_config=self._prepare_display_label_config(row_data=row_data),
        )

    def _prepare_display_label_config(self, *, row_data):
        config_keys = (
            "display_label_enabled",
            "display_label_max_length",
            "display_label_use_choice_label",
            "display_label_empty_value_policy",
            "display_label_template_vi",
            "display_label_fallback_template_vi",
            "display_label_empty_text_vi",
            "display_label_template_en",
            "display_label_fallback_template_en",
            "display_label_empty_text_en",
        )
        if all(self._as_text(row_data.get(key)) == "" for key in config_keys):
            return None

        return {
            "is_enabled": self._coerce_bool_with_default(row_data.get("display_label_enabled"), default=True),
            "max_length": self._coerce_int(
                row_data.get("display_label_max_length"),
                field_label="Display Label Max Length",
                default=120,
            ),
            "use_choice_display_label": self._coerce_bool_with_default(
                row_data.get("display_label_use_choice_label"),
                default=True,
            ),
            "empty_value_policy": self._normalize_choice(
                raw_value=row_data.get("display_label_empty_value_policy"),
                aliases=self.empty_value_policy_aliases,
                field_label="Display Label Empty Value Policy",
                allow_blank=True,
            ) or "FALLBACK",
            "translations": {
                "vi": {
                    "label_template": self._required_text(
                        row_data.get("display_label_template_vi"),
                        field_label="Display Label Template VI",
                    ),
                    "fallback_template": self._required_text(
                        row_data.get("display_label_fallback_template_vi"),
                        field_label="Display Label Fallback Template VI",
                    ),
                    "empty_value_text": self._as_text(row_data.get("display_label_empty_text_vi")),
                },
                "en": {
                    "label_template": self._required_text(
                        row_data.get("display_label_template_en"),
                        field_label="Display Label Template EN",
                    ),
                    "fallback_template": self._required_text(
                        row_data.get("display_label_fallback_template_en"),
                        field_label="Display Label Fallback Template EN",
                    ),
                    "empty_value_text": self._as_text(row_data.get("display_label_empty_text_en")),
                },
            },
        }

    def _reset_bindings_for_import(self, *, prepared_rows, actor_user_id):
        target_event_definition_ids = sorted({int(row.event_definition.pk) for row in prepared_rows})
        if not target_event_definition_ids:
            return
        self.repository.soft_delete_event_form_bindings_for_import(
            event_definition_ids=target_event_definition_ids,
            actor_user_id=actor_user_id,
            updated_at=self._now(),
        )

    def _import_prepared_row(self, *, study_id, prepared_row, actor_user_id):
        now = self._now()
        defaults = {
            "study_id": study_id,
            "study_version": prepared_row.event_definition.study_version,
            "display_order": prepared_row.display_order,
            "is_repeatable_within_event": prepared_row.is_repeatable_within_event,
            "role_scope": prepared_row.role_scope,
            "entry_mode": prepared_row.entry_mode,
            "is_required": True,
            "is_enabled": True,
            "deleted": False,
            "updated_at": now,
            "updated_by_id": actor_user_id,
        }

        with transaction.atomic():
            binding = self.repository.get_event_form_binding(
                event_definition_id=prepared_row.event_definition.pk,
                form_definition_id=prepared_row.form_definition.pk,
            )

            if binding is None:
                binding = self.repository.create_event_form_binding(
                    event_definition_id=prepared_row.event_definition.pk,
                    form_definition_id=prepared_row.form_definition.pk,
                    created_at=now,
                    created_by_id=actor_user_id,
                    **defaults,
                )
                self._sync_display_label_config(
                    binding=binding,
                    prepared_row=prepared_row,
                    actor_user_id=actor_user_id,
                )
                return "created"

            for field_name, value in defaults.items():
                setattr(binding, field_name, value)
            self.repository.save_event_form_binding(binding, update_fields=list(defaults.keys()))
            self._sync_display_label_config(
                binding=binding,
                prepared_row=prepared_row,
                actor_user_id=actor_user_id,
            )
            return "updated"

    def _sync_display_label_config(self, *, binding, prepared_row, actor_user_id):
        if prepared_row.display_label_config is None:
            return
        try:
            self.display_label_service.save_config(
                binding_id=int(binding.pk),
                actor_user_id=actor_user_id,
                is_enabled=prepared_row.display_label_config["is_enabled"],
                max_length=prepared_row.display_label_config["max_length"],
                use_choice_display_label=prepared_row.display_label_config["use_choice_display_label"],
                empty_value_policy=prepared_row.display_label_config["empty_value_policy"],
                translations=prepared_row.display_label_config["translations"],
            )
        except EventFormDisplayLabelValidationError as exc:
            error_messages = "; ".join(error.message for error in exc.errors) or str(exc)
            raise EventFormBindingImportFormatError(
                f"Display label config is invalid: {error_messages}"
            ) from exc

    def _resolve_event_definition(self, *, study_id, event_code):
        event_definitions = list(self.repository.list_active_event_definitions_by_code(
            study_id=study_id,
            code=event_code,
        ))
        count = len(event_definitions)
        if count == 0:
            raise EventFormBindingImportFormatError("Event Code was not found.")
        if count > 1:
            raise EventFormBindingImportFormatError(
                "Event Code matched multiple event definitions. Please make the code unique across study versions."
            )
        return event_definitions[0]

    def _resolve_form_definition(self, *, study_id, form_code):
        try:
            return self.crf_context_adapter.resolve_unique_template_by_code(
                study_id=study_id,
                code=form_code,
                case_insensitive=True,
            )
        except CrfTemplateNotFoundError as exc:
            raise EventFormBindingImportFormatError("Form Code was not found.") from exc
        except CrfTemplateAmbiguousError as exc:
            raise EventFormBindingImportFormatError(
                "Form Code matched multiple CRF templates. Please make the code unique across template versions."
            ) from exc

    def _load_rows_from_workbook(self, *, file_name, file_content):
        file_name_lower = (file_name or "").strip().lower()
        if file_name_lower.endswith(".xlsx"):
            return self._load_xlsx_rows(file_content)
        if file_name_lower.endswith(".xls"):
            return self._load_xls_rows(file_content)
        raise EventFormBindingImportFormatError("Only .xlsx and .xls files are supported.")

    def _load_xlsx_rows(self, file_content):
        try:
            from openpyxl import load_workbook
        except ModuleNotFoundError as exc:
            raise EventFormBindingImportDependencyError(
                "Excel import support for .xlsx is not installed. Install openpyxl first."
            ) from exc

        workbook = load_workbook(filename=BytesIO(file_content), data_only=True, read_only=True)
        worksheet = workbook.worksheets[0]
        rows = list(worksheet.iter_rows(values_only=True))
        return self._map_rows(rows)

    def _load_xls_rows(self, file_content):
        try:
            import xlrd
        except ModuleNotFoundError as exc:
            raise EventFormBindingImportDependencyError(
                "Excel import support for .xls is not installed. Install xlrd first."
            ) from exc

        workbook = xlrd.open_workbook(file_contents=file_content)
        worksheet = workbook.sheet_by_index(0)
        rows = [worksheet.row_values(index) for index in range(worksheet.nrows)]
        return self._map_rows(rows)

    def _map_rows(self, rows):
        if not rows:
            raise EventFormBindingImportFormatError("The workbook is empty.")

        headers = [self._normalize_header(value) for value in rows[0]]
        missing_headers = [
            header
            for header in self.expected_columns
            if self._normalize_header(header) not in headers
        ]
        if missing_headers:
            raise EventFormBindingImportFormatError(
                "Missing required columns: " + ", ".join(missing_headers)
            )

        header_keys = [self.expected_header_map.get(header) for header in headers]
        mapped_rows = []
        for row_index, row_values in enumerate(rows[1:], start=2):
            if self._is_blank_row(row_values):
                continue

            row_data = {}
            for column_index, header_key in enumerate(header_keys):
                if not header_key:
                    continue
                row_data[header_key] = row_values[column_index] if column_index < len(row_values) else None

            mapped_rows.append((row_index, row_data))

        return mapped_rows

    @staticmethod
    def _normalize_header(value):
        return " ".join(str(value or "").strip().lower().split())

    @staticmethod
    def _is_blank_row(row_values):
        return all(str(value or "").strip() == "" for value in row_values)

    @staticmethod
    def _as_text(value):
        if value is None:
            return ""
        if isinstance(value, float) and value.is_integer():
            return str(int(value))
        return str(value).strip()

    def _nullable_text(self, value):
        normalized = self._as_text(value)
        return normalized or None

    def _coerce_bool(self, value):
        normalized = self._as_text(value).lower()
        if normalized in self.true_values:
            return True
        if normalized in self.false_values:
            return False
        raise EventFormBindingImportFormatError(f"Invalid boolean value: {value!r}")

    def _coerce_bool_with_default(self, value, *, default):
        normalized = self._as_text(value)
        if not normalized:
            return default
        return self._coerce_bool(normalized)

    def _coerce_int(self, value, *, field_label, default=None):
        normalized = self._as_text(value)
        if not normalized:
            if default is not None:
                return default
            raise EventFormBindingImportFormatError(f"{field_label} is required.")

        try:
            return int(float(normalized))
        except ValueError as exc:
            raise EventFormBindingImportFormatError(f"{field_label} must be a number.") from exc

    def _required_text(self, value, *, field_label):
        normalized = self._as_text(value)
        if not normalized:
            raise EventFormBindingImportFormatError(f"{field_label} is required.")
        return normalized

    def _normalize_choice(self, *, raw_value, aliases, field_label, allow_blank=False):
        normalized = self._as_text(raw_value).lower().replace("-", " ").replace("_", " ")
        normalized = " ".join(normalized.split())
        if not normalized:
            if allow_blank:
                return None
            raise EventFormBindingImportFormatError(f"{field_label} is required.")

        resolved_value = aliases.get(normalized)
        if resolved_value is None:
            raise EventFormBindingImportFormatError(f"Invalid {field_label}: {raw_value!r}")
        return resolved_value

    @staticmethod
    def _now():
        return timezone.now()
