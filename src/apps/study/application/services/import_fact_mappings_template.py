from io import BytesIO

from django.db import transaction
from django.utils import timezone

from apps.crf.domain.exceptions import StudyScopeViolationError
from apps.crf.public import (
    CrfContextAdapter,
    CrfTemplateAmbiguousError,
    CrfTemplateNotFoundError,
)
from apps.study.application.commands.import_fact_mappings_template import (
    FactMappingImportDependencyError,
    FactMappingImportFormatError,
    FactMappingImportIssue,
    ImportStudyFactMappingsTemplateCommand,
    ImportStudyFactMappingsTemplateResult,
)
from apps.study.infrastructure.repositories import DjangoStudyEventRepository


class ImportStudyFactMappingsTemplateService:
    crf_context_adapter_class = CrfContextAdapter
    repository_class = DjangoStudyEventRepository
    expected_columns = (
        "Event Code",
        "Form Code",
        "Field Code",
        "Source Path",
        "Fact Key",
        "Operator",
        "Expected Value",
        "Value Type",
        "Default Value",
        "Display Order",
        "Attestation Policy Code",
        "Attestation Action Kind",
        "Attestation Gate Code",
        "Attestation Display Order",
        "Attestation Statement Code",
        "Attestation Statement Version",
        "Attestation Permission Code",
        "Attestation Role Code",
        "Attestation Delegation Task Code",
        "Attestation Condition Code",
        "Attestation Requires Confirmation",
        "Attestation Requires Signature",
        "Attestation Requires Reauth",
        "Attestation Invalidate On Data Change",
        "Attestation Invalidate On Scope Change",
        "Attestation Required For Lock",
        "Attestation Enabled",
        "Attestation Dialog Title Vi",
        "Attestation Action Label Vi",
        "Attestation Statement Text Vi",
        "Attestation Confirmation Label Vi",
        "Attestation Success Message Vi",
        "Attestation Dialog Title En",
        "Attestation Action Label En",
        "Attestation Statement Text En",
        "Attestation Confirmation Label En",
        "Attestation Success Message En",
    )
    expected_header_map = {
        "event code": "event_code",
        "form code": "form_code",
        "field code": "field_code",
        "source path": "source_path",
        "fact key": "fact_key",
        "operator": "operator",
        "expected value": "expected_value",
        "value type": "value_type",
        "default value": "default_value",
        "display order": "display_order",
        "attestation policy code": "attestation_policy_code",
        "attestation action kind": "attestation_action_kind",
        "attestation gate code": "attestation_gate_code",
        "attestation display order": "attestation_display_order",
        "attestation statement code": "attestation_statement_code",
        "attestation statement version": "attestation_statement_version",
        "attestation permission code": "attestation_permission_code",
        "attestation role code": "attestation_role_code",
        "attestation delegation task code": "attestation_delegation_task_code",
        "attestation condition code": "attestation_condition_code",
        "attestation requires confirmation": "attestation_requires_confirmation",
        "attestation requires signature": "attestation_requires_signature",
        "attestation requires reauth": "attestation_requires_reauth",
        "attestation invalidate on data change": "attestation_invalidate_on_data_change",
        "attestation invalidate on scope change": "attestation_invalidate_on_scope_change",
        "attestation required for lock": "attestation_required_for_lock",
        "attestation enabled": "attestation_enabled",
        "attestation dialog title vi": "attestation_dialog_title_vi",
        "attestation action label vi": "attestation_action_label_vi",
        "attestation statement text vi": "attestation_statement_text_vi",
        "attestation confirmation label vi": "attestation_confirmation_label_vi",
        "attestation success message vi": "attestation_success_message_vi",
        "attestation dialog title en": "attestation_dialog_title_en",
        "attestation action label en": "attestation_action_label_en",
        "attestation statement text en": "attestation_statement_text_en",
        "attestation confirmation label en": "attestation_confirmation_label_en",
        "attestation success message en": "attestation_success_message_en",
    }
    operators = {
        "equals",
        "not_equals",
        "in",
        "not_in",
        "gt",
        "gte",
        "lt",
        "lte",
        "exists",
        "not_exists",
        "is_true",
        "is_false",
        "is_empty",
        "is_not_empty",
    }
    value_types = {"string", "boolean", "number", "decimal", "integer", "json"}
    attestation_action_kinds = {"REVIEW_COMPLETION", "CERTIFICATION"}

    def __init__(self, crf_context_adapter=None, datacapture_fact_mapping_config_adapter=None, repository=None):
        self.crf_context_adapter = crf_context_adapter or self.crf_context_adapter_class()
        self.datacapture_fact_mapping_config_adapter = (
            datacapture_fact_mapping_config_adapter or self._build_datacapture_fact_mapping_config_adapter()
        )
        self.repository = repository or self.repository_class()

    @staticmethod
    def _build_datacapture_fact_mapping_config_adapter():
        from apps.datacapture.public import DataCaptureFactMappingConfigAdapter

        return DataCaptureFactMappingConfigAdapter()

    @staticmethod
    def _normalize_selected_study_id(selected_study_id):
        if selected_study_id is None:
            raise StudyScopeViolationError("No study is selected in the current context.")
        return int(selected_study_id)

    @classmethod
    def _ensure_current_study_scope(cls, *, selected_study_id, study_id):
        selected_study_id = cls._normalize_selected_study_id(selected_study_id)
        if int(study_id) != selected_study_id:
            raise StudyScopeViolationError("Command study scope does not match the selected study.")
        return selected_study_id

    def execute(self, command: ImportStudyFactMappingsTemplateCommand) -> ImportStudyFactMappingsTemplateResult:
        self._ensure_current_study_scope(
            selected_study_id=command.selected_study_id,
            study_id=command.study_id,
        )
        workbook_rows = self._load_rows_from_workbook(
            file_name=command.file_name,
            file_content=command.file_content,
        )

        created_count = 0
        updated_count = 0
        issues = []
        now = timezone.now()

        for row_number, row_data in workbook_rows:
            try:
                import_outcome = self._import_row(
                    study_id=command.study_id,
                    row_data=row_data,
                    actor_user_id=command.actor_user_id,
                    now=now,
                )
            except FactMappingImportFormatError as exc:
                issues.append(
                    FactMappingImportIssue(
                        row_number=row_number,
                        event_code=str(row_data.get("event_code", "") or ""),
                        form_code=str(row_data.get("form_code", "") or ""),
                        fact_key=str(row_data.get("fact_key", "") or ""),
                        reason=str(exc),
                    )
                )
                continue

            if import_outcome == "created":
                created_count += 1
            else:
                updated_count += 1

        return ImportStudyFactMappingsTemplateResult(
            total_rows=len(workbook_rows),
            created_count=created_count,
            updated_count=updated_count,
            skipped_count=len(issues),
            issues=tuple(issues),
            warnings=(),
        )

    def _import_row(self, *, study_id, row_data, actor_user_id, now):
        event_code = self._require_text(row_data.get("event_code"), field_label="Event Code")
        form_code = self._require_text(row_data.get("form_code"), field_label="Form Code")
        fact_key = self._require_text(row_data.get("fact_key"), field_label="Fact Key", max_length=128)
        field_code = self._nullable_text(row_data.get("field_code"), max_length=128)
        source_path = self._nullable_text(row_data.get("source_path"), max_length=512) or field_code
        if not source_path:
            raise FactMappingImportFormatError("Source Path is required when Field Code is blank.")

        event_definition = self._resolve_event_definition(study_id=study_id, event_code=event_code)
        form_definition = self._resolve_form_definition(study_id=study_id, form_code=form_code)
        operator = self._normalize_allowed_value(
            value=row_data.get("operator"),
            allowed_values=self.operators,
            field_label="Operator",
            default="equals",
        )
        value_type = self._normalize_allowed_value(
            value=row_data.get("value_type"),
            allowed_values=self.value_types,
            field_label="Value Type",
            default="string",
        )
        display_order = self._coerce_int(
            row_data.get("display_order"),
            field_label="Display Order",
            default=1,
        )
        if display_order < 1:
            raise FactMappingImportFormatError("Display Order must be greater than 0.")

        with transaction.atomic():
            result = self.datacapture_fact_mapping_config_adapter.upsert_fact_mapping(
                study_id=study_id,
                study_version=event_definition.study_version,
                event_definition_id=event_definition.pk,
                crf_template_id=form_definition.pk,
                field_code=field_code,
                source_path=source_path,
                fact_key=fact_key,
                operator=operator,
                expected_value=self._nullable_text(row_data.get("expected_value")),
                value_type=value_type,
                default_value=self._nullable_text(row_data.get("default_value")),
                display_order=display_order,
                actor_user_id=actor_user_id,
                now=now,
            )
            self._upsert_attestation_policy_from_row(
                study_id=study_id,
                study_version=event_definition.study_version,
                event_definition=event_definition,
                row_data=row_data,
                actor_user_id=actor_user_id,
                now=now,
            )
        return result.outcome

    def _upsert_attestation_policy_from_row(
        self,
        *,
        study_id,
        study_version,
        event_definition,
        row_data,
        actor_user_id,
        now,
    ):
        if not self._has_attestation_policy_payload(row_data):
            return None

        policy_code = self._require_text(
            row_data.get("attestation_policy_code"),
            field_label="Attestation Policy Code",
            max_length=64,
        )
        action_kind = self._normalize_attestation_action_kind(
            row_data.get("attestation_action_kind"),
        )
        gate_code = self._require_text(
            row_data.get("attestation_gate_code"),
            field_label="Attestation Gate Code",
            max_length=64,
        )
        display_order = self._coerce_int(
            row_data.get("attestation_display_order"),
            field_label="Attestation Display Order",
            default=1,
        )
        if display_order < 1:
            raise FactMappingImportFormatError("Attestation Display Order must be greater than 0.")
        statement_code = (
            self._nullable_text(row_data.get("attestation_statement_code"), max_length=64)
            or policy_code
        )
        condition_code = self._nullable_text(row_data.get("attestation_condition_code"), max_length=64)
        condition_definition = None
        if condition_code:
            condition_definition = self.repository.get_condition_definition(
                study_id=study_id,
                study_version=study_version,
                code=condition_code,
            )
            if condition_definition is None:
                raise FactMappingImportFormatError("Attestation Condition Code was not found.")

        defaults = {
            "updated_at": now,
            "deleted": False,
            "action_kind": action_kind,
            "display_order": display_order,
            "statement_code": statement_code,
            "statement_version": (
                self._nullable_text(row_data.get("attestation_statement_version"), max_length=20)
                or "1"
            ),
            "required_permission_code": self._require_text(
                row_data.get("attestation_permission_code"),
                field_label="Attestation Permission Code",
                max_length=100,
            ),
            "required_role_code": self._nullable_text(row_data.get("attestation_role_code"), max_length=100),
            "delegation_task_code": self._nullable_text(
                row_data.get("attestation_delegation_task_code"),
                max_length=64,
            ),
            "condition_definition": condition_definition,
            "gate_code": gate_code,
            "requires_confirmation_checkbox": self._coerce_bool(
                row_data.get("attestation_requires_confirmation"),
                field_label="Attestation Requires Confirmation",
                default=True,
            ),
            "requires_signature": self._coerce_bool(
                row_data.get("attestation_requires_signature"),
                field_label="Attestation Requires Signature",
                default=False,
            ),
            "requires_reauthentication": self._coerce_bool(
                row_data.get("attestation_requires_reauth"),
                field_label="Attestation Requires Reauth",
                default=False,
            ),
            "invalidate_on_data_change": self._coerce_bool(
                row_data.get("attestation_invalidate_on_data_change"),
                field_label="Attestation Invalidate On Data Change",
                default=True,
            ),
            "invalidate_on_scope_change": self._coerce_bool(
                row_data.get("attestation_invalidate_on_scope_change"),
                field_label="Attestation Invalidate On Scope Change",
                default=True,
            ),
            "is_required_for_lock": self._coerce_bool(
                row_data.get("attestation_required_for_lock"),
                field_label="Attestation Required For Lock",
                default=False,
            ),
            "is_enabled": self._coerce_bool(
                row_data.get("attestation_enabled"),
                field_label="Attestation Enabled",
                default=True,
            ),
            "updated_by_id": actor_user_id,
        }

        policy = self.repository.get_attestation_policy_for_import(
            study_id=study_id,
            study_version=study_version,
            event_definition=event_definition,
            code=policy_code,
        )
        if policy is None:
            policy = self.repository.create_attestation_policy(
                study_id=study_id,
                study_version=study_version,
                event_definition=event_definition,
                code=policy_code,
                created_at=now,
                created_by_id=actor_user_id,
                **defaults,
            )
        else:
            for field_name, value in defaults.items():
                setattr(policy, field_name, value)
            self.repository.save_attestation_policy(policy, update_fields=list(defaults.keys()))

        for language_code in ("vi", "en"):
            self.repository.upsert_attestation_policy_translation(
                attestation_policy=policy,
                language_code=language_code,
                defaults=self._attestation_translation_defaults(
                    row_data=row_data,
                    language_code=language_code,
                    policy_code=policy_code,
                    action_kind=action_kind,
                ),
            )
        return policy

    def _has_attestation_policy_payload(self, row_data):
        return bool(self._as_text(row_data.get("attestation_policy_code")))

    def _normalize_attestation_action_kind(self, value):
        normalized_value = self._as_text(value).upper().replace("-", "_").replace(" ", "_")
        if normalized_value not in self.attestation_action_kinds:
            raise FactMappingImportFormatError(f"Invalid Attestation Action Kind: {value!r}")
        return normalized_value

    def _attestation_translation_defaults(self, *, row_data, language_code, policy_code, action_kind):
        suffix = language_code.lower()
        statement_text = self._require_text(
            row_data.get(f"attestation_statement_text_{suffix}"),
            field_label=f"Attestation Statement Text {suffix.upper()}",
        )
        return {
            "dialog_title": (
                self._nullable_text(row_data.get(f"attestation_dialog_title_{suffix}"), max_length=255)
                or policy_code
            ),
            "action_label": (
                self._nullable_text(row_data.get(f"attestation_action_label_{suffix}"), max_length=100)
                or action_kind.replace("_", " ").title()
            ),
            "statement_text": statement_text,
            "confirmation_label": self._nullable_text(
                row_data.get(f"attestation_confirmation_label_{suffix}"),
                max_length=255,
            ),
            "success_message": self._nullable_text(
                row_data.get(f"attestation_success_message_{suffix}"),
                max_length=500,
            ),
        }

    def _resolve_event_definition(self, *, study_id, event_code):
        event_definitions = list(
            self.repository.list_active_event_definitions_by_code(
                study_id=study_id,
                code=event_code,
            )
        )
        count = len(event_definitions)
        if count == 0:
            raise FactMappingImportFormatError("Event Code was not found.")
        if count > 1:
            raise FactMappingImportFormatError(
                "Event Code matched multiple event definitions. Please make the code unique across study versions."
            )
        return event_definitions[0]

    def _resolve_form_definition(self, *, study_id, form_code):
        try:
            return self.crf_context_adapter.resolve_unique_template_by_code(
                study_id=study_id,
                code=form_code,
                case_insensitive=True,
            )
        except CrfTemplateNotFoundError as exc:
            raise FactMappingImportFormatError("Form Code was not found.") from exc
        except CrfTemplateAmbiguousError as exc:
            raise FactMappingImportFormatError(
                "Form Code matched multiple CRF templates. Please make the code unique across template versions."
            ) from exc

    def _load_rows_from_workbook(self, *, file_name, file_content):
        file_name_lower = (file_name or "").strip().lower()
        if file_name_lower.endswith(".xlsx"):
            return self._load_xlsx_rows(file_content)
        if file_name_lower.endswith(".xls"):
            return self._load_xls_rows(file_content)
        raise FactMappingImportFormatError("Only .xlsx and .xls files are supported.")

    def _load_xlsx_rows(self, file_content):
        try:
            from openpyxl import load_workbook
        except ModuleNotFoundError as exc:
            raise FactMappingImportDependencyError(
                "Excel import support for .xlsx is not installed. Install openpyxl first."
            ) from exc

        workbook = load_workbook(filename=BytesIO(file_content), data_only=True, read_only=True)
        worksheet = workbook.worksheets[0]
        rows = list(worksheet.iter_rows(values_only=True))
        return self._map_rows(rows)

    def _load_xls_rows(self, file_content):
        try:
            import xlrd
        except ModuleNotFoundError as exc:
            raise FactMappingImportDependencyError(
                "Excel import support for .xls is not installed. Install xlrd first."
            ) from exc

        workbook = xlrd.open_workbook(file_contents=file_content)
        worksheet = workbook.sheet_by_index(0)
        rows = [worksheet.row_values(index) for index in range(worksheet.nrows)]
        return self._map_rows(rows)

    def _map_rows(self, rows):
        if not rows:
            raise FactMappingImportFormatError("The workbook is empty.")

        headers = [self._normalize_header(value) for value in rows[0]]
        missing_headers = [
            header
            for header in self.expected_columns
            if self._normalize_header(header) not in headers
        ]
        if missing_headers:
            raise FactMappingImportFormatError(
                "Missing required columns: " + ", ".join(missing_headers)
            )

        header_keys = [self.expected_header_map.get(header) for header in headers]
        mapped_rows = []
        for row_index, row_values in enumerate(rows[1:], start=2):
            if self._is_blank_row(row_values):
                continue

            row_data = {}
            for column_index, header_key in enumerate(header_keys):
                if not header_key:
                    continue
                row_data[header_key] = row_values[column_index] if column_index < len(row_values) else None

            mapped_rows.append((row_index, row_data))

        return mapped_rows

    @staticmethod
    def _normalize_header(value):
        return " ".join(str(value or "").strip().lower().split())

    @staticmethod
    def _is_blank_row(row_values):
        return all(str(value or "").strip() == "" for value in row_values)

    @staticmethod
    def _as_text(value):
        if value is None:
            return ""
        if isinstance(value, float) and value.is_integer():
            return str(int(value))
        return str(value).strip()

    def _require_text(self, value, *, field_label, max_length=None):
        normalized_value = self._as_text(value)
        if not normalized_value:
            raise FactMappingImportFormatError(f"{field_label} is required.")
        if max_length is not None and len(normalized_value) > max_length:
            raise FactMappingImportFormatError(f"{field_label} must be at most {max_length} characters.")
        return normalized_value

    def _nullable_text(self, value, *, max_length=None):
        normalized_value = self._as_text(value)
        if not normalized_value:
            return None
        if max_length is not None and len(normalized_value) > max_length:
            raise FactMappingImportFormatError(f"Value must be at most {max_length} characters.")
        return normalized_value

    def _normalize_allowed_value(self, *, value, allowed_values, field_label, default):
        normalized_value = self._as_text(value).lower().replace("-", "_").replace(" ", "_")
        normalized_value = normalized_value or default
        if normalized_value not in allowed_values:
            raise FactMappingImportFormatError(f"Invalid {field_label}: {value!r}")
        return normalized_value

    def _coerce_bool(self, value, *, field_label, default):
        normalized_value = self._as_text(value).lower()
        if not normalized_value:
            return default
        truthy = {"1", "true", "yes", "y"}
        falsy = {"0", "false", "no", "n"}
        if normalized_value in truthy:
            return True
        if normalized_value in falsy:
            return False
        raise FactMappingImportFormatError(
            f"{field_label} must be one of: 1/0, true/false, yes/no."
        )

    def _coerce_int(self, value, *, field_label, default=None):
        normalized = self._as_text(value)
        if not normalized:
            if default is not None:
                return default
            raise FactMappingImportFormatError(f"{field_label} is required.")

        try:
            return int(float(normalized))
        except ValueError as exc:
            raise FactMappingImportFormatError(f"{field_label} must be a number.") from exc


__all__ = ["ImportStudyFactMappingsTemplateService"]
