from io import BytesIO

from django.db import transaction
from django.utils import timezone

from apps.crf.domain.exceptions import StudyScopeViolationError
from apps.study.application.commands.import_event_attestation_policies_template import (
    EventAttestationPolicyImportDependencyError,
    EventAttestationPolicyImportFormatError,
    EventAttestationPolicyImportIssue,
    ImportStudyEventAttestationPoliciesTemplateCommand,
    ImportStudyEventAttestationPoliciesTemplateResult,
)
from apps.study.infrastructure.repositories import DjangoStudyEventRepository


class ImportStudyEventAttestationPoliciesTemplateService:
    sheet_name = "Attestation Policies"
    expected_columns = {
        sheet_name: (
            "Event Code",
            "Study Version",
            "Attestation Policy Code",
            "Attestation Action Kind",
            "Attestation Gate Code",
            "Attestation Display Order",
            "Attestation Statement Code",
            "Attestation Statement Version",
            "Attestation Permission Code",
            "Attestation Role Code",
            "Attestation Delegation Task Code",
            "Attestation Condition Code",
            "Attestation Requires Confirmation",
            "Attestation Requires Signature",
            "Attestation Requires Reauth",
            "Attestation Invalidate On Data Change",
            "Attestation Invalidate On Scope Change",
            "Attestation Required For Lock",
            "Attestation Enabled",
            "Attestation Dialog Title Vi",
            "Attestation Action Label Vi",
            "Attestation Statement Text Vi",
            "Attestation Confirmation Label Vi",
            "Attestation Success Message Vi",
            "Attestation Dialog Title En",
            "Attestation Action Label En",
            "Attestation Statement Text En",
            "Attestation Confirmation Label En",
            "Attestation Success Message En",
        ),
    }
    expected_header_map = {
        sheet_name: {
            "event code": "event_code",
            "study version": "study_version",
            "attestation policy code": "attestation_policy_code",
            "attestation action kind": "attestation_action_kind",
            "attestation gate code": "attestation_gate_code",
            "attestation display order": "attestation_display_order",
            "attestation statement code": "attestation_statement_code",
            "attestation statement version": "attestation_statement_version",
            "attestation permission code": "attestation_permission_code",
            "attestation role code": "attestation_role_code",
            "attestation delegation task code": "attestation_delegation_task_code",
            "attestation condition code": "attestation_condition_code",
            "attestation requires confirmation": "attestation_requires_confirmation",
            "attestation requires signature": "attestation_requires_signature",
            "attestation requires reauth": "attestation_requires_reauth",
            "attestation invalidate on data change": "attestation_invalidate_on_data_change",
            "attestation invalidate on scope change": "attestation_invalidate_on_scope_change",
            "attestation invalidate on query change": "attestation_invalidate_on_query_change",
            "attestation required for lock": "attestation_required_for_lock",
            "attestation enabled": "attestation_enabled",
            "attestation dialog title vi": "attestation_dialog_title_vi",
            "attestation action label vi": "attestation_action_label_vi",
            "attestation statement text vi": "attestation_statement_text_vi",
            "attestation confirmation label vi": "attestation_confirmation_label_vi",
            "attestation success message vi": "attestation_success_message_vi",
            "attestation dialog title en": "attestation_dialog_title_en",
            "attestation action label en": "attestation_action_label_en",
            "attestation statement text en": "attestation_statement_text_en",
            "attestation confirmation label en": "attestation_confirmation_label_en",
            "attestation success message en": "attestation_success_message_en",
        },
    }
    sheet_aliases = {
        sheet_name: (
            sheet_name,
            "Event Attestation Policies",
            "Review Certification Policies",
            "Attestation",
        ),
    }
    attestation_action_kinds = {"REVIEW_COMPLETION", "CERTIFICATION"}
    repository_class = DjangoStudyEventRepository

    def __init__(self, repository=None):
        self.repository = repository or self.repository_class()

    @staticmethod
    def _normalize_selected_study_id(selected_study_id):
        if selected_study_id is None:
            raise StudyScopeViolationError("No study is selected in the current context.")
        return int(selected_study_id)

    @classmethod
    def _ensure_current_study_scope(cls, *, selected_study_id, study_id):
        selected_study_id = cls._normalize_selected_study_id(selected_study_id)
        if int(study_id) != selected_study_id:
            raise StudyScopeViolationError("Command study scope does not match the selected study.")
        return selected_study_id

    def execute(
        self,
        command: ImportStudyEventAttestationPoliciesTemplateCommand,
    ) -> ImportStudyEventAttestationPoliciesTemplateResult:
        selected_study_id = self._ensure_current_study_scope(
            selected_study_id=command.selected_study_id,
            study_id=command.study_id,
        )
        rows_by_sheet = self._load_rows_from_workbook(
            file_name=command.file_name,
            file_content=command.file_content,
        )
        rows = rows_by_sheet[self.sheet_name]

        created_count = 0
        updated_count = 0
        issues = []
        now = timezone.now()
        for row_number, row_data in rows:
            identifier = self._build_identifier(row_data)
            try:
                import_outcome = self._import_row(
                    study_id=selected_study_id,
                    row_data=row_data,
                    actor_user_id=command.actor_user_id,
                    now=now,
                )
            except EventAttestationPolicyImportFormatError as exc:
                issues.append(
                    EventAttestationPolicyImportIssue(
                        sheet_name=self.sheet_name,
                        row_number=row_number,
                        identifier=identifier,
                        reason=str(exc),
                    )
                )
                continue

            if import_outcome == "created":
                created_count += 1
            else:
                updated_count += 1

        return ImportStudyEventAttestationPoliciesTemplateResult(
            total_rows=len(rows),
            created_count=created_count,
            updated_count=updated_count,
            skipped_count=len(issues),
            issues=tuple(issues),
            warnings=(),
        )

    def _import_row(self, *, study_id, row_data, actor_user_id, now):
        event_definition = self._resolve_event_definition(
            study_id=study_id,
            event_code=self._require_text(
                row_data.get("event_code"),
                field_label="Event Code",
                max_length=64,
            ),
            study_version=self._nullable_text(row_data.get("study_version"), max_length=20),
        )
        policy_code = self._require_text(
            row_data.get("attestation_policy_code"),
            field_label="Attestation Policy Code",
            max_length=64,
        )
        action_kind = self._normalize_attestation_action_kind(row_data.get("attestation_action_kind"))
        condition_code = self._nullable_text(row_data.get("attestation_condition_code"), max_length=64)
        condition_definition = None
        if condition_code:
            condition_definition = self.repository.get_condition_definition(
                study_id=study_id,
                study_version=event_definition.study_version,
                code=condition_code,
            )
            if condition_definition is None:
                raise EventAttestationPolicyImportFormatError(
                    "Attestation Condition Code was not found."
                )

        defaults = {
            "updated_at": now,
            "deleted": False,
            "action_kind": action_kind,
            "display_order": self._coerce_int(
                row_data.get("attestation_display_order"),
                field_label="Attestation Display Order",
                default=1,
            ),
            "statement_code": (
                self._nullable_text(row_data.get("attestation_statement_code"), max_length=64)
                or policy_code
            ),
            "statement_version": (
                self._nullable_text(row_data.get("attestation_statement_version"), max_length=20)
                or "1"
            ),
            "required_permission_code": self._require_text(
                row_data.get("attestation_permission_code"),
                field_label="Attestation Permission Code",
                max_length=100,
            ),
            "required_role_code": self._nullable_text(
                row_data.get("attestation_role_code"),
                max_length=100,
            ),
            "delegation_task_code": self._nullable_text(
                row_data.get("attestation_delegation_task_code"),
                max_length=64,
            ),
            "condition_definition": condition_definition,
            "gate_code": self._require_text(
                row_data.get("attestation_gate_code"),
                field_label="Attestation Gate Code",
                max_length=64,
            ),
            "requires_confirmation_checkbox": self._coerce_bool(
                row_data.get("attestation_requires_confirmation"),
                field_label="Attestation Requires Confirmation",
                default=True,
            ),
            "requires_signature": self._coerce_bool(
                row_data.get("attestation_requires_signature"),
                field_label="Attestation Requires Signature",
                default=False,
            ),
            "requires_reauthentication": self._coerce_bool(
                row_data.get("attestation_requires_reauth"),
                field_label="Attestation Requires Reauth",
                default=False,
            ),
            "invalidate_on_data_change": self._coerce_bool(
                row_data.get("attestation_invalidate_on_data_change"),
                field_label="Attestation Invalidate On Data Change",
                default=True,
            ),
            "invalidate_on_scope_change": self._coerce_bool(
                row_data.get("attestation_invalidate_on_scope_change"),
                field_label="Attestation Invalidate On Scope Change",
                default=True,
            ),
            "invalidate_on_query_change": self._coerce_bool(
                row_data.get("attestation_invalidate_on_query_change"),
                field_label="Attestation Invalidate On Query Change",
                default=True,
            ),
            "is_required_for_lock": self._coerce_bool(
                row_data.get("attestation_required_for_lock"),
                field_label="Attestation Required For Lock",
                default=False,
            ),
            "is_enabled": self._coerce_bool(
                row_data.get("attestation_enabled"),
                field_label="Attestation Enabled",
                default=True,
            ),
            "updated_by_id": actor_user_id,
        }
        if defaults["display_order"] < 1:
            raise EventAttestationPolicyImportFormatError(
                "Attestation Display Order must be greater than 0."
            )

        with transaction.atomic():
            policy = self.repository.get_attestation_policy_for_import(
                study_id=study_id,
                study_version=event_definition.study_version,
                event_definition=event_definition,
                code=policy_code,
            )
            if policy is None:
                policy = self.repository.create_attestation_policy(
                    study_id=study_id,
                    study_version=event_definition.study_version,
                    event_definition=event_definition,
                    code=policy_code,
                    created_at=now,
                    created_by_id=actor_user_id,
                    **defaults,
                )
                outcome = "created"
            else:
                for field_name, value in defaults.items():
                    setattr(policy, field_name, value)
                self.repository.save_attestation_policy(
                    policy,
                    update_fields=list(defaults.keys()),
                )
                outcome = "updated"

            for language_code in ("vi", "en"):
                self.repository.upsert_attestation_policy_translation(
                    attestation_policy=policy,
                    language_code=language_code,
                    defaults=self._translation_defaults(
                        row_data=row_data,
                        language_code=language_code,
                        policy_code=policy_code,
                        action_kind=action_kind,
                    ),
                )
        return outcome

    def _resolve_event_definition(self, *, study_id, event_code, study_version):
        if study_version:
            event_definition = self.repository.get_active_event_definition_by_code(
                study_id=study_id,
                study_version=study_version,
                code=event_code,
            )
            if event_definition is None:
                raise EventAttestationPolicyImportFormatError(
                    "Event Code and Study Version were not found."
                )
            return event_definition

        event_definitions = list(
            self.repository.list_active_event_definitions_by_code(
                study_id=study_id,
                code=event_code,
            )
        )
        count = len(event_definitions)
        if count == 0:
            raise EventAttestationPolicyImportFormatError("Event Code was not found.")
        if count > 1:
            raise EventAttestationPolicyImportFormatError(
                "Event Code matched multiple study versions. Please fill Study Version."
            )
        return event_definitions[0]

    def _translation_defaults(self, *, row_data, language_code, policy_code, action_kind):
        suffix = language_code.lower()
        return {
            "dialog_title": (
                self._nullable_text(row_data.get(f"attestation_dialog_title_{suffix}"), max_length=255)
                or policy_code
            ),
            "action_label": (
                self._nullable_text(row_data.get(f"attestation_action_label_{suffix}"), max_length=100)
                or action_kind.replace("_", " ").title()
            ),
            "statement_text": self._require_text(
                row_data.get(f"attestation_statement_text_{suffix}"),
                field_label=f"Attestation Statement Text {suffix.upper()}",
                max_length=10000,
            ),
            "confirmation_label": self._nullable_text(
                row_data.get(f"attestation_confirmation_label_{suffix}"),
                max_length=255,
            ),
            "success_message": self._nullable_text(
                row_data.get(f"attestation_success_message_{suffix}"),
                max_length=500,
            ),
        }

    def _load_rows_from_workbook(self, *, file_name, file_content):
        file_name_lower = (file_name or "").strip().lower()
        if file_name_lower.endswith(".xlsx"):
            return self._load_xlsx_rows(file_content)
        if file_name_lower.endswith(".xls"):
            return self._load_xls_rows(file_content)
        raise EventAttestationPolicyImportFormatError("Only .xlsx and .xls files are supported.")

    def _load_xlsx_rows(self, file_content):
        try:
            from openpyxl import load_workbook
        except ModuleNotFoundError as exc:
            raise EventAttestationPolicyImportDependencyError(
                "Excel import support for .xlsx is not installed. Install openpyxl first."
            ) from exc

        workbook = load_workbook(filename=BytesIO(file_content), data_only=True, read_only=True)
        actual_sheet_name = self._resolve_sheet_name(workbook.sheetnames)
        return {
            self.sheet_name: self._map_rows(
                rows=list(workbook[actual_sheet_name].iter_rows(values_only=True))
            )
        }

    def _load_xls_rows(self, file_content):
        try:
            import xlrd
        except ModuleNotFoundError as exc:
            raise EventAttestationPolicyImportDependencyError(
                "Excel import support for .xls is not installed. Install xlrd first."
            ) from exc

        workbook = xlrd.open_workbook(file_contents=file_content)
        sheet_name = self._resolve_sheet_name(workbook.sheet_names())
        worksheet = workbook.sheet_by_name(sheet_name)
        return {
            self.sheet_name: self._map_rows(
                rows=[worksheet.row_values(index) for index in range(worksheet.nrows)]
            )
        }

    def _resolve_sheet_name(self, workbook_sheet_names):
        normalized_sheet_map = {
            self._normalize_header(sheet_name): sheet_name
            for sheet_name in workbook_sheet_names
        }
        for candidate_name in self.sheet_aliases[self.sheet_name]:
            resolved_sheet_name = normalized_sheet_map.get(self._normalize_header(candidate_name))
            if resolved_sheet_name:
                return resolved_sheet_name

        accepted_names = ", ".join(self.sheet_aliases[self.sheet_name])
        raise EventAttestationPolicyImportFormatError(
            f"Missing required worksheet: {self.sheet_name}. Accepted names: {accepted_names}"
        )

    def _map_rows(self, rows):
        if not rows:
            raise EventAttestationPolicyImportFormatError("The workbook is empty.")

        headers = [self._normalize_header(value) for value in rows[0]]
        missing_headers = [
            header
            for header in self.expected_columns[self.sheet_name]
            if self._normalize_header(header) not in headers
        ]
        if missing_headers:
            raise EventAttestationPolicyImportFormatError(
                "Missing required columns: " + ", ".join(missing_headers)
            )

        header_keys = [self.expected_header_map[self.sheet_name].get(header) for header in headers]
        mapped_rows = []
        for row_index, row_values in enumerate(rows[1:], start=2):
            if self._is_blank_row(row_values):
                continue
            row_data = {}
            for column_index, header_key in enumerate(header_keys):
                if not header_key:
                    continue
                row_data[header_key] = row_values[column_index] if column_index < len(row_values) else None
            mapped_rows.append((row_index, row_data))
        return mapped_rows

    def _normalize_attestation_action_kind(self, value):
        normalized_value = self._as_text(value).upper().replace("-", "_").replace(" ", "_")
        if normalized_value not in self.attestation_action_kinds:
            raise EventAttestationPolicyImportFormatError(
                f"Invalid Attestation Action Kind: {value!r}"
            )
        return normalized_value

    def _build_identifier(self, row_data):
        return " / ".join(
            part
            for part in (
                self._as_text(row_data.get("event_code")),
                self._as_text(row_data.get("study_version")),
                self._as_text(row_data.get("attestation_policy_code")),
            )
            if part
        )

    @staticmethod
    def _normalize_header(value):
        return " ".join(str(value or "").strip().lower().split())

    @staticmethod
    def _is_blank_row(row_values):
        return all(str(value or "").strip() == "" for value in row_values)

    @staticmethod
    def _as_text(value):
        if value is None:
            return ""
        if isinstance(value, float) and value.is_integer():
            return str(int(value))
        return str(value).strip()

    def _require_text(self, value, *, field_label, max_length):
        normalized_value = self._as_text(value)
        if not normalized_value:
            raise EventAttestationPolicyImportFormatError(f"{field_label} is required.")
        if len(normalized_value) > max_length:
            raise EventAttestationPolicyImportFormatError(
                f"{field_label} must be {max_length} characters or fewer."
            )
        return normalized_value

    def _nullable_text(self, value, *, max_length):
        normalized_value = self._as_text(value)
        if not normalized_value:
            return None
        if len(normalized_value) > max_length:
            raise EventAttestationPolicyImportFormatError(
                f"Value must be {max_length} characters or fewer."
            )
        return normalized_value

    def _coerce_bool(self, value, *, field_label, default):
        normalized_value = self._as_text(value).lower()
        if not normalized_value:
            return default
        truthy = {"1", "true", "yes", "y"}
        falsy = {"0", "false", "no", "n"}
        if normalized_value in truthy:
            return True
        if normalized_value in falsy:
            return False
        raise EventAttestationPolicyImportFormatError(
            f"{field_label} must be one of: 1/0, true/false, yes/no."
        )

    def _coerce_int(self, value, *, field_label, default):
        normalized = self._as_text(value)
        if not normalized:
            return default
        try:
            return int(float(normalized))
        except ValueError as exc:
            raise EventAttestationPolicyImportFormatError(
                f"{field_label} must be a number."
            ) from exc


__all__ = ["ImportStudyEventAttestationPoliciesTemplateService"]
