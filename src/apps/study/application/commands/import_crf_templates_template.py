from dataclasses import dataclass
from io import BytesIO

from django.utils import timezone

from apps.crf.public import (
    CrfContextAdapter,
)
from apps.crf.domain.exceptions import StudyScopeViolationError
from apps.shared.context_processors import StudyDropdownHandler


@dataclass(frozen=True)
class ImportStudyCrfTemplatesTemplateCommand:
    actor_user_id: int
    study_id: int
    file_name: str
    file_content: bytes


@dataclass(frozen=True)
class CrfTemplateImportIssue:
    sheet_name: str
    row_number: int
    identifier: str
    reason: str


@dataclass(frozen=True)
class ImportStudyCrfTemplatesTemplateResult:
    total_rows: int
    created_count: int
    updated_count: int
    skipped_count: int
    issues: tuple[CrfTemplateImportIssue, ...] = ()
    warnings: tuple[str, ...] = ()


class CrfTemplateImportTemplateError(Exception):
    """Base error raised for CRF template import failures."""


class CrfTemplateImportDependencyError(CrfTemplateImportTemplateError):
    """Raised when the Excel parser dependency is missing."""


class CrfTemplateImportFormatError(CrfTemplateImportTemplateError):
    """Raised when the uploaded workbook shape is invalid."""


class ImportStudyCrfTemplatesTemplateService:
    crf_context_adapter_class = CrfContextAdapter
    form_templates_sheet_name = "Form Templates"
    section_templates_sheet_name = "Section Templates"
    legacy_form_templates_sheet_name = "CRF Templates"
    expected_columns = {
        form_templates_sheet_name: (
            "Code",
            "Vi Name",
            "En Name",
            "Version",
        ),
        section_templates_sheet_name: (
            "Form Code",
            "Code",
            "Vi Name",
            "En Name",
            "Order",
            "Required",
            "Repeated",
            "Min Repeat",
            "Max Repeat",
        ),
    }
    expected_header_map = {
        form_templates_sheet_name: {
            "code": "code",
            "vi name": "vi_name",
            "en name": "en_name",
            "version": "version",
        },
        section_templates_sheet_name: {
            "form code": "form_code",
            "code": "code",
            "vi name": "vi_name",
            "en name": "en_name",
            "order": "order",
            "required": "required",
            "repeated": "repeated",
            "min repeat": "min_repeat",
            "max repeat": "max_repeat",
        },
    }
    sheet_aliases = {
        form_templates_sheet_name: (
            form_templates_sheet_name,
            "Form Template",
            legacy_form_templates_sheet_name,
        ),
        section_templates_sheet_name: (
            section_templates_sheet_name,
            "Section Template",
            "sectiontemplate",
        ),
    }

    def __init__(self, crf_context_adapter=None):
        self.crf_context_adapter = crf_context_adapter or self.crf_context_adapter_class()

    @staticmethod
    def _resolve_selected_study_id(request):
        try:
            selected_study_id = StudyDropdownHandler(request=request).build().selected_id
        except Exception as exc:
            raise StudyScopeViolationError("No study is selected in the current context.") from exc
        if selected_study_id is None:
            raise StudyScopeViolationError("No study is selected in the current context.")
        return int(selected_study_id)

    @classmethod
    def _ensure_current_study_scope(cls, *, request, study_id):
        selected_study_id = cls._resolve_selected_study_id(request)
        if int(study_id) != selected_study_id:
            raise StudyScopeViolationError("Command study scope does not match the selected study.")
        return selected_study_id

    def execute(self, command: ImportStudyCrfTemplatesTemplateCommand, *, request) -> ImportStudyCrfTemplatesTemplateResult:
        self._ensure_current_study_scope(request=request, study_id=command.study_id)
        workbook_rows_by_sheet = self._load_rows_from_workbook(
            file_name=command.file_name,
            file_content=command.file_content,
        )

        created_count = 0
        updated_count = 0
        issues = []
        imported_template_ids_by_code = {}

        for row_number, row_data in workbook_rows_by_sheet[self.form_templates_sheet_name]:
            identifier = self._build_form_template_identifier(row_data)
            try:
                import_outcome, imported_template_id, imported_template_code = self._import_form_template_row(
                    request=request,
                    study_id=command.study_id,
                    row_data=row_data,
                    row_number=row_number,
                    actor_user_id=command.actor_user_id,
                )
            except CrfTemplateImportFormatError as exc:
                issues.append(
                    CrfTemplateImportIssue(
                        sheet_name=self.form_templates_sheet_name,
                        row_number=row_number,
                        identifier=identifier,
                        reason=str(exc),
                    )
                )
                continue

            imported_template_ids_by_code.setdefault(imported_template_code, set()).add(imported_template_id)

            if import_outcome == "created":
                created_count += 1
            else:
                updated_count += 1

        for row_number, row_data in workbook_rows_by_sheet[self.section_templates_sheet_name]:
            identifier = self._build_section_template_identifier(row_data)
            try:
                row_created_count, row_updated_count = self._import_section_template_row(
                    request=request,
                    study_id=command.study_id,
                    row_data=row_data,
                    row_number=row_number,
                    actor_user_id=command.actor_user_id,
                    imported_template_ids_by_code=imported_template_ids_by_code,
                )
            except CrfTemplateImportFormatError as exc:
                issues.append(
                    CrfTemplateImportIssue(
                        sheet_name=self.section_templates_sheet_name,
                        row_number=row_number,
                        identifier=identifier,
                        reason=str(exc),
                    )
                )
                continue

            created_count += row_created_count
            updated_count += row_updated_count

        total_rows = (
            len(workbook_rows_by_sheet[self.form_templates_sheet_name])
            + len(workbook_rows_by_sheet[self.section_templates_sheet_name])
        )
        return ImportStudyCrfTemplatesTemplateResult(
            total_rows=total_rows,
            created_count=created_count,
            updated_count=updated_count,
            skipped_count=len(issues),
            issues=tuple(issues),
            warnings=(),
        )

    def _import_form_template_row(self, *, request, study_id, row_data, row_number, actor_user_id):
        code = self._require_text(
            row_data.get("code"),
            field_label="Code",
            max_length=64,
        )
        vi_name = self._require_text(
            row_data.get("vi_name"),
            field_label="Vi Name",
            max_length=255,
        )
        en_name = self._require_text(
            row_data.get("en_name"),
            field_label="En Name",
            max_length=255,
        )
        version = self._require_text(
            row_data.get("version"),
            field_label="Version",
            max_length=32,
        )

        import_outcome = self.crf_context_adapter.upsert_crf_template(
            request=request,
            study_id=study_id,
            code=code,
            version=version,
            vi_name=vi_name,
            en_name=en_name,
            actor_user_id=actor_user_id,
            now=self._now(),
        )
        imported_template = self.crf_context_adapter.resolve_unique_template_by_code_version(
            study_id=study_id,
            code=code,
            version=version,
        )
        return import_outcome, imported_template.pk, imported_template.code

    def _import_section_template_row(
        self,
        *,
        request,
        study_id,
        row_data,
        row_number,
        actor_user_id,
        imported_template_ids_by_code,
    ):
        form_code = self._require_text(
            row_data.get("form_code"),
            field_label="Form Code",
            max_length=64,
        )
        section_code = self._require_text(
            row_data.get("code"),
            field_label="Code",
            max_length=64,
        )
        vi_name = self._require_text(
            row_data.get("vi_name"),
            field_label="Vi Name",
            max_length=255,
        )
        en_name = self._require_text(
            row_data.get("en_name"),
            field_label="En Name",
            max_length=255,
        )
        display_order = self._coerce_positive_int(
            row_data.get("order"),
            field_label="Order",
        )
        is_required = self._coerce_bool(
            row_data.get("required"),
            field_label="Required",
            default=True,
        )
        is_repeatable = self._coerce_bool(
            row_data.get("repeated"),
            field_label="Repeated",
            default=False,
        )
        min_repeats = self._coerce_non_negative_int(
            row_data.get("min_repeat"),
            field_label="Min Repeat",
            default=0,
        )
        max_repeats = self._coerce_optional_non_negative_int(
            row_data.get("max_repeat"),
            field_label="Max Repeat",
        )

        if max_repeats is not None and max_repeats < min_repeats:
            raise CrfTemplateImportFormatError("Max Repeat must be greater than or equal to Min Repeat.")

        template_ids = sorted(imported_template_ids_by_code.get(form_code, ()))
        if not template_ids:
            matched_templates = self.crf_context_adapter.list_study_templates_by_code(
                study_id=study_id,
                code=form_code,
            )
            if not matched_templates:
                raise CrfTemplateImportFormatError(
                    "Form Code was not found in study CRF templates."
                )
            if len(matched_templates) > 1:
                raise CrfTemplateImportFormatError(
                    "Form Code is ambiguous because multiple versions exist. "
                    "Include matching rows in Form Templates sheet for the desired version."
                )
            template_ids = [matched_templates[0].pk]

        created_count = 0
        updated_count = 0
        for template_id in template_ids:
            import_outcome = self.crf_context_adapter.upsert_section_template(
                request=request,
                crf_template_id=template_id,
                section_code=section_code,
                vi_name=vi_name,
                en_name=en_name,
                display_order=display_order,
                is_required=is_required,
                is_repeatable=is_repeatable,
                min_repeats=min_repeats,
                max_repeats=max_repeats,
                actor_user_id=actor_user_id,
                now=self._now(),
            )
            if import_outcome == "created":
                created_count += 1
            else:
                updated_count += 1

        return created_count, updated_count

    def _load_rows_from_workbook(self, *, file_name, file_content):
        file_name_lower = (file_name or "").strip().lower()
        if file_name_lower.endswith(".xlsx"):
            return self._load_xlsx_rows(file_content)
        if file_name_lower.endswith(".xls"):
            return self._load_xls_rows(file_content)
        raise CrfTemplateImportFormatError("Only .xlsx and .xls files are supported.")

    def _load_xlsx_rows(self, file_content):
        try:
            from openpyxl import load_workbook
        except ModuleNotFoundError as exc:
            raise CrfTemplateImportDependencyError(
                "Excel import support for .xlsx is not installed. Install openpyxl first."
            ) from exc

        workbook = load_workbook(filename=BytesIO(file_content), data_only=True, read_only=True)
        rows_by_sheet = {}
        for logical_sheet_name in self.expected_columns:
            actual_sheet_name = self._resolve_sheet_name(
                workbook_sheet_names=workbook.sheetnames,
                logical_sheet_name=logical_sheet_name,
            )
            rows_by_sheet[logical_sheet_name] = self._map_rows(
                rows=list(workbook[actual_sheet_name].iter_rows(values_only=True)),
                sheet_name=logical_sheet_name,
            )
        return rows_by_sheet

    def _load_xls_rows(self, file_content):
        try:
            import xlrd
        except ModuleNotFoundError as exc:
            raise CrfTemplateImportDependencyError(
                "Excel import support for .xls is not installed. Install xlrd first."
            ) from exc

        workbook = xlrd.open_workbook(file_contents=file_content)
        rows_by_sheet = {}
        for logical_sheet_name in self.expected_columns:
            actual_sheet_name = self._resolve_sheet_name(
                workbook_sheet_names=workbook.sheet_names(),
                logical_sheet_name=logical_sheet_name,
            )
            try:
                worksheet = workbook.sheet_by_name(actual_sheet_name)
            except xlrd.biffh.XLRDError as exc:
                raise CrfTemplateImportFormatError(
                    f"Missing required worksheet: {logical_sheet_name}"
                ) from exc
            rows_by_sheet[logical_sheet_name] = self._map_rows(
                rows=[worksheet.row_values(index) for index in range(worksheet.nrows)],
                sheet_name=logical_sheet_name,
            )
        return rows_by_sheet

    def _resolve_sheet_name(self, *, workbook_sheet_names, logical_sheet_name):
        normalized_sheet_map = {
            self._normalize_header(sheet_name): sheet_name
            for sheet_name in workbook_sheet_names
        }
        for candidate_name in self.sheet_aliases.get(logical_sheet_name, (logical_sheet_name,)):
            normalized_candidate = self._normalize_header(candidate_name)
            resolved_sheet_name = normalized_sheet_map.get(normalized_candidate)
            if resolved_sheet_name:
                return resolved_sheet_name

        accepted_names = ", ".join(self.sheet_aliases.get(logical_sheet_name, (logical_sheet_name,)))
        raise CrfTemplateImportFormatError(
            f"Missing required worksheet: {logical_sheet_name}. Accepted names: {accepted_names}"
        )

    def _map_rows(self, *, rows, sheet_name):
        if sheet_name not in self.expected_columns:
            raise CrfTemplateImportFormatError(f"Unexpected worksheet: {sheet_name}")
        if not rows:
            return []

        headers = [self._normalize_header(value) for value in rows[0]]
        missing_headers = [
            header
            for header in self.expected_columns[sheet_name]
            if self._normalize_header(header) not in headers
        ]
        if missing_headers:
            raise CrfTemplateImportFormatError(
                f"Worksheet '{sheet_name}' is missing required columns: " + ", ".join(missing_headers)
            )

        header_keys = [self.expected_header_map[sheet_name].get(header) for header in headers]
        mapped_rows = []
        for row_index, row_values in enumerate(rows[1:], start=2):
            if self._is_blank_row(row_values):
                continue

            row_data = {}
            for column_index, header_key in enumerate(header_keys):
                if not header_key:
                    continue
                row_data[header_key] = row_values[column_index] if column_index < len(row_values) else None

            mapped_rows.append((row_index, row_data))

        return mapped_rows

    @staticmethod
    def _normalize_header(value):
        return " ".join(str(value or "").strip().lower().split())

    @staticmethod
    def _is_blank_row(row_values):
        return all(str(value or "").strip() == "" for value in row_values)

    @staticmethod
    def _as_text(value):
        if value is None:
            return ""
        if isinstance(value, float) and value.is_integer():
            return str(int(value))
        return str(value).strip()

    def _require_text(self, value, *, field_label, max_length):
        normalized_value = self._as_text(value)
        if not normalized_value:
            raise CrfTemplateImportFormatError(f"{field_label} is required.")
        if len(normalized_value) > max_length:
            raise CrfTemplateImportFormatError(
                f"{field_label} must be {max_length} characters or fewer."
            )
        return normalized_value

    def _coerce_positive_int(self, value, *, field_label):
        normalized_value = self._as_text(value)
        if not normalized_value:
            raise CrfTemplateImportFormatError(f"{field_label} is required.")
        try:
            parsed = int(normalized_value)
        except ValueError as exc:
            raise CrfTemplateImportFormatError(f"{field_label} must be an integer.") from exc
        if parsed <= 0:
            raise CrfTemplateImportFormatError(f"{field_label} must be greater than 0.")
        return parsed

    def _coerce_non_negative_int(self, value, *, field_label, default=None):
        normalized_value = self._as_text(value)
        if not normalized_value:
            if default is None:
                raise CrfTemplateImportFormatError(f"{field_label} is required.")
            return default
        try:
            parsed = int(normalized_value)
        except ValueError as exc:
            raise CrfTemplateImportFormatError(f"{field_label} must be an integer.") from exc
        if parsed < 0:
            raise CrfTemplateImportFormatError(f"{field_label} must be greater than or equal to 0.")
        return parsed

    def _coerce_optional_non_negative_int(self, value, *, field_label):
        normalized_value = self._as_text(value)
        if not normalized_value:
            return None
        try:
            parsed = int(normalized_value)
        except ValueError as exc:
            raise CrfTemplateImportFormatError(f"{field_label} must be an integer.") from exc
        if parsed < 0:
            raise CrfTemplateImportFormatError(f"{field_label} must be greater than or equal to 0.")
        return parsed

    def _coerce_bool(self, value, *, field_label, default=None):
        normalized_value = self._as_text(value).lower()
        if normalized_value == "":
            if default is None:
                raise CrfTemplateImportFormatError(f"{field_label} is required.")
            return default

        truthy = {"1", "true", "yes", "y"}
        falsy = {"0", "false", "no", "n"}
        if normalized_value in truthy:
            return True
        if normalized_value in falsy:
            return False
        raise CrfTemplateImportFormatError(
            f"{field_label} must be one of: 1/0, true/false, yes/no."
        )

    def _build_form_template_identifier(self, row_data):
        return " / ".join(
            part for part in (
                self._as_text(row_data.get("code")),
                self._as_text(row_data.get("version")),
            ) if part
        )

    def _build_section_template_identifier(self, row_data):
        return " / ".join(
            part for part in (
                self._as_text(row_data.get("form_code")),
                self._as_text(row_data.get("code")),
                self._as_text(row_data.get("order")),
                self._as_text(row_data.get("repeated")),
            ) if part
        )

    @staticmethod
    def _now():
        return timezone.now()
