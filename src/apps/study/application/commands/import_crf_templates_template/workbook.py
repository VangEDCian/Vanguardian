from io import BytesIO

from django.utils import timezone

from apps.study.application.commands.import_crf_templates_template.types import (
    CrfTemplateImportDependencyError,
    CrfTemplateImportFormatError,
)


class CrfTemplateWorkbookMixin:
    def _load_rows_from_workbook(self, *, file_name, file_content):
        file_name_lower = (file_name or "").strip().lower()
        if file_name_lower.endswith(".xlsx"):
            return self._load_xlsx_rows(file_content)
        if file_name_lower.endswith(".xls"):
            return self._load_xls_rows(file_content)
        raise CrfTemplateImportFormatError("Only .xlsx and .xls files are supported.")

    def _load_xlsx_rows(self, file_content):
        try:
            from openpyxl import load_workbook
        except ModuleNotFoundError as exc:
            raise CrfTemplateImportDependencyError(
                "Excel import support for .xlsx is not installed. Install openpyxl first."
            ) from exc

        workbook = load_workbook(filename=BytesIO(file_content), data_only=True, read_only=True)
        rows_by_sheet = {}
        for logical_sheet_name in self.expected_columns:
            actual_sheet_name = self._resolve_sheet_name(
                workbook_sheet_names=workbook.sheetnames,
                logical_sheet_name=logical_sheet_name,
            )
            rows_by_sheet[logical_sheet_name] = self._map_rows(
                rows=list(workbook[actual_sheet_name].iter_rows(values_only=True)),
                sheet_name=logical_sheet_name,
            )
        return rows_by_sheet

    def _load_xls_rows(self, file_content):
        try:
            import xlrd
        except ModuleNotFoundError as exc:
            raise CrfTemplateImportDependencyError(
                "Excel import support for .xls is not installed. Install xlrd first."
            ) from exc

        workbook = xlrd.open_workbook(file_contents=file_content)
        rows_by_sheet = {}
        for logical_sheet_name in self.expected_columns:
            actual_sheet_name = self._resolve_sheet_name(
                workbook_sheet_names=workbook.sheet_names(),
                logical_sheet_name=logical_sheet_name,
            )
            try:
                worksheet = workbook.sheet_by_name(actual_sheet_name)
            except xlrd.biffh.XLRDError as exc:
                raise CrfTemplateImportFormatError(
                    f"Missing required worksheet: {logical_sheet_name}"
                ) from exc
            rows_by_sheet[logical_sheet_name] = self._map_rows(
                rows=[worksheet.row_values(index) for index in range(worksheet.nrows)],
                sheet_name=logical_sheet_name,
            )
        return rows_by_sheet

    def _resolve_sheet_name(self, *, workbook_sheet_names, logical_sheet_name):
        normalized_sheet_map = {
            self._normalize_header(sheet_name): sheet_name
            for sheet_name in workbook_sheet_names
        }
        for candidate_name in self.sheet_aliases.get(logical_sheet_name, (logical_sheet_name,)):
            normalized_candidate = self._normalize_header(candidate_name)
            resolved_sheet_name = normalized_sheet_map.get(normalized_candidate)
            if resolved_sheet_name:
                return resolved_sheet_name

        accepted_names = ", ".join(self.sheet_aliases.get(logical_sheet_name, (logical_sheet_name,)))
        raise CrfTemplateImportFormatError(
            f"Missing required worksheet: {logical_sheet_name}. Accepted names: {accepted_names}"
        )

    def _map_rows(self, *, rows, sheet_name):
        if sheet_name not in self.expected_columns:
            raise CrfTemplateImportFormatError(f"Unexpected worksheet: {sheet_name}")
        if not rows:
            return []

        headers = [self._normalize_header(value) for value in rows[0]]
        missing_headers = [
            header
            for header in self.expected_columns[sheet_name]
            if self._normalize_header(header) not in headers
        ]
        if missing_headers:
            raise CrfTemplateImportFormatError(
                f"Worksheet '{sheet_name}' is missing required columns: " + ", ".join(missing_headers)
            )

        header_keys = [self.expected_header_map[sheet_name].get(header) for header in headers]
        mapped_rows = []
        for row_index, row_values in enumerate(rows[1:], start=2):
            if self._is_blank_row(row_values):
                continue

            row_data = {}
            for column_index, header_key in enumerate(header_keys):
                if not header_key:
                    continue
                row_data[header_key] = row_values[column_index] if column_index < len(row_values) else None

            mapped_rows.append((row_index, row_data))

        return mapped_rows

    @staticmethod
    def _normalize_header(value):
        return " ".join(str(value or "").strip().lower().split())

    @staticmethod
    def _is_blank_row(row_values):
        return all(str(value or "").strip() == "" for value in row_values)

    @staticmethod
    def _as_text(value):
        if value is None:
            return ""
        if isinstance(value, float) and value.is_integer():
            return str(int(value))
        return str(value).strip()

    def _require_text(self, value, *, field_label, max_length):
        normalized_value = self._as_text(value)
        if not normalized_value:
            raise CrfTemplateImportFormatError(f"{field_label} is required.")
        if len(normalized_value) > max_length:
            raise CrfTemplateImportFormatError(
                f"{field_label} must be {max_length} characters or fewer."
            )
        return normalized_value

    def _coerce_positive_int(self, value, *, field_label):
        normalized_value = self._as_text(value)
        if not normalized_value:
            raise CrfTemplateImportFormatError(f"{field_label} is required.")
        try:
            parsed = int(normalized_value)
        except ValueError as exc:
            raise CrfTemplateImportFormatError(f"{field_label} must be an integer.") from exc
        if parsed <= 0:
            raise CrfTemplateImportFormatError(f"{field_label} must be greater than 0.")
        return parsed

    def _coerce_non_negative_int(self, value, *, field_label, default=None):
        normalized_value = self._as_text(value)
        if not normalized_value:
            if default is None:
                raise CrfTemplateImportFormatError(f"{field_label} is required.")
            return default
        try:
            parsed = int(normalized_value)
        except ValueError as exc:
            raise CrfTemplateImportFormatError(f"{field_label} must be an integer.") from exc
        if parsed < 0:
            raise CrfTemplateImportFormatError(f"{field_label} must be greater than or equal to 0.")
        return parsed

    def _coerce_optional_non_negative_int(self, value, *, field_label):
        normalized_value = self._as_text(value)
        if not normalized_value:
            return None
        try:
            parsed = int(normalized_value)
        except ValueError as exc:
            raise CrfTemplateImportFormatError(f"{field_label} must be an integer.") from exc
        if parsed < 0:
            raise CrfTemplateImportFormatError(f"{field_label} must be greater than or equal to 0.")
        return parsed

    def _coerce_bool(self, value, *, field_label, default=None):
        normalized_value = self._as_text(value).lower()
        if normalized_value == "":
            if default is None:
                raise CrfTemplateImportFormatError(f"{field_label} is required.")
            return default

        truthy = {"1", "true", "yes", "y"}
        falsy = {"0", "false", "no", "n"}
        if normalized_value in truthy:
            return True
        if normalized_value in falsy:
            return False
        raise CrfTemplateImportFormatError(
            f"{field_label} must be one of: 1/0, true/false, yes/no."
        )

    def _build_form_template_identifier(self, row_data):
        return " / ".join(
            part for part in (
                self._as_text(row_data.get("code")),
                self._as_text(row_data.get("version")),
            ) if part
        )

    def _build_section_template_identifier(self, row_data):
        return " / ".join(
            part for part in (
                self._as_text(row_data.get("form_code")),
                self._as_text(row_data.get("code")),
                self._as_text(row_data.get("order")),
                self._as_text(row_data.get("repeated")),
            ) if part
        )

    @staticmethod
    def _now():
        return timezone.now()
