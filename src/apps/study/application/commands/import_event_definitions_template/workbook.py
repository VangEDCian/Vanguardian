from io import BytesIO

from django.utils import timezone

from apps.study.application.commands.import_event_definitions_template.types import (
    EventDefinitionImportDependencyError,
    EventDefinitionImportFormatError,
)


class EventDefinitionWorkbookMixin:
    def _load_rows_from_workbook(self, *, file_name, file_content):
        file_name_lower = (file_name or "").strip().lower()
        if file_name_lower.endswith(".xlsx"):
            return self._load_xlsx_rows(file_content)
        if file_name_lower.endswith(".xls"):
            return self._load_xls_rows(file_content)
        raise EventDefinitionImportFormatError("Only .xlsx and .xls files are supported.")

    def _load_xlsx_rows(self, file_content):
        try:
            from openpyxl import load_workbook
        except ModuleNotFoundError as exc:
            raise EventDefinitionImportDependencyError(
                "Excel import support for .xlsx is not installed. Install openpyxl first."
            ) from exc

        workbook = load_workbook(filename=BytesIO(file_content), data_only=True, read_only=True)
        worksheet = workbook.worksheets[0]
        rows = list(worksheet.iter_rows(values_only=True))
        return self._map_rows(rows)

    def _load_xls_rows(self, file_content):
        try:
            import xlrd
        except ModuleNotFoundError as exc:
            raise EventDefinitionImportDependencyError(
                "Excel import support for .xls is not installed. Install xlrd first."
            ) from exc

        workbook = xlrd.open_workbook(file_contents=file_content)
        worksheet = workbook.sheet_by_index(0)
        rows = [worksheet.row_values(index) for index in range(worksheet.nrows)]
        return self._map_rows(rows)

    def _map_rows(self, rows):
        if not rows:
            raise EventDefinitionImportFormatError("The workbook is empty.")

        headers = [self._normalize_header(value) for value in rows[0]]
        missing_headers = [
            header
            for header in self.expected_columns
            if self._normalize_header(header) not in headers
        ]
        if missing_headers:
            raise EventDefinitionImportFormatError(
                "Missing required columns: " + ", ".join(missing_headers)
            )

        header_keys = [self.expected_header_map.get(header) for header in headers]
        mapped_rows = []
        for row_index, row_values in enumerate(rows[1:], start=2):
            if self._is_blank_row(row_values):
                continue

            row_data = {}
            for column_index, header_key in enumerate(header_keys):
                if not header_key:
                    continue
                row_data[header_key] = row_values[column_index] if column_index < len(row_values) else None

            mapped_rows.append((row_index, row_data))

        return mapped_rows

    @staticmethod
    def _normalize_header(value):
        return " ".join(str(value or "").strip().lower().split())

    @staticmethod
    def _is_blank_row(row_values):
        return all(str(value or "").strip() == "" for value in row_values)

    @staticmethod
    def _as_text(value):
        if value is None:
            return ""
        if isinstance(value, float) and value.is_integer():
            return str(int(value))
        return str(value).strip()

    def _nullable_text(self, value):
        normalized = self._as_text(value)
        return normalized or None

    def _resolve_study_version(self, *, study_id, raw_study_version):
        normalized_study_version = self._as_text(raw_study_version)
        existing_study_version = self.repository.find_study_version(
            study_id=study_id,
            study_version=normalized_study_version,
        )
        if existing_study_version:
            return existing_study_version
        return normalized_study_version

    def _coerce_bool(self, value, *, allow_blank=False, default=None):
        normalized = self._as_text(value).lower()
        if not normalized:
            if allow_blank:
                return default
            raise EventDefinitionImportFormatError(f"Invalid boolean value: {value!r}")
        if normalized in self.true_values:
            return True
        if normalized in self.false_values:
            return False
        raise EventDefinitionImportFormatError(f"Invalid boolean value: {value!r}")

    def _coerce_int(self, value, *, field_label, allow_blank=False, default=None):
        normalized = self._as_text(value)
        if not normalized:
            if allow_blank:
                return None
            if default is not None:
                return default
            raise EventDefinitionImportFormatError(f"{field_label} is required.")

        try:
            return int(float(normalized))
        except ValueError as exc:
            raise EventDefinitionImportFormatError(f"{field_label} must be a number.") from exc

    def _normalize_choice(self, *, raw_value, aliases, field_label, allow_blank=False):
        normalized = self._as_text(raw_value).lower().replace("-", " ").replace("_", " ")
        normalized = " ".join(normalized.split())
        if not normalized:
            if allow_blank:
                return None
            raise EventDefinitionImportFormatError(f"{field_label} is required.")

        resolved_value = aliases.get(normalized)
        if resolved_value is None:
            raise EventDefinitionImportFormatError(f"Invalid {field_label}: {raw_value!r}")
        return resolved_value

    @staticmethod
    def _now():
        return timezone.now()
