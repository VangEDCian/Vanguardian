from dataclasses import dataclass
from io import BytesIO

from django.db import transaction
from django.utils import timezone

from apps.core.choices import (
    EventDefinitionCategoryChoices,
    EventDefinitionTimingModeChoices,
    EventDefinitionTypeChoices,
    EventExecutionModeChoices,
    EventTransitionConditionScopeChoices,
    EventTransitionTypeChoices,
)
from apps.study.infrastructure.persistence.models import EventDefinition, EventTransitionRule


@dataclass(frozen=True)
class ImportStudyEventDefinitionsTemplateCommand:
    actor_user_id: int
    study_id: int
    file_name: str
    file_content: bytes


@dataclass(frozen=True)
class EventDefinitionImportIssue:
    row_number: int
    study_code: str
    code: str
    reason: str


@dataclass(frozen=True)
class ImportStudyEventDefinitionsTemplateResult:
    total_rows: int
    created_count: int
    updated_count: int
    skipped_count: int
    issues: tuple[EventDefinitionImportIssue, ...] = ()
    warnings: tuple[str, ...] = ()


class EventDefinitionImportTemplateError(Exception):
    """Base error raised for event definition template import failures."""


class EventDefinitionImportDependencyError(EventDefinitionImportTemplateError):
    """Raised when the Excel parser dependency is missing."""


class EventDefinitionImportFormatError(EventDefinitionImportTemplateError):
    """Raised when the uploaded workbook shape is invalid."""


class ImportStudyEventDefinitionsTemplateService:
    expected_columns = (
        "Study Version",
        "Code",
        "Name",
        "Description",
        "Event Type",
        "Timing Mode",
        "Event Category",
        "Execution Mode",
        "Sequence No",
        "Phase Code",
        "Repeated",
        "Max Repeats",
        "Required",
        "Precondition",
        "Transition Type",
        "Condition Scope",
        "Condition Code",
        "Condition Expression",
        "Offset Days",
        "Window Before Days",
        "Window After Days",
        "Auto Open",
        "Auto Create",
        "Requires Previous Completion",
        "Allow Skip",
    )
    expected_header_map = {
        "study code": "study_code",
        "study version": "study_version",
        "code": "code",
        "name": "name",
        "description": "description",
        "event type": "event_type",
        "timing mode": "timing_mode",
        "event category": "event_category",
        "execution mode": "execution_mode",
        "sequence no": "sequence_no",
        "phase code": "phase_code",
        "repeated": "repeated",
        "max repeats": "max_repeats",
        "required": "required",
        "precondition": "precondition",
        "transition type": "transition_type",
        "condition scope": "condition_scope",
        "condition code": "condition_code",
        "condition expression": "condition_expression",
        "offset days": "offset_days",
        "window before days": "window_before_days",
        "window after days": "window_after_days",
        "auto open": "auto_open",
        "auto create": "auto_create",
        "requires previous completion": "requires_previous_completion",
        "allow skip": "allow_skip",
    }
    event_type_aliases = {
        "visit_based": EventDefinitionTypeChoices.VISIT_BASED,
        "visit based": EventDefinitionTypeChoices.VISIT_BASED,
        "visit-based": EventDefinitionTypeChoices.VISIT_BASED,
        "common": EventDefinitionTypeChoices.COMMON,
        "operational": EventDefinitionTypeChoices.OPERATIONAL,
    }
    timing_mode_aliases = {
        "scheduled": EventDefinitionTimingModeChoices.SCHEDULED,
        "unscheduled": EventDefinitionTimingModeChoices.UNSCHEDULED,
        "conditional": EventDefinitionTimingModeChoices.CONDITIONAL,
    }
    event_category_aliases = {
        "screening": EventDefinitionCategoryChoices.SCREENING,
        "randomization": EventDefinitionCategoryChoices.RANDOMIZATION,
        "treatment": EventDefinitionCategoryChoices.TREATMENT,
        "washout": EventDefinitionCategoryChoices.WASHOUT,
        "follow_up": EventDefinitionCategoryChoices.FOLLOW_UP,
        "follow up": EventDefinitionCategoryChoices.FOLLOW_UP,
        "follow-up": EventDefinitionCategoryChoices.FOLLOW_UP,
        "eos": EventDefinitionCategoryChoices.EOS,
        "end of study": EventDefinitionCategoryChoices.EOS,
        "unscheduled": EventDefinitionCategoryChoices.UNSCHEDULED,
    }
    execution_mode_aliases = {
        "form_entry": EventExecutionModeChoices.FORM_ENTRY,
        "form entry": EventExecutionModeChoices.FORM_ENTRY,
        "form-entry": EventExecutionModeChoices.FORM_ENTRY,
        "workflow_action": EventExecutionModeChoices.WORKFLOW_ACTION,
        "workflow action": EventExecutionModeChoices.WORKFLOW_ACTION,
        "workflow-action": EventExecutionModeChoices.WORKFLOW_ACTION,
        "hybrid": EventExecutionModeChoices.HYBRID,
    }
    transition_type_aliases = {
        "sequential": EventTransitionTypeChoices.SEQUENTIAL,
        "conditional": EventTransitionTypeChoices.CONDITIONAL,
        "manual": EventTransitionTypeChoices.MANUAL,
        "automatic": EventTransitionTypeChoices.AUTOMATIC,
    }
    transition_condition_scope_aliases = {
        "subject": EventTransitionConditionScopeChoices.SUBJECT,
        "subject_event": EventTransitionConditionScopeChoices.SUBJECT_EVENT,
        "subject event": EventTransitionConditionScopeChoices.SUBJECT_EVENT,
        "subject-event": EventTransitionConditionScopeChoices.SUBJECT_EVENT,
        "subject_period": EventTransitionConditionScopeChoices.SUBJECT_PERIOD,
        "subject period": EventTransitionConditionScopeChoices.SUBJECT_PERIOD,
        "subject-period": EventTransitionConditionScopeChoices.SUBJECT_PERIOD,
        "randomization": EventTransitionConditionScopeChoices.RANDOMIZATION,
        "eligibility": EventTransitionConditionScopeChoices.ELIGIBILITY,
    }
    true_values = {"true", "1", "yes", "y"}
    false_values = {"", "false", "0", "no", "n"}

    def execute(self, command: ImportStudyEventDefinitionsTemplateCommand) -> ImportStudyEventDefinitionsTemplateResult:
        workbook_rows = self._load_rows_from_workbook(
            file_name=command.file_name,
            file_content=command.file_content,
        )

        created_count = 0
        updated_count = 0
        issues = []

        for row_number, row_data in workbook_rows:
            try:
                import_outcome = self._import_row(
                    study_id=command.study_id,
                    row_data=row_data,
                    row_number=row_number,
                    actor_user_id=command.actor_user_id,
                )
            except EventDefinitionImportFormatError as exc:
                issues.append(
                    EventDefinitionImportIssue(
                        row_number=row_number,
                        study_code=str(row_data.get("study_code", "") or ""),
                        code=str(row_data.get("code", "") or ""),
                        reason=str(exc),
                    )
                )
                continue

            if import_outcome == "created":
                created_count += 1
            else:
                updated_count += 1

        return ImportStudyEventDefinitionsTemplateResult(
            total_rows=len(workbook_rows),
            created_count=created_count,
            updated_count=updated_count,
            skipped_count=len(issues),
            issues=tuple(issues),
            warnings=(),
        )

    def _import_row(self, *, study_id, row_data, row_number, actor_user_id):
        study_version_input = self._as_text(row_data.get("study_version"))
        if not study_version_input:
            raise EventDefinitionImportFormatError("Study Version is required.")
        if len(study_version_input) > 20:
            raise EventDefinitionImportFormatError("Study Version must be 20 characters or fewer.")

        study_version = self._resolve_study_version(
            study_id=study_id,
            raw_study_version=study_version_input,
        )

        code = self._as_text(row_data.get("code"))
        if not code:
            raise EventDefinitionImportFormatError("Code is required.")

        name = self._as_text(row_data.get("name"))
        if not name:
            raise EventDefinitionImportFormatError("Name is required.")

        event_type = self._normalize_choice(
            raw_value=row_data.get("event_type"),
            aliases=self.event_type_aliases,
            field_label="Event Type",
        )
        timing_mode = self._normalize_choice(
            raw_value=row_data.get("timing_mode"),
            aliases=self.timing_mode_aliases,
            field_label="Timing Mode",
        )
        event_category = self._normalize_choice(
            raw_value=row_data.get("event_category"),
            aliases=self.event_category_aliases,
            field_label="Event Category",
            allow_blank=True,
        )
        execution_mode = self._normalize_choice(
            raw_value=row_data.get("execution_mode"),
            aliases=self.execution_mode_aliases,
            field_label="Execution Mode",
            allow_blank=True,
        ) or EventExecutionModeChoices.FORM_ENTRY
        precondition_code = self._nullable_text(row_data.get("precondition"))

        sequence_no = self._coerce_int(row_data.get("sequence_no"), field_label="Sequence No", default=1)
        max_repeats = self._coerce_int(row_data.get("max_repeats"), field_label="Max Repeats", allow_blank=True)

        if precondition_code:
            transition_type = self._normalize_choice(
                raw_value=row_data.get("transition_type"),
                aliases=self.transition_type_aliases,
                field_label="Transition Type",
                allow_blank=True,
            ) or EventTransitionTypeChoices.SEQUENTIAL
            condition_scope = self._normalize_choice(
                raw_value=row_data.get("condition_scope"),
                aliases=self.transition_condition_scope_aliases,
                field_label="Condition Scope",
                allow_blank=True,
            ) or EventTransitionConditionScopeChoices.SUBJECT_EVENT
            condition_code = self._nullable_text(row_data.get("condition_code"))
            condition_expression = self._nullable_text(row_data.get("condition_expression"))
            offset_days = self._coerce_int(row_data.get("offset_days"), field_label="Offset Days", allow_blank=True)
            window_before_days = self._coerce_int(
                row_data.get("window_before_days"),
                field_label="Window Before Days",
                allow_blank=True,
            )
            window_after_days = self._coerce_int(
                row_data.get("window_after_days"),
                field_label="Window After Days",
                allow_blank=True,
            )
            auto_open = self._coerce_bool(row_data.get("auto_open"), allow_blank=True, default=False)
            auto_create = self._coerce_bool(row_data.get("auto_create"), allow_blank=True, default=False)
            requires_previous_completion = self._coerce_bool(
                row_data.get("requires_previous_completion"),
                allow_blank=True,
                default=True,
            )
            allow_skip = self._coerce_bool(row_data.get("allow_skip"), allow_blank=True, default=False)
        else:
            transition_type = None
            condition_scope = None
            condition_code = None
            condition_expression = None
            offset_days = None
            window_before_days = None
            window_after_days = None
            auto_open = False
            auto_create = False
            requires_previous_completion = True
            allow_skip = False

        now = self._now()
        defaults = {
            "study_id": study_id,
            "study_version": study_version,
            "name": name,
            "description": self._nullable_text(row_data.get("description")),
            "event_type": event_type,
            "timing_mode": timing_mode,
            "event_category": event_category,
            "execution_mode": execution_mode,
            "sequence_no": sequence_no,
            "phase_code": self._nullable_text(row_data.get("phase_code")),
            "is_repeating": self._coerce_bool(row_data.get("repeated")),
            "max_repeats": max_repeats,
            "is_enabled": True,
            "is_required": self._coerce_bool(row_data.get("required")),
            "deleted": False,
            "updated_at": now,
            "updated_by_id": actor_user_id,
        }

        with transaction.atomic():
            event_definition = EventDefinition.objects.filter(
                study_id=study_id,
                study_version=study_version,
                code=code,
            ).first()

            if event_definition is None:
                event_definition = EventDefinition.objects.create(
                    code=code,
                    created_at=now,
                    created_by_id=actor_user_id,
                    **defaults,
                )
                self._sync_transition_rule(
                    event_definition=event_definition,
                    study_id=study_id,
                    study_version=study_version,
                    sequence_no=sequence_no,
                    precondition_code=precondition_code,
                    transition_type=transition_type,
                    condition_scope=condition_scope,
                    condition_code=condition_code,
                    condition_expression=condition_expression,
                    offset_days=offset_days,
                    window_before_days=window_before_days,
                    window_after_days=window_after_days,
                    auto_open=auto_open,
                    auto_create=auto_create,
                    requires_previous_completion=requires_previous_completion,
                    allow_skip=allow_skip,
                    actor_user_id=actor_user_id,
                    now=now,
                )
                return "created"

            for field_name, value in defaults.items():
                setattr(event_definition, field_name, value)
            event_definition.save(update_fields=list(defaults.keys()))
            self._sync_transition_rule(
                event_definition=event_definition,
                study_id=study_id,
                study_version=study_version,
                sequence_no=sequence_no,
                precondition_code=precondition_code,
                transition_type=transition_type,
                condition_scope=condition_scope,
                condition_code=condition_code,
                condition_expression=condition_expression,
                offset_days=offset_days,
                window_before_days=window_before_days,
                window_after_days=window_after_days,
                auto_open=auto_open,
                auto_create=auto_create,
                requires_previous_completion=requires_previous_completion,
                allow_skip=allow_skip,
                actor_user_id=actor_user_id,
                now=now,
            )
            return "updated"

    def _sync_transition_rule(
        self,
        *,
        event_definition,
        study_id,
        study_version,
        sequence_no,
        precondition_code,
        transition_type,
        condition_scope,
        condition_code,
        condition_expression,
        offset_days,
        window_before_days,
        window_after_days,
        auto_open,
        auto_create,
        requires_previous_completion,
        allow_skip,
        actor_user_id,
        now,
    ):
        existing_transition_rules = EventTransitionRule.objects.filter(
            study_id=study_id,
            study_version=study_version,
            to_event_definition=event_definition,
        )

        if not precondition_code:
            existing_transition_rules.filter(deleted=False).update(
                deleted=True,
                updated_at=now,
                updated_by_id=actor_user_id,
            )
            return

        from_event_definition = EventDefinition.objects.filter(
            study_id=study_id,
            study_version=study_version,
            code__iexact=precondition_code,
            deleted=False,
        ).first()
        if from_event_definition is None:
            raise EventDefinitionImportFormatError(
                f"Precondition event {precondition_code!r} was not found in study version {study_version!r}."
            )

        existing_transition_rules.exclude(from_event_definition=from_event_definition).filter(deleted=False).update(
            deleted=True,
            updated_at=now,
            updated_by_id=actor_user_id,
        )

        transition_defaults = {
            "transition_type": transition_type,
            "condition_scope": condition_scope,
            "condition_code": condition_code,
            "condition_expression": condition_expression,
            "offset_days": offset_days,
            "window_before_days": window_before_days,
            "window_after_days": window_after_days,
            "auto_open": auto_open,
            "auto_create": auto_create,
            "requires_previous_completion": requires_previous_completion,
            "allow_skip": allow_skip,
            "display_order": sequence_no,
            "is_enabled": True,
            "deleted": False,
            "updated_at": now,
            "updated_by_id": actor_user_id,
        }

        transition_rule = existing_transition_rules.filter(
            from_event_definition=from_event_definition,
        ).first()

        if transition_rule is None:
            EventTransitionRule.objects.create(
                study_id=study_id,
                study_version=study_version,
                from_event_definition=from_event_definition,
                to_event_definition=event_definition,
                created_at=now,
                created_by_id=actor_user_id,
                **transition_defaults,
            )
            return

        for field_name, value in transition_defaults.items():
            setattr(transition_rule, field_name, value)
        transition_rule.save(update_fields=list(transition_defaults.keys()))

    def _load_rows_from_workbook(self, *, file_name, file_content):
        file_name_lower = (file_name or "").strip().lower()
        if file_name_lower.endswith(".xlsx"):
            return self._load_xlsx_rows(file_content)
        if file_name_lower.endswith(".xls"):
            return self._load_xls_rows(file_content)
        raise EventDefinitionImportFormatError("Only .xlsx and .xls files are supported.")

    def _load_xlsx_rows(self, file_content):
        try:
            from openpyxl import load_workbook
        except ModuleNotFoundError as exc:
            raise EventDefinitionImportDependencyError(
                "Excel import support for .xlsx is not installed. Install openpyxl first."
            ) from exc

        workbook = load_workbook(filename=BytesIO(file_content), data_only=True, read_only=True)
        worksheet = workbook.worksheets[0]
        rows = list(worksheet.iter_rows(values_only=True))
        return self._map_rows(rows)

    def _load_xls_rows(self, file_content):
        try:
            import xlrd
        except ModuleNotFoundError as exc:
            raise EventDefinitionImportDependencyError(
                "Excel import support for .xls is not installed. Install xlrd first."
            ) from exc

        workbook = xlrd.open_workbook(file_contents=file_content)
        worksheet = workbook.sheet_by_index(0)
        rows = [worksheet.row_values(index) for index in range(worksheet.nrows)]
        return self._map_rows(rows)

    def _map_rows(self, rows):
        if not rows:
            raise EventDefinitionImportFormatError("The workbook is empty.")

        headers = [self._normalize_header(value) for value in rows[0]]
        missing_headers = [
            header
            for header in self.expected_columns
            if self._normalize_header(header) not in headers
        ]
        if missing_headers:
            raise EventDefinitionImportFormatError(
                "Missing required columns: " + ", ".join(missing_headers)
            )

        header_keys = [self.expected_header_map.get(header) for header in headers]
        mapped_rows = []
        for row_index, row_values in enumerate(rows[1:], start=2):
            if self._is_blank_row(row_values):
                continue

            row_data = {}
            for column_index, header_key in enumerate(header_keys):
                if not header_key:
                    continue
                row_data[header_key] = row_values[column_index] if column_index < len(row_values) else None

            mapped_rows.append((row_index, row_data))

        return mapped_rows

    @staticmethod
    def _normalize_header(value):
        return " ".join(str(value or "").strip().lower().split())

    @staticmethod
    def _is_blank_row(row_values):
        return all(str(value or "").strip() == "" for value in row_values)

    @staticmethod
    def _as_text(value):
        if value is None:
            return ""
        if isinstance(value, float) and value.is_integer():
            return str(int(value))
        return str(value).strip()

    def _nullable_text(self, value):
        normalized = self._as_text(value)
        return normalized or None

    def _resolve_study_version(self, *, study_id, raw_study_version):
        normalized_study_version = self._as_text(raw_study_version)
        existing_study_version = (
            EventDefinition.objects.filter(
                study_id=study_id,
                study_version__iexact=normalized_study_version,
            )
            .values_list("study_version", flat=True)
            .first()
        )
        if existing_study_version:
            return existing_study_version
        return normalized_study_version

    def _coerce_bool(self, value, *, allow_blank=False, default=None):
        normalized = self._as_text(value).lower()
        if not normalized:
            if allow_blank:
                return default
            raise EventDefinitionImportFormatError(f"Invalid boolean value: {value!r}")
        if normalized in self.true_values:
            return True
        if normalized in self.false_values:
            return False
        raise EventDefinitionImportFormatError(f"Invalid boolean value: {value!r}")

    def _coerce_int(self, value, *, field_label, allow_blank=False, default=None):
        normalized = self._as_text(value)
        if not normalized:
            if allow_blank:
                return None
            if default is not None:
                return default
            raise EventDefinitionImportFormatError(f"{field_label} is required.")

        try:
            return int(float(normalized))
        except ValueError as exc:
            raise EventDefinitionImportFormatError(f"{field_label} must be a number.") from exc

    def _normalize_choice(self, *, raw_value, aliases, field_label, allow_blank=False):
        normalized = self._as_text(raw_value).lower().replace("-", " ").replace("_", " ")
        normalized = " ".join(normalized.split())
        if not normalized:
            if allow_blank:
                return None
            raise EventDefinitionImportFormatError(f"{field_label} is required.")

        resolved_value = aliases.get(normalized)
        if resolved_value is None:
            raise EventDefinitionImportFormatError(f"Invalid {field_label}: {raw_value!r}")
        return resolved_value

    @staticmethod
    def _now():
        return timezone.now()
