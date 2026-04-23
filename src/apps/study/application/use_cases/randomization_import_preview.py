from __future__ import annotations

import csv
import json
from dataclasses import dataclass
from datetime import date, datetime
from io import BytesIO, StringIO
from typing import Any

from django.conf import settings
from django.utils import timezone
from django.utils.translation import gettext_lazy as _

from apps.core.choices.study import RandomizationSchemeStatusChoice


@dataclass(frozen=True)
class RandomizationImportColumn:
    key: str
    label: str
    data_type: str = "string"
    required: bool = True
    max_length: int | None = None
    aliases: tuple[str, ...] = ()


@dataclass(frozen=True)
class RandomizationImportIssue:
    row_number: int
    identifier: str
    column_label: str
    reason: str


@dataclass(frozen=True)
class RandomizationImportParsedRow:
    row_number: int
    identifier: str
    values: dict[str, Any]


@dataclass(frozen=True)
class RandomizationImportPreviewRow:
    row_number: int
    values: tuple[Any, ...]


@dataclass(frozen=True)
class RandomizationImportPreviewResult:
    columns: tuple[RandomizationImportColumn, ...]
    preview_rows: tuple[RandomizationImportPreviewRow, ...]
    parsed_rows: tuple[RandomizationImportParsedRow, ...]
    total_rows: int
    issues: tuple[RandomizationImportIssue, ...] = ()


class RandomizationImportDependencyError(Exception):
    """Raised when a workbook parser dependency is unavailable."""


class RandomizationImportFormatError(Exception):
    """Raised when the uploaded import file structure is invalid."""


class BaseRandomizationImportPreviewUseCase:
    columns: tuple[RandomizationImportColumn, ...] = ()

    def execute(self, *, file_name: str, file_content: bytes) -> RandomizationImportPreviewResult:
        rows = self._load_rows_from_file(file_name=file_name, file_content=file_content)
        mapped_rows = self._map_rows(rows)
        return self._build_preview(mapped_rows)

    def _build_preview(self, mapped_rows):
        issues: list[RandomizationImportIssue] = []
        preview_rows: list[RandomizationImportPreviewRow] = []
        parsed_rows: list[RandomizationImportParsedRow] = []
        seen_duplicate_keys: set[tuple[Any, ...]] = set()

        for row_number, row_data in mapped_rows:
            cleaned_values: dict[str, Any] = {}
            preview_values: list[Any] = []
            row_has_error = False

            for column in self.columns:
                raw_value = row_data.get(column.key)
                try:
                    cleaned_value = self._coerce_value(raw_value=raw_value, column=column)
                    cleaned_values[column.key] = cleaned_value
                    preview_values.append(cleaned_value)
                except RandomizationImportFormatError as exc:
                    row_has_error = True
                    preview_values.append(self._preview_value(raw_value))
                    issues.append(
                        RandomizationImportIssue(
                            row_number=row_number,
                            identifier=self._build_identifier_from_row_data(row_data),
                            column_label=column.label,
                            reason=str(exc),
                        )
                    )

            identifier = self._build_identifier(cleaned_values, row_data)
            preview_rows.append(
                RandomizationImportPreviewRow(
                    row_number=row_number,
                    values=tuple(preview_values),
                )
            )

            if row_has_error:
                continue

            duplicate_key = self._build_duplicate_key(cleaned_values)
            if duplicate_key in seen_duplicate_keys:
                issues.append(
                    RandomizationImportIssue(
                        row_number=row_number,
                        identifier=identifier,
                        column_label=self._get_duplicate_column_label(),
                        reason=str(_("This row duplicates another row in the import file.")),
                    )
                )
                continue

            seen_duplicate_keys.add(duplicate_key)
            parsed_rows.append(
                RandomizationImportParsedRow(
                    row_number=row_number,
                    identifier=identifier,
                    values=cleaned_values,
                )
            )

        return RandomizationImportPreviewResult(
            columns=self.columns,
            preview_rows=tuple(preview_rows),
            parsed_rows=tuple(parsed_rows),
            total_rows=len(mapped_rows),
            issues=tuple(issues),
        )

    def _load_rows_from_file(self, *, file_name, file_content):
        file_name_lower = (file_name or "").strip().lower()
        if file_name_lower.endswith(".xlsx"):
            return self._load_xlsx_rows(file_content)
        if file_name_lower.endswith(".xls"):
            return self._load_xls_rows(file_content)
        if file_name_lower.endswith(".csv"):
            return self._load_csv_rows(file_content)
        raise RandomizationImportFormatError(
            str(_("Only .xlsx, .xls, and .csv files are supported."))
        )

    def _load_xlsx_rows(self, file_content):
        try:
            from openpyxl import load_workbook
        except ModuleNotFoundError as exc:
            raise RandomizationImportDependencyError(
                str(_("Excel import support for .xlsx is not installed. Install openpyxl first."))
            ) from exc

        workbook = load_workbook(filename=BytesIO(file_content), data_only=True, read_only=True)
        worksheet = workbook.worksheets[0]
        return list(worksheet.iter_rows(values_only=True))

    def _load_xls_rows(self, file_content):
        try:
            import xlrd
        except ModuleNotFoundError as exc:
            raise RandomizationImportDependencyError(
                str(_("Excel import support for .xls is not installed. Install xlrd first."))
            ) from exc

        workbook = xlrd.open_workbook(file_contents=file_content)
        worksheet = workbook.sheet_by_index(0)
        return [worksheet.row_values(index) for index in range(worksheet.nrows)]

    def _load_csv_rows(self, file_content):
        decoded_content = None
        for encoding in ("utf-8-sig", "utf-8", "latin-1"):
            try:
                decoded_content = file_content.decode(encoding)
                break
            except UnicodeDecodeError:
                continue

        if decoded_content is None:
            raise RandomizationImportFormatError(
                str(_("The CSV file encoding is not supported."))
            )

        return list(csv.reader(StringIO(decoded_content)))

    def _map_rows(self, rows):
        if not rows:
            raise RandomizationImportFormatError(str(_("The workbook is empty.")))

        normalized_headers = [self._normalize_header(value) for value in rows[0]]
        header_index_map = {
            normalized_header: index
            for index, normalized_header in enumerate(normalized_headers)
            if normalized_header
        }

        missing_headers = [
            column.label
            for column in self.columns
            if column.required and self._find_header_index(column, header_index_map) is None
        ]
        if missing_headers:
            raise RandomizationImportFormatError(
                str(_("Missing required columns: %(columns)s") % {"columns": ", ".join(missing_headers)})
            )

        mapped_rows = []
        for row_index, row_values in enumerate(rows[1:], start=2):
            if self._is_blank_row(row_values):
                continue

            row_data = {}
            for column in self.columns:
                header_index = self._find_header_index(column, header_index_map)
                if header_index is None:
                    row_data[column.key] = None
                    continue
                column_index = int(header_index)
                row_data[column.key] = row_values[column_index] if column_index < len(row_values) else None

            mapped_rows.append((row_index, row_data))

        if not mapped_rows:
            raise RandomizationImportFormatError(
                str(_("The workbook does not contain any data rows."))
            )

        return mapped_rows

    def _find_header_index(self, column, header_index_map):
        for candidate in (column.label, *column.aliases):
            normalized_candidate = self._normalize_header(candidate)
            if normalized_candidate in header_index_map:
                return header_index_map[normalized_candidate]
        return None

    @staticmethod
    def _normalize_header(value):
        return " ".join(str(value or "").strip().lower().split())

    @staticmethod
    def _is_blank_row(row_values):
        return all(str(value or "").strip() == "" for value in row_values)

    @staticmethod
    def _as_text(value):
        if value is None:
            return ""
        if isinstance(value, float) and value.is_integer():
            return str(int(value))
        return str(value).strip()

    def _preview_value(self, value):
        if value is None:
            return ""
        if isinstance(value, (dict, list)):
            return json.dumps(value, ensure_ascii=True)
        if isinstance(value, bool):
            return value
        if isinstance(value, (int, float)):
            return value
        if isinstance(value, datetime):
            return value.isoformat(sep=" ", timespec="seconds")
        if isinstance(value, date):
            return value.isoformat()
        return self._as_text(value)

    def _coerce_value(self, *, raw_value, column):
        if self._is_empty_value(raw_value):
            if column.required:
                raise RandomizationImportFormatError(
                    str(_("%(column)s is required.") % {"column": column.label})
                )
            return ""

        if column.data_type == "integer":
            return self._coerce_int(raw_value, field_label=column.label)
        if column.data_type == "boolean":
            return self._coerce_bool(raw_value, field_label=column.label)

        normalized_text = self._as_text(raw_value)
        if column.max_length and len(normalized_text) > column.max_length:
            raise RandomizationImportFormatError(
                str(
                    _("%(column)s must be %(max_length)s characters or fewer.")
                    % {"column": column.label, "max_length": column.max_length}
                )
            )
        return normalized_text

    @staticmethod
    def _is_empty_value(value):
        return value is None or str(value).strip() == ""

    def _coerce_int(self, value, *, field_label):
        if isinstance(value, bool):
            raise RandomizationImportFormatError(
                str(_("%(column)s must be a whole number.") % {"column": field_label})
            )
        if isinstance(value, int):
            return value
        if isinstance(value, float):
            if value.is_integer():
                return int(value)
            raise RandomizationImportFormatError(
                str(_("%(column)s must be a whole number.") % {"column": field_label})
            )

        normalized_value = self._as_text(value).replace(",", "")
        try:
            if "." in normalized_value:
                parsed_value = float(normalized_value)
                if not parsed_value.is_integer():
                    raise ValueError
                return int(parsed_value)
            return int(normalized_value)
        except ValueError as exc:
            raise RandomizationImportFormatError(
                str(_("%(column)s must be a whole number.") % {"column": field_label})
            ) from exc

    def _coerce_bool(self, value, *, field_label):
        if isinstance(value, bool):
            return value
        if isinstance(value, (int, float)) and value in (0, 1):
            return bool(value)

        normalized_value = self._as_text(value).lower()
        if normalized_value in {"true", "1", "yes", "y", "on"}:
            return True
        if normalized_value in {"false", "0", "no", "n", "off"}:
            return False
        raise RandomizationImportFormatError(
            str(_("%(column)s must be Yes/No or True/False.") % {"column": field_label})
        )

    def _build_identifier(self, cleaned_values, row_data):
        raise NotImplementedError

    def _build_identifier_from_row_data(self, row_data):
        raise NotImplementedError

    def _build_duplicate_key(self, cleaned_values):
        raise NotImplementedError

    def _get_duplicate_column_label(self):
        raise NotImplementedError


class RandomizationSchemeImportPreviewUseCase(BaseRandomizationImportPreviewUseCase):
    allowed_status_values = {choice.value for choice in RandomizationSchemeStatusChoice}

    columns = (
        RandomizationImportColumn("code", "Code", max_length=64),
        RandomizationImportColumn("name", "Name", max_length=255),
        RandomizationImportColumn(
            "randomization_type",
            "Type",
            max_length=32,
            aliases=("Randomization Type",),
        ),
        RandomizationImportColumn(
            "allocation_ratio_json",
            "Allocation Ratio",
            required=False,
            aliases=("Allocation",),
        ),
        RandomizationImportColumn(
            "target_randomized_total",
            "Target Randomized Total",
            data_type="integer",
            aliases=("Target Total",),
        ),
        RandomizationImportColumn(
            "is_open_label",
            "Is Open Label",
            data_type="boolean",
            aliases=("Open Label",),
        ),
        RandomizationImportColumn(
            "requires_screening_pass",
            "Requires Screening Pass",
            data_type="boolean",
            aliases=("Requires Screening",),
        ),
        RandomizationImportColumn(
            "eligibility_rule_code",
            "Eligibility Rule Code",
            required=False,
            max_length=64,
            aliases=("Eligibility Rule",),
        ),
        RandomizationImportColumn(
            "effective_from",
            "Effective From",
            required=False,
        ),
        RandomizationImportColumn(
            "effective_to",
            "Effective To",
            required=False,
        ),
        RandomizationImportColumn(
            "status",
            "Status",
            required=False,
            max_length=32,
            aliases=("Scheme Status", "Randomization Status"),
        ),
        RandomizationImportColumn(
            "notes",
            "Notes",
            required=False,
        ),
    )

    def _coerce_value(self, *, raw_value, column):
        if column.key == "allocation_ratio_json":
            return self._coerce_allocation_ratio(raw_value, field_label=column.label)
        if column.key in {"effective_from", "effective_to"}:
            return self._coerce_datetime(raw_value, field_label=column.label)
        if column.key == "status":
            return self._coerce_status(raw_value, field_label=column.label)
        return super()._coerce_value(raw_value=raw_value, column=column)

    def _coerce_allocation_ratio(self, value, *, field_label):
        if self._is_empty_value(value):
            return ""

        ratio_mapping = self._parse_allocation_ratio(value, field_label=field_label)
        normalized_mapping = {}
        for arm_code, ratio in ratio_mapping.items():
            normalized_code = self._as_text(arm_code)
            if not normalized_code:
                raise RandomizationImportFormatError(
                    str(_("%(column)s contains an empty ARM code.") % {"column": field_label})
                )
            normalized_mapping[normalized_code] = self._coerce_int(
                ratio,
                field_label=field_label,
            )

        for arm_code, ratio in normalized_mapping.items():
            if ratio <= 0:
                raise RandomizationImportFormatError(
                    str(
                        _("%(column)s ratio for %(arm_code)s must be greater than 0.")
                        % {"column": field_label, "arm_code": arm_code}
                    )
                )

        if not normalized_mapping:
            raise RandomizationImportFormatError(
                str(_("%(column)s must contain at least one ARM ratio.") % {"column": field_label})
            )

        return normalized_mapping

    def _parse_allocation_ratio(self, value, *, field_label) -> dict[str, Any]:
        if isinstance(value, dict):
            return value

        text_value = self._as_text(value)
        if text_value.startswith("{"):
            try:
                parsed = json.loads(text_value)
            except json.JSONDecodeError as exc:
                raise RandomizationImportFormatError(
                    str(
                        _(
                            "%(column)s must be valid JSON object or shorthand like ARM-A:2;ARM-B:1.",
                        )
                        % {"column": field_label}
                    )
                ) from exc
            if not isinstance(parsed, dict):
                raise RandomizationImportFormatError(
                    str(_("%(column)s must be a JSON object.") % {"column": field_label})
                )
            return parsed

        pairs = [segment.strip() for segment in text_value.replace(",", ";").split(";") if segment.strip()]
        mapping: dict[str, Any] = {}
        for pair in pairs:
            if ":" not in pair:
                raise RandomizationImportFormatError(
                    str(
                        _(
                            "%(column)s must be valid JSON object or shorthand like ARM-A:2;ARM-B:1.",
                        )
                        % {"column": field_label}
                    )
                )
            arm_code, ratio = pair.split(":", 1)
            mapping[self._as_text(arm_code)] = self._as_text(ratio)
        return mapping

    def _coerce_datetime(self, value, *, field_label):
        if self._is_empty_value(value):
            return ""

        if isinstance(value, datetime):
            return self._ensure_aware_datetime(value)
        if isinstance(value, date):
            return self._ensure_aware_datetime(
                datetime.combine(value, datetime.min.time()),
            )

        text_value = self._as_text(value).replace("T", " ")
        try:
            return self._ensure_aware_datetime(datetime.fromisoformat(text_value))
        except ValueError as exc:
            raise RandomizationImportFormatError(
                str(
                    _("%(column)s must be a valid datetime. Example: 2026-04-21 08:30")
                    % {"column": field_label}
                )
            ) from exc

    @staticmethod
    def _ensure_aware_datetime(value):
        if not settings.USE_TZ or timezone.is_aware(value):
            return value
        return timezone.make_aware(value, timezone.get_current_timezone())

    def _coerce_status(self, value, *, field_label):
        if self._is_empty_value(value):
            return ""

        normalized_value = self._as_text(value).lower()
        if normalized_value in self.allowed_status_values:
            return normalized_value
        raise RandomizationImportFormatError(
            str(
                _("%(column)s must be one of: %(choices)s")
                % {
                    "column": field_label,
                    "choices": ", ".join(sorted(self.allowed_status_values)),
                }
            )
        )

    def _build_identifier(self, cleaned_values, row_data):
        return str(cleaned_values.get("code") or self._build_identifier_from_row_data(row_data))

    def _build_identifier_from_row_data(self, row_data):
        return self._as_text(row_data.get("code")) or str(_("(no code)"))

    def _build_duplicate_key(self, cleaned_values):
        return (str(cleaned_values.get("code", "")).lower(),)

    def _get_duplicate_column_label(self):
        return "Code"


class RandomizationArmImportPreviewUseCase(BaseRandomizationImportPreviewUseCase):
    columns = (
        RandomizationImportColumn(
            "scheme_code",
            "Scheme Code",
            max_length=64,
            aliases=("Scheme",),
        ),
        RandomizationImportColumn(
            "arm_code",
            "Code",
            max_length=32,
            aliases=("Arm Code",),
        ),
        RandomizationImportColumn(
            "arm_name",
            "Name",
            max_length=255,
            aliases=("Arm Name",),
        ),
        RandomizationImportColumn(
            "target_count",
            "Target Count",
            data_type="integer",
        ),
        RandomizationImportColumn(
            "display_order",
            "Display Order",
            data_type="integer",
            aliases=("Order",),
        ),
        RandomizationImportColumn(
            "is_active",
            "Is Active",
            data_type="boolean",
            required=False,
            aliases=("Active",),
        ),
        RandomizationImportColumn(
            "notes",
            "Notes",
            required=False,
        ),
    )

    def _build_identifier(self, cleaned_values, row_data):
        scheme_code = cleaned_values.get("scheme_code") or self._as_text(row_data.get("scheme_code"))
        arm_code = cleaned_values.get("arm_code") or self._as_text(row_data.get("arm_code"))
        return " / ".join(part for part in (str(scheme_code).strip(), str(arm_code).strip()) if part)

    def _build_identifier_from_row_data(self, row_data):
        return " / ".join(
            part
            for part in (
                self._as_text(row_data.get("scheme_code")),
                self._as_text(row_data.get("arm_code")),
            )
            if part
        ) or str(_("(no code)"))

    def _build_duplicate_key(self, cleaned_values):
        return (
            str(cleaned_values.get("scheme_code", "")).lower(),
            str(cleaned_values.get("arm_code", "")).lower(),
        )

    def _get_duplicate_column_label(self):
        return "Scheme Code / Code"

