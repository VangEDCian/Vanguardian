from io import BytesIO

from django.core.files.uploadedfile import SimpleUploadedFile
from django.test import TestCase
from openpyxl import Workbook

from apps.identity.application.services.role_permission_import import IdentityRolePermissionImportService
from apps.identity.models import IdentityPermission, Role, RoleScopeLevel


class IdentityRolePermissionImportServiceTests(TestCase):
    def _build_workbook_upload(self):
        workbook = Workbook()
        worksheet = workbook.active
        worksheet.append(
            [
                "role_name",
                "scope_level",
                "description",
                "permission",
            ]
        )
        worksheet.append(
            [
                "Site Coordinator",
                "STUDY_SITE",
                "Input/edit/data; Query response",
                "dashboard.test_import_permission",
            ]
        )
        worksheet.append(
            [
                "Site Coordinator",
                "STUDY_SITE",
                "Input/edit/data; Query response",
                "dashboard.missing_permission",
            ]
        )

        buffer = BytesIO()
        workbook.save(buffer)
        return SimpleUploadedFile(
            "role_permission_proposal.xlsx",
            buffer.getvalue(),
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )

    def test_import_creates_study_role_and_skips_missing_permission(self):
        permission = IdentityPermission.objects.create(
            app_label="dashboard",
            codename="test_import_permission",
            name="Can test import permission",
        )

        result = IdentityRolePermissionImportService().import_workbook(
            study_id=3,
            import_file=self._build_workbook_upload(),
        )

        self.assertEqual(result["total_rows"], 2)
        self.assertEqual(result["imported_rows"], 1)
        self.assertEqual(result["skipped_rows"], 1)
        self.assertEqual(result["created_roles"], 1)
        self.assertEqual(result["role_permission_links"], 1)
        self.assertIn("Row 3: permission 'dashboard.missing_permission' does not exist.", result["issues"])

        role = Role.objects.get(study_id=3, name="Site Coordinator")
        self.assertEqual(role.description, "Input/edit/data; Query response")
        self.assertEqual(role.scope_level, RoleScopeLevel.STUDY_SITE)
        self.assertEqual(list(role.permissions.values_list("pk", flat=True)), [permission.pk])
        self.assertFalse(Role.objects.filter(study_id=4, name="Site Coordinator").exists())

    def test_import_skips_invalid_scope_level(self):
        workbook = Workbook()
        worksheet = workbook.active
        worksheet.append(["role_name", "scope_level", "description", "permission"])
        worksheet.append(["Study Reviewer", "SITE", "Review source", "dashboard.test_import_permission"])
        buffer = BytesIO()
        workbook.save(buffer)
        upload = SimpleUploadedFile(
            "role_permission_proposal.xlsx",
            buffer.getvalue(),
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )

        result = IdentityRolePermissionImportService().import_workbook(
            study_id=3,
            import_file=upload,
        )

        self.assertEqual(result["imported_rows"], 0)
        self.assertEqual(result["skipped_rows"], 1)
        self.assertIn("Row 2: invalid scope_level 'SITE'.", result["issues"])

    def test_import_updates_existing_role_scope_level(self):
        IdentityPermission.objects.create(
            app_label="dashboard",
            codename="test_import_permission",
            name="Can test import permission",
        )
        Role.objects.create(
            study_id=3,
            name="Study Reviewer",
            code="STUDY_REVIEWER",
            scope_level=RoleScopeLevel.STUDY_SITE,
        )
        workbook = Workbook()
        worksheet = workbook.active
        worksheet.append(["role_name", "scope_level", "description", "permission"])
        worksheet.append(
            [
                "Study Reviewer",
                "STUDY",
                "Study-level reviewer",
                "dashboard.test_import_permission",
            ]
        )
        buffer = BytesIO()
        workbook.save(buffer)
        upload = SimpleUploadedFile(
            "role_permission_proposal.xlsx",
            buffer.getvalue(),
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )

        result = IdentityRolePermissionImportService().import_workbook(
            study_id=3,
            import_file=upload,
        )

        role = Role.objects.get(study_id=3, name="Study Reviewer")
        self.assertEqual(result["updated_roles"], 1)
        self.assertEqual(role.scope_level, RoleScopeLevel.STUDY)
        self.assertEqual(role.description, "Study-level reviewer")

    def test_import_rejects_csv_upload(self):
        upload = SimpleUploadedFile(
            "role_permission_proposal.csv",
            b"role_name,description,permission\n",
            content_type="text/csv",
        )

        with self.assertRaisesMessage(ValueError, "Only .xlsx and .xlsm files are supported."):
            IdentityRolePermissionImportService().import_workbook(
                study_id=3,
                import_file=upload,
            )

    def test_create_role_links_existing_permissions(self):
        permission = IdentityPermission.objects.create(
            app_label="study",
            codename="create_role_permission",
            name="Can create role permission",
        )

        result = IdentityRolePermissionImportService().create_role(
            study_id=5,
            name="Study Reviewer",
            code="STUDY_REVIEWER",
            description="Review source data",
            scope_level=RoleScopeLevel.STUDY,
            permission_ids=(permission.pk,),
        )

        role = Role.objects.get(study_id=5, name="Study Reviewer")
        self.assertEqual(result["permission_count"], 1)
        self.assertEqual(role.code, "STUDY_REVIEWER")
        self.assertEqual(role.scope_level, RoleScopeLevel.STUDY)
        self.assertEqual(list(role.permissions.values_list("pk", flat=True)), [permission.pk])

    def test_create_role_rejects_duplicate_study_role_name(self):
        Role.objects.create(
            study_id=5,
            name="Study Reviewer",
            code="STUDY_REVIEWER",
            scope_level=RoleScopeLevel.STUDY,
        )

        with self.assertRaises(ValueError):
            IdentityRolePermissionImportService().create_role(
                study_id=5,
                name="Study Reviewer",
            )
