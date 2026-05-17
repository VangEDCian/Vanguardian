import json
from pathlib import Path
from types import SimpleNamespace
from unittest.mock import ANY, MagicMock, patch

from django.test import SimpleTestCase
from openpyxl import load_workbook

from apps.study.application import (
    RandomizationImportDependencyError,
    RandomizationImportFormatError,
)
from apps.study.application.commands.import_randomization import (
    CommitRandomizationImportCommand,
    PreviewRandomizationImportCommand,
    RandomizationImportValidationError,
)
from apps.study.application.services import (
    CommitStudyRandomizationArmsImportService,
    CommitStudyRandomizationSchemesImportService,
    PreviewStudyRandomizationArmsImportService,
)
from apps.study.application.services.randomization_slot_generation import (
    RandomizationSlotGenerationError,
)
from apps.study.application.use_cases.randomization_import_preview import (
    RandomizationArmImportPreviewUseCase,
    RandomizationImportIssue,
    RandomizationImportParsedRow,
    RandomizationImportPreviewResult,
    RandomizationSchemeImportPreviewUseCase,
)
from apps.study.presentation.web.views.randomization import (
    StudyRandomizationCommitBaseView,
    StudyRandomizationImportBaseView,
)


class RandomizationSchemeImportPreviewUseCaseTests(SimpleTestCase):
    def setUp(self):
        self.use_case = RandomizationSchemeImportPreviewUseCase()

    def test_static_template_headers_match_expected_columns(self):
        workbook = load_workbook(
            Path(__file__).resolve().parents[2] / "src/staticfiles/study/templates/randomization_scheme_template.xlsx",
            read_only=True,
        )
        worksheet = workbook.worksheets[0]
        headers = list(next(worksheet.iter_rows(values_only=True)))

        self.assertEqual(
            [str(header).strip() for header in headers],
            [column.label.strip() for column in self.use_case.columns],
        )

    def test_execute_parses_csv_and_converts_integer_and_boolean_columns(self):
        csv_content = "\n".join(
            [
                "Code,Name,Type,Allocation Ratio,Target Randomized Total,Is Open Label,Requires Screening Pass,Eligibility Rule Code",
                "SCH-001,Main Scheme,block,,100,Yes,No,ELIG-01",
            ]
        ).encode("utf-8")

        result = self.use_case.execute(
            file_name="randomization_scheme.csv",
            file_content=csv_content,
        )

        self.assertEqual(result.total_rows, 1)
        self.assertEqual(result.issues, ())
        self.assertEqual(result.preview_rows[0].values[3], "")
        self.assertEqual(result.preview_rows[0].values[4], 100)
        self.assertTrue(result.preview_rows[0].values[5])
        self.assertFalse(result.preview_rows[0].values[6])
        self.assertEqual(result.parsed_rows[0].values["target_randomized_total"], 100)

    def test_execute_reports_duplicate_scheme_codes(self):
        csv_content = "\n".join(
            [
                "Code,Name,Type,Target Randomized Total,Is Open Label,Requires Screening Pass,Eligibility Rule Code",
                "SCH-001,Main Scheme,block,100,Yes,No,ELIG-01",
                "SCH-001,Backup Scheme,block,120,No,Yes,ELIG-02",
            ]
        ).encode("utf-8")

        result = self.use_case.execute(
            file_name="randomization_scheme.csv",
            file_content=csv_content,
        )

        self.assertEqual(result.total_rows, 2)
        self.assertEqual(len(result.issues), 1)
        self.assertEqual(result.issues[0].row_number, 3)
        self.assertEqual(result.issues[0].column_label, "Code")

    def test_execute_accepts_scheme_status_alias_header(self):
        csv_content = "\n".join(
            [
                "Code,Name,Type,Target Randomized Total,Is Open Label,Requires Screening Pass,Scheme Status",
                "SCH-001,Main Scheme,block,100,Yes,No,closed",
            ]
        ).encode("utf-8")

        result = self.use_case.execute(
            file_name="randomization_scheme.csv",
            file_content=csv_content,
        )

        self.assertEqual(result.issues, ())
        self.assertEqual(result.parsed_rows[0].values["status"], "closed")


class RandomizationArmImportPreviewTests(SimpleTestCase):
    def test_static_template_headers_match_expected_columns(self):
        workbook = load_workbook(
            Path(__file__).resolve().parents[2] / "src/staticfiles/study/templates/randomization_arms_template.xlsx",
            read_only=True,
        )
        worksheet = workbook.worksheets[0]
        headers = list(next(worksheet.iter_rows(values_only=True)))

        self.assertEqual(
            [str(header).strip() for header in headers],
            [column.label.strip() for column in RandomizationArmImportPreviewUseCase.columns],
        )

    def test_preview_service_adds_issue_for_unknown_scheme_code(self):
        repository = MagicMock()
        repository.list_active_scheme_map.return_value = {"sch-existing": SimpleNamespace(code="SCH-EXISTING")}
        repository.list_arm_map.return_value = {}

        preview_service = PreviewStudyRandomizationArmsImportService(repository=repository)
        command = PreviewRandomizationImportCommand(
            actor_user_id=9,
            study_id=3,
            file_name="randomization_arms.csv",
            file_content="\n".join(
                [
                    "Scheme Code,Code,Name,Target Count,Display Order",
                    "SCH-MISSING,ARM-A,Active Comparator,50,1",
                ]
            ).encode("utf-8"),
        )

        result = preview_service.execute(command)

        self.assertEqual(result.total_rows, 1)
        self.assertEqual(len(result.issues), 1)
        self.assertEqual(result.issues[0].column_label, "Scheme Code")
        self.assertIn("SCH-MISSING", result.issues[0].reason)


class CommitStudyRandomizationSchemesImportServiceTests(SimpleTestCase):
    def setUp(self):
        self.service = CommitStudyRandomizationSchemesImportService()
        self.command = CommitRandomizationImportCommand(
            actor_user_id=9,
            study_id=3,
            file_name="scheme.xlsx",
            file_content=b"xlsx",
        )

    def test_execute_raises_validation_error_when_preview_has_issues(self):
        preview_result = RandomizationImportPreviewResult(
            columns=(),
            preview_rows=(),
            parsed_rows=(),
            total_rows=1,
            issues=(
                RandomizationImportIssue(
                    row_number=2,
                    identifier="SCH-001",
                    column_label="Code",
                    reason="Code is required.",
                ),
            ),
        )
        self.service.preview_service = MagicMock()
        self.service.preview_service.execute.return_value = preview_result

        with self.assertRaises(RandomizationImportValidationError):
            self.service.execute(self.command)

    @patch("apps.study.application.services.import_randomization_commit.transaction.atomic")
    def test_execute_counts_created_and_updated_rows(self, mock_atomic):
        mock_atomic.return_value.__enter__.return_value = None
        mock_atomic.return_value.__exit__.return_value = None
        preview_result = RandomizationImportPreviewResult(
            columns=(),
            preview_rows=(),
            parsed_rows=(
                RandomizationImportParsedRow(2, "SCH-001", {"code": "SCH-001"}),
                RandomizationImportParsedRow(3, "SCH-002", {"code": "SCH-002"}),
            ),
            total_rows=2,
            issues=(),
        )
        self.service.preview_service = MagicMock()
        self.service.preview_service.execute.return_value = preview_result
        created_scheme = SimpleNamespace(pk=101, status="draft")
        updated_scheme = SimpleNamespace(pk=102, status="draft")
        self.service._upsert_scheme = MagicMock(
            side_effect=[
                ("created", created_scheme, {}),
                ("updated", updated_scheme, {"code": "SCH-002"}),
            ]
        )
        self.service.randomization_audit_service = MagicMock()

        result = self.service.execute(self.command)

        self.assertEqual(result.total_rows, 2)
        self.assertEqual(result.created_count, 1)
        self.assertEqual(result.updated_count, 1)
        self.service.randomization_audit_service.record_scheme_inserted_by_import.assert_called_once_with(
            scheme=created_scheme,
            actor_user_id=9,
        )
        self.service.randomization_audit_service.record_scheme_updated_by_import.assert_called_once_with(
            scheme=updated_scheme,
            actor_user_id=9,
            before_data={"code": "SCH-002"},
        )

    def test_upsert_scheme_create_uses_expected_payload(self):
        repository = MagicMock()
        repository.get_scheme_by_code.return_value = None
        service = CommitStudyRandomizationSchemesImportService(repository=repository)
        parsed_row = RandomizationImportParsedRow(
            row_number=2,
            identifier="SCH-001",
            values={
                "code": "SCH-001",
                "name": "Main",
                "randomization_type": "block",
                "target_randomized_total": 100,
                "eligibility_rule_code": "RULE-1",
                "requires_screening_pass": True,
                "is_open_label": False,
            },
        )

        outcome = service._upsert_scheme(
            study_id=3,
            parsed_row=parsed_row,
            actor_user_id=9,
            now=ANY,
        )

        self.assertEqual(outcome[0], "created")
        self.assertEqual(outcome[2], {})
        repository.create_scheme.assert_called_once_with(
            study_id=3,
            code="SCH-001",
            status="draft",
            allocation_ratio_json=None,
            effective_from=None,
            effective_to=None,
            approved_by_id=None,
            notes=None,
            created_at=ANY,
            created_by_id=9,
            name="Main",
            randomization_type="block",
            target_randomized_total=100,
            eligibility_rule_code="RULE-1",
            requires_screening_pass=True,
            is_open_label=False,
            deleted=False,
            updated_at=ANY,
        )

    def test_upsert_scheme_update_preserves_existing_status_when_import_status_empty(self):
        existing_scheme = SimpleNamespace(
            study_id=3,
            code="SCH-001",
            name="Main",
            randomization_type="block",
            allocation_ratio_json=None,
            target_randomized_total=100,
            eligibility_rule_code="RULE-1",
            requires_screening_pass=True,
            is_open_label=False,
            status="active",
            effective_from=None,
            effective_to=None,
            deleted=False,
            notes=None,
            save=MagicMock(),
        )
        repository = MagicMock()
        repository.get_scheme_by_code.return_value = existing_scheme
        service = CommitStudyRandomizationSchemesImportService(repository=repository)
        parsed_row = RandomizationImportParsedRow(
            row_number=2,
            identifier="SCH-001",
            values={
                "code": "SCH-001",
                "name": "Main",
                "randomization_type": "block",
                "target_randomized_total": 100,
                "eligibility_rule_code": "RULE-1",
                "requires_screening_pass": True,
                "is_open_label": False,
                "status": "",
            },
        )

        outcome = service._upsert_scheme(
            study_id=3,
            parsed_row=parsed_row,
            actor_user_id=9,
            now=ANY,
        )

        self.assertEqual(outcome[0], "updated")
        self.assertEqual(outcome[1], existing_scheme)
        self.assertEqual(outcome[2]["status"], "active")
        self.assertEqual(existing_scheme.status, "active")
        repository.save_scheme.assert_called_once()

    def test_upsert_scheme_update_overrides_status_when_import_has_status(self):
        existing_scheme = SimpleNamespace(
            study_id=3,
            code="SCH-001",
            name="Main",
            randomization_type="block",
            allocation_ratio_json=None,
            target_randomized_total=100,
            eligibility_rule_code="RULE-1",
            requires_screening_pass=True,
            is_open_label=False,
            status="draft",
            effective_from=None,
            effective_to=None,
            deleted=False,
            notes=None,
            save=MagicMock(),
        )
        repository = MagicMock()
        repository.get_scheme_by_code.return_value = existing_scheme
        service = CommitStudyRandomizationSchemesImportService(repository=repository)
        parsed_row = RandomizationImportParsedRow(
            row_number=2,
            identifier="SCH-001",
            values={
                "code": "SCH-001",
                "name": "Main",
                "randomization_type": "block",
                "target_randomized_total": 100,
                "eligibility_rule_code": "RULE-1",
                "requires_screening_pass": True,
                "is_open_label": False,
                "status": "active",
            },
        )

        outcome = service._upsert_scheme(
            study_id=3,
            parsed_row=parsed_row,
            actor_user_id=9,
            now=ANY,
        )

        self.assertEqual(outcome[0], "updated")
        self.assertEqual(outcome[1], existing_scheme)
        self.assertEqual(outcome[2]["status"], "draft")
        self.assertEqual(existing_scheme.status, "active")
        repository.save_scheme.assert_called_once()


class CommitStudyRandomizationArmsImportServiceTests(SimpleTestCase):
    def setUp(self):
        self.service = CommitStudyRandomizationArmsImportService()
        self.command = CommitRandomizationImportCommand(
            actor_user_id=9,
            study_id=3,
            file_name="arms.xlsx",
            file_content=b"xlsx",
        )

    def test_execute_raises_validation_error_when_preview_has_issues(self):
        preview_result = RandomizationImportPreviewResult(
            columns=(),
            preview_rows=(),
            parsed_rows=(),
            total_rows=1,
            issues=(
                RandomizationImportIssue(
                    row_number=2,
                    identifier="SCH-001 / ARM-A",
                    column_label="Scheme Code",
                    reason="Scheme code SCH-001 was not found.",
                ),
            ),
        )
        self.service.preview_service = MagicMock()
        self.service.preview_service.execute.return_value = preview_result

        with self.assertRaises(RandomizationImportValidationError):
            self.service.execute(self.command)

    @patch("apps.study.application.services.import_randomization_commit.transaction.atomic")
    def test_execute_counts_created_and_updated_rows(self, mock_atomic):
        mock_atomic.return_value.__enter__.return_value = None
        mock_atomic.return_value.__exit__.return_value = None
        preview_result = RandomizationImportPreviewResult(
            columns=(),
            preview_rows=(),
            parsed_rows=(
                RandomizationImportParsedRow(2, "SCH-001 / ARM-A", {"scheme_code": "SCH-001", "arm_code": "ARM-A"}),
                RandomizationImportParsedRow(3, "SCH-001 / ARM-B", {"scheme_code": "SCH-001", "arm_code": "ARM-B"}),
            ),
            total_rows=2,
            issues=(),
        )
        self.service.preview_service = MagicMock()
        self.service.preview_service.execute.return_value = preview_result
        self.service._build_scheme_map = MagicMock(return_value={"sch-001": SimpleNamespace(pk=11)})
        scheme = SimpleNamespace(pk=11, status="draft")
        created_arm = SimpleNamespace(pk=201, is_active=True)
        updated_arm = SimpleNamespace(pk=202, is_active=True)
        self.service._upsert_arm = MagicMock(
            side_effect=[
                ("created", scheme, created_arm, {}),
                ("updated", scheme, updated_arm, {"arm_code": "ARM-B"}),
            ]
        )
        self.service._generate_slots_for_impacted_scheme = MagicMock()
        self.service.randomization_audit_service = MagicMock()

        result = self.service.execute(self.command)

        self.assertEqual(result.total_rows, 2)
        self.assertEqual(result.created_count, 1)
        self.assertEqual(result.updated_count, 1)
        self.service.randomization_audit_service.record_arm_inserted_by_import.assert_called_once_with(
            arm=created_arm,
            actor_user_id=9,
        )
        self.service.randomization_audit_service.record_arm_updated_by_import.assert_called_once_with(
            arm=updated_arm,
            actor_user_id=9,
            before_data={"arm_code": "ARM-B"},
        )
        self.service._generate_slots_for_impacted_scheme.assert_called_once_with(
            scheme=scheme,
            parsed_row=preview_result.parsed_rows[1],
        )

    @patch("apps.study.application.services.import_randomization_commit.transaction.atomic")
    def test_execute_raises_validation_error_when_slot_generation_fails(self, mock_atomic):
        mock_atomic.return_value.__enter__.return_value = None
        mock_atomic.return_value.__exit__.return_value = None
        preview_result = RandomizationImportPreviewResult(
            columns=(),
            preview_rows=(),
            parsed_rows=(
                RandomizationImportParsedRow(2, "SCH-001 / ARM-A", {"scheme_code": "SCH-001", "arm_code": "ARM-A"}),
            ),
            total_rows=1,
            issues=(),
        )
        self.service.preview_service = MagicMock()
        self.service.preview_service.execute.return_value = preview_result
        self.service._build_scheme_map = MagicMock(return_value={"sch-001": SimpleNamespace(pk=11)})
        scheme = SimpleNamespace(pk=11, status="active")
        created_arm = SimpleNamespace(pk=201, is_active=True)
        self.service._upsert_arm = MagicMock(
            return_value=("created", scheme, created_arm, {}),
        )
        self.service._generate_slots_for_impacted_scheme = MagicMock(
            side_effect=RandomizationImportValidationError(
                (
                    RandomizationImportIssue(
                        row_number=2,
                        identifier="SCH-001 / ARM-A",
                        column_label="Scheme Code",
                        reason="Allocation Ratio references inactive or missing arm(s): B.",
                    ),
                ),
            )
        )

        with self.assertRaises(RandomizationImportValidationError) as exc_info:
            self.service.execute(self.command)

        self.assertEqual(exc_info.exception.issues[0].column_label, "Scheme Code")
        self.assertIn("missing arm", exc_info.exception.issues[0].reason)

    def test_upsert_arm_create_uses_expected_payload(self):
        repository = MagicMock()
        repository.get_arm_by_code.return_value = None
        service = CommitStudyRandomizationArmsImportService(repository=repository)
        parsed_row = RandomizationImportParsedRow(
            row_number=2,
            identifier="SCH-001 / ARM-A",
            values={
                "scheme_code": "SCH-001",
                "arm_code": "ARM-A",
                "arm_name": "Comparator",
                "target_count": 50,
                "display_order": 1,
            },
        )

        outcome = service._upsert_arm(
            parsed_row=parsed_row,
            scheme_map={"sch-001": SimpleNamespace(pk=11)},
            now=ANY,
        )

        self.assertEqual(outcome[0], "created")
        self.assertEqual(outcome[3], {})
        repository.create_arm.assert_called_once_with(
            scheme=ANY,
            arm_code="ARM-A",
            current_count=0,
            is_active=True,
            notes=None,
            created_at=ANY,
            arm_name="Comparator",
            target_count=50,
            display_order=1,
            deleted=False,
            updated_at=ANY,
        )

    def test_generate_slots_for_impacted_scheme_wraps_domain_error(self):
        repository = MagicMock()
        repository.list_active_arms_for_scheme.return_value = [SimpleNamespace(pk=201, arm_code="ARM-A")]
        slot_generation_service = MagicMock()
        slot_generation_service.generate_slots_for_scheme_arm.side_effect = RandomizationSlotGenerationError(
            "Allocation Ratio references inactive or missing arm(s): B.",
        )
        service = CommitStudyRandomizationArmsImportService(
            repository=repository,
            slot_generation_service=slot_generation_service,
        )

        with self.assertRaises(RandomizationImportValidationError) as exc_info:
            service._generate_slots_for_impacted_scheme(
                scheme=SimpleNamespace(pk=11, status="active"),
                parsed_row=RandomizationImportParsedRow(
                    row_number=2,
                    identifier="SCH-001 / ARM-A",
                    values={"scheme_code": "SCH-001", "arm_code": "ARM-A"},
                ),
            )

        self.assertEqual(exc_info.exception.issues[0].row_number, 2)
        self.assertEqual(exc_info.exception.issues[0].column_label, "Scheme Code")
        self.assertIn("B", exc_info.exception.issues[0].reason)


class StudyRandomizationImportBaseViewTests(SimpleTestCase):
    def test_render_format_error_hides_raw_exception_message(self):
        response = StudyRandomizationImportBaseView.render_format_error(
            RandomizationImportFormatError("db_password=secret"),
        )

        payload = json.loads(response.content.decode("utf-8"))
        self.assertEqual(response.status_code, 400)
        self.assertEqual(
            payload["detail"],
            "The uploaded file format is invalid. Please review the template and try again.",
        )
        self.assertNotIn("db_password", payload["detail"])

    def test_render_format_error_returns_safe_dependency_message(self):
        response = StudyRandomizationImportBaseView.render_format_error(
            RandomizationImportDependencyError("openpyxl missing on C:/venv/private"),
        )

        payload = json.loads(response.content.decode("utf-8"))
        self.assertEqual(response.status_code, 400)
        self.assertEqual(
            payload["detail"],
            "Import processing is temporarily unavailable. Please contact support.",
        )
        self.assertNotIn("openpyxl", payload["detail"])

    def test_serialize_validation_error_hides_exception_message_and_keeps_issues(self):
        view = StudyRandomizationCommitBaseView()
        exc = RandomizationImportValidationError(
            (
                RandomizationImportIssue(
                    row_number=2,
                    identifier="SCH-001",
                    column_label="Code",
                    reason="Code is required.",
                ),
            ),
        )

        response = view.serialize_validation_error(exc)

        payload = json.loads(response.content.decode("utf-8"))
        self.assertEqual(response.status_code, 400)
        self.assertEqual(payload["detail"], "The uploaded file contains validation issues.")
        self.assertEqual(len(payload["issues"]), 1)
        self.assertEqual(payload["issues"][0]["column_label"], "Code")
        self.assertEqual(payload["issues"][0]["reason"], "Code is required.")
