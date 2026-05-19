from contextlib import nullcontext
from pathlib import Path
from unittest.mock import ANY, MagicMock, patch

from django.test import SimpleTestCase
from openpyxl import load_workbook

from apps.core.choices import (
    EventDefinitionCategoryChoices,
    EventDefinitionTimingModeChoices,
    EventDefinitionTypeChoices,
    EventExecutionModeChoices,
    EventTransitionConditionScopeChoices,
    EventTransitionTypeChoices,
    StudyConditionDefinitionScopeChoices,
    StudyConditionDefinitionStatusChoices,
)
from apps.study.application.services import ImportStudyEventDefinitionsTemplateService


class ImportStudyEventDefinitionsTemplateServiceTests(SimpleTestCase):
    def setUp(self):
        self.repository = MagicMock()
        self.service = ImportStudyEventDefinitionsTemplateService(repository=self.repository)
        self.event_definition = MagicMock(name="event_definition")

    def test_static_template_headers_match_expected_columns(self):
        workbook = load_workbook(
            Path(__file__).resolve().parents[2] / "src/staticfiles/study/templates/event_definitions_import_template.xlsx",
            read_only=True,
        )
        worksheet = workbook.worksheets[0]
        headers = list(next(worksheet.iter_rows(values_only=True)))

        self.assertEqual(headers, list(self.service.expected_columns))

    @patch("apps.study.application.services.import_event_definitions_template.transaction.atomic")
    @patch.object(ImportStudyEventDefinitionsTemplateService, "_sync_condition_definition")
    @patch.object(ImportStudyEventDefinitionsTemplateService, "_sync_transition_rule")
    @patch.object(ImportStudyEventDefinitionsTemplateService, "_resolve_study_version", return_value="v1.0")
    def test_import_row_maps_new_event_definition_and_transition_fields(
        self,
        mock_resolve_study_version,
        mock_sync_transition_rule,
        mock_sync_condition_definition,
        mock_atomic,
    ):
        mock_atomic.return_value = nullcontext()
        created_event_definition = MagicMock(name="created_event_definition")
        condition_definition = MagicMock(name="condition_definition")
        self.service.repository.get_event_definition_for_import.return_value = None
        self.service.repository.create_event_definition.return_value = created_event_definition
        mock_sync_condition_definition.return_value = condition_definition

        outcome = self.service._import_row(
            study_id=3,
            row_data={
                "study_version": "v1.0",
                "code": "RAND",
                "name": "Randomization",
                "description": "Randomize subject",
                "event_type": "operational",
                "timing_mode": "conditional",
                "event_category": "randomization",
                "execution_mode": "workflow action",
                "sequence_no": "7",
                "phase_code": "TREATMENT",
                "repeated": "no",
                "max_repeats": "",
                "required": "yes",
                "precondition": "SCREENING",
                "transition_type": "automatic",
                "condition_scope": "eligibility",
                "condition_code": "eligible",
                "condition_expression": '{"fact":"eligible"}',
                "offset_days": "2",
                "window_before_days": "1",
                "window_after_days": "3",
                "auto_open": "yes",
                "auto_create": "yes",
                "requires_previous_completion": "no",
                "allow_skip": "yes",
            },
            row_number=2,
            actor_user_id=99,
        )

        self.assertEqual(outcome, "created")
        mock_resolve_study_version.assert_called_once_with(study_id=3, raw_study_version="v1.0")
        self.service.repository.create_event_definition.assert_called_once_with(
            code="RAND",
            created_at=ANY,
            created_by_id=99,
            study_id=3,
            study_version="v1.0",
            name="Randomization",
            description="Randomize subject",
            event_type=EventDefinitionTypeChoices.OPERATIONAL,
            timing_mode=EventDefinitionTimingModeChoices.CONDITIONAL,
            event_category=EventDefinitionCategoryChoices.RANDOMIZATION,
            execution_mode=EventExecutionModeChoices.WORKFLOW_ACTION,
            sequence_no=7,
            phase_code="TREATMENT",
            is_repeating=False,
            max_repeats=None,
            is_enabled=True,
            is_required=True,
            deleted=False,
            updated_at=ANY,
            updated_by_id=99,
        )
        mock_sync_condition_definition.assert_called_once_with(
            study_id=3,
            study_version="v1.0",
            condition_code="eligible",
            condition_scope=StudyConditionDefinitionScopeChoices.ELIGIBILITY,
            condition_expression='{"fact":"eligible"}',
            actor_user_id=99,
            now=ANY,
        )
        mock_sync_transition_rule.assert_called_once_with(
            event_definition=created_event_definition,
            study_id=3,
            study_version="v1.0",
            sequence_no=7,
            precondition_code="SCREENING",
            transition_type=EventTransitionTypeChoices.AUTOMATIC,
            condition_scope=EventTransitionConditionScopeChoices.ELIGIBILITY,
            condition_code=None,
            condition_definition=condition_definition,
            offset_days=2,
            window_before_days=1,
            window_after_days=3,
            auto_open=True,
            auto_create=True,
            requires_previous_completion=False,
            allow_skip=True,
            actor_user_id=99,
            now=ANY,
        )

    @patch("apps.study.application.services.import_event_definitions_template.transaction.atomic")
    @patch.object(ImportStudyEventDefinitionsTemplateService, "_sync_transition_rule")
    @patch.object(ImportStudyEventDefinitionsTemplateService, "_resolve_study_version", return_value="v1.0")
    def test_import_row_ignores_transition_columns_when_precondition_is_blank(
        self,
        mock_resolve_study_version,
        mock_sync_transition_rule,
        mock_atomic,
    ):
        mock_atomic.return_value = nullcontext()
        created_event_definition = MagicMock(name="created_event_definition")
        self.service.repository.get_event_definition_for_import.return_value = None
        self.service.repository.create_event_definition.return_value = created_event_definition

        outcome = self.service._import_row(
            study_id=3,
            row_data={
                "study_version": "v1.0",
                "code": "VISIT1",
                "name": "Visit 1",
                "description": "",
                "event_type": "visit based",
                "timing_mode": "scheduled",
                "event_category": "",
                "execution_mode": "",
                "sequence_no": "2",
                "phase_code": "",
                "repeated": "no",
                "max_repeats": "",
                "required": "yes",
                "precondition": "",
                "transition_type": "not-a-real-type",
                "condition_scope": "still-invalid",
                "condition_code": "ignored",
                "offset_days": "bad-number",
                "window_before_days": "bad-number",
                "window_after_days": "bad-number",
                "auto_open": "maybe",
                "auto_create": "maybe",
                "requires_previous_completion": "maybe",
                "allow_skip": "maybe",
            },
            row_number=2,
            actor_user_id=99,
        )

        self.assertEqual(outcome, "created")
        mock_resolve_study_version.assert_called_once_with(study_id=3, raw_study_version="v1.0")
        mock_sync_transition_rule.assert_called_once_with(
            event_definition=created_event_definition,
            study_id=3,
            study_version="v1.0",
            sequence_no=2,
            precondition_code=None,
            transition_type=None,
            condition_scope=None,
            condition_code=None,
            condition_definition=None,
            offset_days=None,
            window_before_days=None,
            window_after_days=None,
            auto_open=False,
            auto_create=False,
            requires_previous_completion=True,
            allow_skip=False,
            actor_user_id=99,
            now=ANY,
        )

    def test_sync_transition_rule_soft_deletes_existing_rules_when_precondition_removed(self):
        self.service._sync_transition_rule(
            event_definition=self.event_definition,
            study_id=3,
            study_version="v1.0",
            sequence_no=2,
            precondition_code=None,
            transition_type=None,
            condition_scope=None,
            condition_code=None,
            condition_definition=None,
            offset_days=None,
            window_before_days=None,
            window_after_days=None,
            auto_open=False,
            auto_create=False,
            requires_previous_completion=True,
            allow_skip=False,
            actor_user_id=42,
            now="2026-04-13T21:45:00",
        )

        self.service.repository.soft_delete_transition_rules_for_event.assert_called_once_with(
            study_id=3,
            study_version="v1.0",
            to_event_definition=self.event_definition,
            actor_user_id=42,
            updated_at="2026-04-13T21:45:00",
        )

    def test_sync_transition_rule_replaces_stale_incoming_rule_when_precondition_changes(
        self,
    ):
        from_event_definition = MagicMock(name="from_event_definition")
        transition_rule = MagicMock()
        self.service.repository.get_active_event_definition_by_code.return_value = from_event_definition
        self.service.repository.get_transition_rule.return_value = transition_rule

        self.service._sync_transition_rule(
            event_definition=self.event_definition,
            study_id=3,
            study_version="v1.0",
            sequence_no=4,
            precondition_code="SCREENING",
            transition_type=EventTransitionTypeChoices.AUTOMATIC,
            condition_scope=EventTransitionConditionScopeChoices.ELIGIBILITY,
            condition_code="eligible",
            condition_definition=None,
            offset_days=5,
            window_before_days=1,
            window_after_days=3,
            auto_open=True,
            auto_create=True,
            requires_previous_completion=False,
            allow_skip=True,
            actor_user_id=84,
            now="2026-04-13T21:46:00",
        )

        self.service.repository.soft_delete_transition_rules_for_event.assert_called_once_with(
            study_id=3,
            study_version="v1.0",
            to_event_definition=self.event_definition,
            exclude_from_event_definition=from_event_definition,
            actor_user_id=84,
            updated_at="2026-04-13T21:46:00",
        )
        self.service.repository.get_transition_rule.assert_called_once_with(
            study_id=3,
            study_version="v1.0",
            to_event_definition=self.event_definition,
            from_event_definition=from_event_definition,
        )
        self.assertEqual(transition_rule.transition_type, EventTransitionTypeChoices.AUTOMATIC)
        self.assertEqual(transition_rule.condition_scope, EventTransitionConditionScopeChoices.ELIGIBILITY)
        self.assertEqual(transition_rule.condition_code, "eligible")
        self.assertIsNone(transition_rule.condition_definition)
        self.assertEqual(transition_rule.offset_days, 5)
        self.assertEqual(transition_rule.window_before_days, 1)
        self.assertEqual(transition_rule.window_after_days, 3)
        self.assertTrue(transition_rule.auto_open)
        self.assertTrue(transition_rule.auto_create)
        self.assertFalse(transition_rule.requires_previous_completion)
        self.assertTrue(transition_rule.allow_skip)
        self.service.repository.save_transition_rule.assert_called_once()

    def test_sync_condition_definition_creates_active_definition_from_template_columns(self):
        self.service.repository.get_condition_definition.return_value = None
        created_condition_definition = MagicMock(name="created_condition_definition")
        self.service.repository.create_condition_definition.return_value = created_condition_definition

        result = self.service._sync_condition_definition(
            study_id=3,
            study_version="v1.0",
            condition_code="eligible",
            condition_scope=StudyConditionDefinitionScopeChoices.ELIGIBILITY,
            condition_expression='{"fact":"eligible"}',
            actor_user_id=99,
            now="2026-05-19T10:00:00",
        )

        self.assertIs(result, created_condition_definition)
        self.service.repository.create_condition_definition.assert_called_once_with(
            study_id=3,
            study_version="v1.0",
            code="eligible",
            created_at="2026-05-19T10:00:00",
            created_by_id=99,
            scope=StudyConditionDefinitionScopeChoices.ELIGIBILITY,
            expression_json='{"fact":"eligible"}',
            status=StudyConditionDefinitionStatusChoices.ACTIVE,
            deleted=False,
            updated_at="2026-05-19T10:00:00",
            updated_by_id=99,
        )
