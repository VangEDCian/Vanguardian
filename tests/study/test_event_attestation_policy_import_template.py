from contextlib import nullcontext
from pathlib import Path
from types import SimpleNamespace
from unittest.mock import MagicMock, patch

from django.test import SimpleTestCase
from openpyxl import load_workbook

from apps.study.application.services import ImportStudyEventAttestationPoliciesTemplateService


CERTIFICATION_STATEMENT = (
    "I certify that the data entered in this eCRF are complete, accurate, "
    "and ae supported by the source documents"
)


class ImportStudyEventAttestationPoliciesTemplateServiceTests(SimpleTestCase):
    def setUp(self):
        self.repository = MagicMock()
        self.service = ImportStudyEventAttestationPoliciesTemplateService(
            repository=self.repository,
        )

    def test_static_template_headers_match_expected_columns(self):
        workbook = load_workbook(
            Path(__file__).resolve().parents[2]
            / "src/staticfiles/study/templates/event_attestation_policies_import_template.xlsx",
            read_only=True,
        )
        worksheet = workbook[self.service.sheet_name]
        headers = list(next(worksheet.iter_rows(values_only=True)))

        self.assertEqual(
            headers,
            list(self.service.expected_columns[self.service.sheet_name]),
        )

    def test_static_template_contains_screening_and_visit_certification_rows(self):
        workbook = load_workbook(
            Path(__file__).resolve().parents[2]
            / "src/staticfiles/study/templates/event_attestation_policies_import_template.xlsx",
            read_only=True,
            data_only=True,
        )
        worksheet = workbook[self.service.sheet_name]
        rows = list(worksheet.iter_rows(values_only=True))
        headers = rows[0]
        row_data = [dict(zip(headers, row, strict=True)) for row in rows[1:]]

        self.assertEqual(
            [row["Event Code"] for row in row_data],
            [
                "SCREENING",
                "VISIT1",
                "VISIT2",
                "VISIT3",
                "VISIT4",
                "VISIT5",
                "VISIT6",
                "VISIT7",
                "VISIT8",
                "VISIT9",
                "VISIT10",
                "VISIT11",
                "VISIT12",
                "VISIT13",
                "VISIT14",
                "VISIT15",
                "VISIT16",
                "FU",
                "ET",
            ],
        )
        for row in row_data:
            self.assertEqual(row["Study Version"], "v1.0")
            self.assertEqual(row["Attestation Action Kind"], "CERTIFICATION")
            self.assertEqual(row["Attestation Permission Code"], "EVENT_CERTIFICATION.CERTIFY")
            self.assertEqual(row["Attestation Requires Confirmation"], "TRUE")
            self.assertEqual(row["Attestation Invalidate On Data Change"], "TRUE")
            self.assertEqual(row["Attestation Invalidate On Scope Change"], "TRUE")
            self.assertEqual(row["Attestation Statement Text Vi"], CERTIFICATION_STATEMENT)
            self.assertEqual(row["Attestation Statement Text En"], CERTIFICATION_STATEMENT)

    @patch("apps.study.application.services.import_event_attestation_policies_template.transaction.atomic")
    def test_import_row_creates_certification_policy(self, mock_atomic):
        mock_atomic.return_value = nullcontext()
        event_definition = SimpleNamespace(pk=61, study_version="v1.0")
        attestation_policy = SimpleNamespace(pk=88)
        now = object()
        self.repository.list_active_event_definitions_by_code.return_value = [event_definition]
        self.repository.get_attestation_policy_for_import.return_value = None
        self.repository.create_attestation_policy.return_value = attestation_policy

        outcome = self.service._import_row(
            study_id=3,
            row_data={
                "event_code": "VISIT1",
                "study_version": "",
                "attestation_policy_code": "VISIT1_CERT",
                "attestation_action_kind": "certification",
                "attestation_gate_code": "DEFAULT",
                "attestation_display_order": "2",
                "attestation_permission_code": "EVENT_CERTIFICATION.CERTIFY",
                "attestation_requires_confirmation": "true",
                "attestation_statement_text_vi": "Toi xac nhan du lieu Visit 1 la day du.",
                "attestation_statement_text_en": "I certify Visit 1 data is complete.",
            },
            actor_user_id=99,
            now=now,
        )

        self.assertEqual(outcome, "created")
        self.repository.create_attestation_policy.assert_called_once_with(
            study_id=3,
            study_version="v1.0",
            event_definition=event_definition,
            code="VISIT1_CERT",
            created_at=now,
            created_by_id=99,
            updated_at=now,
            deleted=False,
            action_kind="CERTIFICATION",
            display_order=2,
            statement_code="VISIT1_CERT",
            statement_version="1",
            required_permission_code="EVENT_CERTIFICATION.CERTIFY",
            required_role_code=None,
            delegation_task_code=None,
            condition_definition=None,
            gate_code="DEFAULT",
            requires_confirmation_checkbox=True,
            requires_signature=False,
            requires_reauthentication=False,
            invalidate_on_data_change=True,
            invalidate_on_scope_change=True,
            invalidate_on_query_change=True,
            is_required_for_lock=False,
            is_enabled=True,
            updated_by_id=99,
        )
        self.assertEqual(self.repository.upsert_attestation_policy_translation.call_count, 2)
        self.repository.upsert_attestation_policy_translation.assert_any_call(
            attestation_policy=attestation_policy,
            language_code="vi",
            defaults={
                "dialog_title": "VISIT1_CERT",
                "action_label": "Certification",
                "statement_text": "Toi xac nhan du lieu Visit 1 la day du.",
                "confirmation_label": None,
                "success_message": None,
            },
        )
        self.repository.upsert_attestation_policy_translation.assert_any_call(
            attestation_policy=attestation_policy,
            language_code="en",
            defaults={
                "dialog_title": "VISIT1_CERT",
                "action_label": "Certification",
                "statement_text": "I certify Visit 1 data is complete.",
                "confirmation_label": None,
                "success_message": None,
            },
        )
