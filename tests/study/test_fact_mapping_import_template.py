from contextlib import nullcontext
from pathlib import Path
from types import SimpleNamespace
from unittest.mock import ANY, MagicMock, patch

from django.test import SimpleTestCase
from openpyxl import load_workbook

from apps.study.application.commands import (
    FactMappingImportConflictError,
    FactMappingImportFormatError,
    ImportStudyFactMappingsTemplateCommand,
)
from apps.study.application.services import ImportStudyFactMappingsTemplateService


class ImportStudyFactMappingsTemplateServiceTests(SimpleTestCase):
    def setUp(self):
        self.repository = MagicMock()
        self.crf_context_adapter = MagicMock()
        self.datacapture_fact_mapping_config_adapter = MagicMock()
        self.service = ImportStudyFactMappingsTemplateService(
            crf_context_adapter=self.crf_context_adapter,
            datacapture_fact_mapping_config_adapter=self.datacapture_fact_mapping_config_adapter,
            repository=self.repository,
        )

    def test_static_template_headers_match_expected_columns(self):
        workbook = load_workbook(
            Path(__file__).resolve().parents[2] / "src/staticfiles/study/templates/fact_mappings_import_template.xlsx",
            read_only=True,
        )
        worksheet = workbook.worksheets[0]
        headers = list(next(worksheet.iter_rows(values_only=True)))

        self.assertEqual(headers, list(self.service.expected_columns))

    @patch("apps.study.application.services.import_fact_mappings_template.transaction.atomic")
    def test_import_row_resolves_event_and_form_then_upserts_fact_mapping(self, mock_atomic):
        mock_atomic.return_value = nullcontext()
        event_definition = SimpleNamespace(pk=61, study_version="v1.0")
        form_definition = SimpleNamespace(pk=71)
        self.repository.list_active_event_definitions_by_code.return_value = [event_definition]
        self.crf_context_adapter.resolve_unique_template_by_code.return_value = form_definition
        self.datacapture_fact_mapping_config_adapter.upsert_fact_mapping.return_value = SimpleNamespace(
            outcome="created"
        )

        outcome = self.service._import_row(
            study_id=3,
            row_data={
                "event_code": "SCREENING",
                "form_code": "SCREENING_INCLUSION_CRITERIA",
                "field_code": "ELIGIBLE",
                "source_path": "",
                "fact_key": "screening.eligible",
                "operator": "is true",
                "expected_value": "",
                "value_type": "boolean",
                "default_value": "false",
                "display_order": "2",
            },
            actor_user_id=99,
            now=ANY,
        )

        self.assertEqual(outcome, "created")
        self.repository.list_active_event_definitions_by_code.assert_called_once_with(
            study_id=3,
            code="SCREENING",
        )
        self.crf_context_adapter.resolve_unique_template_by_code.assert_called_once_with(
            study_id=3,
            code="SCREENING_INCLUSION_CRITERIA",
            case_insensitive=True,
        )
        self.datacapture_fact_mapping_config_adapter.upsert_fact_mapping.assert_called_once_with(
            study_id=3,
            study_version="v1.0",
            event_definition_id=61,
            crf_template_id=71,
            field_code="ELIGIBLE",
            source_path="ELIGIBLE",
            fact_key="screening.eligible",
            operator="is_true",
            expected_value=None,
            value_type="boolean",
            default_value="false",
            display_order=2,
            actor_user_id=99,
            now=ANY,
        )

    def test_execute_collects_row_level_issues(self):
        command = ImportStudyFactMappingsTemplateCommand(
            actor_user_id=99,
            selected_study_id=3,
            study_id=3,
            file_name="fact_mappings.xlsx",
            file_content=b"content",
        )
        with (
            patch.object(
                self.service,
                "_load_rows_from_workbook",
                return_value=[
                    (2, {"event_code": "", "form_code": "FORM_A", "fact_key": "fact.a"}),
                    (3, {"event_code": "SCREENING", "form_code": "FORM_A", "fact_key": "fact.b"}),
                ],
            ),
            patch.object(
                self.service,
                "_import_row",
                side_effect=[
                    FactMappingImportFormatError("Event Code is required."),
                    "updated",
                ],
            ),
        ):
            result = self.service.execute(command)

        self.assertEqual(result.total_rows, 2)
        self.assertEqual(result.updated_count, 1)
        self.assertEqual(result.skipped_count, 1)
        self.assertEqual(result.issues[0].row_number, 2)
        self.assertEqual(result.issues[0].event_code, "")

    def test_execute_collects_conflict_issues(self):
        command = ImportStudyFactMappingsTemplateCommand(
            actor_user_id=99,
            selected_study_id=3,
            study_id=3,
            file_name="fact_mappings.xlsx",
            file_content=b"content",
        )
        with (
            patch.object(
                self.service,
                "_load_rows_from_workbook",
                return_value=[(2, {"event_code": "SCREENING", "form_code": "FORM_A", "fact_key": "fact.a"})],
            ),
            patch.object(
                self.service,
                "_import_row",
                side_effect=FactMappingImportConflictError("Fact Key already exists in this study/event/version scope and is bound to another form."),
            ),
        ):
            result = self.service.execute(command)

        self.assertEqual(result.total_rows, 1)
        self.assertEqual(result.created_count, 0)
        self.assertEqual(result.updated_count, 0)
        self.assertEqual(result.skipped_count, 1)
        self.assertIn("bound to another form", result.issues[0].reason)
