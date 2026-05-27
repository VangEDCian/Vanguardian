from pathlib import Path
from types import SimpleNamespace
from unittest.mock import MagicMock, call, patch

from django.test import SimpleTestCase
from openpyxl import load_workbook

from apps.study.application.services import (
    ImportStudyCrfSectionLayoutConfigsTemplateService,
    ImportStudyCrfTemplateFieldsTemplateService,
    ImportStudyCrfTemplatesTemplateService,
    ImportStudyCrfValidationRulesTemplateService,
    StudyCrfTemplateDirectoryQueryService,
)


def _make_crf_template(*, pk, code, name, version, is_active=True, updated_at=None):
    return SimpleNamespace(
        pk=pk,
        code=code,
        version=version,
        is_active=is_active,
        updated_at=updated_at,
        safe_translation_getter=lambda field_name, default="", any_language=False: name if field_name == "name" else default,
    )


class StudyCrfTemplateDirectoryQueryServiceTests(SimpleTestCase):
    def test_filters_using_translated_name(self):
        adapter = SimpleNamespace(
            list_study_templates_for_listing=lambda study_id: [
                _make_crf_template(pk=2, code="LAB", name="Laboratory", version="v2"),
                _make_crf_template(pk=1, code="AE", name="Adverse Event", version="v1"),
            ]
        )

        result = StudyCrfTemplateDirectoryQueryService(
            crf_context_adapter=adapter,
        ).list_crf_templates(
            study_id=3,
            search_query="adverse",
            sort_query="name",
        )

        self.assertEqual(result["crf_templates_total"], 1)
        self.assertEqual(
            result["crf_templates"][0].safe_translation_getter("name", default="", any_language=True),
            "Adverse Event",
        )


class ImportStudyCrfTemplatesTemplateServiceTests(SimpleTestCase):
    def setUp(self):
        self.service = ImportStudyCrfTemplatesTemplateService()

    def test_static_template_contains_required_sheet_and_headers(self):
        workbook = load_workbook(
            Path(__file__).resolve().parents[2] / "src/staticfiles/study/templates/crf_templates_import_template.xlsx",
            read_only=True,
        )

        self.assertEqual(workbook.sheetnames, ["Form Templates", "Section Templates"])
        self.assertEqual(
            list(next(workbook["Form Templates"].iter_rows(values_only=True))),
            list(self.service.expected_columns["Form Templates"]),
        )

    @patch.object(
        ImportStudyCrfTemplatesTemplateService,
        "_load_rows_from_workbook",
        return_value={
            "Form Templates": [
                (2, {"code": "AE", "vi_name": "Bien co bat loi", "en_name": "Adverse Event", "version": "v1"})
            ],
            "Section Templates": [],
        },
    )
    def test_execute_imports_crf_sheet(self, mock_load_rows):
        call_order = []

        def import_form_template_row(**kwargs):
            call_order.append(("crf", kwargs["row_number"]))
            return "created", 101, "AE"

        self.service._import_form_template_row = import_form_template_row

        result = self.service.execute(
            command=SimpleNamespace(
                actor_user_id=7,
                selected_study_id=3,
                study_id=3,
                file_name="crf_templates_import_template.xlsx",
                file_content=b"xlsx",
            )
        )

        self.assertEqual(result.total_rows, 1)
        self.assertEqual(result.created_count, 1)
        self.assertEqual(result.updated_count, 0)
        self.assertEqual(call_order, [("crf", 2)])

    def test_import_crf_template_row_calls_context_adapter(self):
        mock_adapter = MagicMock()
        mock_adapter.upsert_crf_template.return_value = "updated"
        mock_adapter.resolve_unique_template_by_code_version.return_value = SimpleNamespace(pk=17, code="AE")

        service = ImportStudyCrfTemplatesTemplateService(
            crf_context_adapter=mock_adapter,
        )

        outcome = service._import_form_template_row(
            selected_study_id=7,
            study_id=7,
            row_data={
                "code": "AE",
                "vi_name": "Bien co bat loi",
                "en_name": "Adverse Event",
                "version": "v1.0",
            },
            row_number=2,
            actor_user_id=11,
        )

        self.assertEqual(outcome, ("updated", 17, "AE"))
        mock_adapter.upsert_crf_template.assert_called_once()


class ImportStudyCrfTemplateFieldsTemplateServiceTests(SimpleTestCase):
    def setUp(self):
        self.service = ImportStudyCrfTemplateFieldsTemplateService()

    def test_static_field_template_contains_required_sheet_and_headers(self):
        workbook = load_workbook(
            Path(__file__).resolve().parents[2]
            / "src/staticfiles/study/templates/crf_template_fields_import_template.xlsx",
            read_only=True,
        )

        self.assertEqual(workbook.sheetnames, ["Template Fields"])
        self.assertEqual(
            list(next(workbook["Template Fields"].iter_rows(values_only=True))),
            list(self.service.expected_columns["Template Fields"]),
        )

    @patch.object(
        ImportStudyCrfTemplateFieldsTemplateService,
        "_load_rows_from_workbook",
        return_value={
            "Template Fields": [
                (
                    2,
                    {
                        "form_name": "AE",
                        "section_name": "General",
                        "field_name": "AETERM",
                        "field_description_vi": "Bien co bat loi",
                        "field_description_en": "Adverse Event Term",
                        "data_type": "TEXT",
                        "display_order": "1",
                        "control_type": "TEXT",
                    },
                )
            ],
        },
    )
    def test_execute_imports_template_field_through_crf_adapter(self, mock_load_rows):
        form_template = SimpleNamespace(pk=17)
        section_template = SimpleNamespace(pk=23)
        mock_adapter = MagicMock()
        mock_adapter.resolve_import_template_by_name_or_code.return_value = form_template
        mock_adapter.resolve_import_section_by_name_or_code.return_value = section_template
        mock_adapter.reset_import_template_fields.return_value = 0
        mock_adapter.upsert_import_template_field.return_value = ("created", SimpleNamespace(pk=31))
        service = ImportStudyCrfTemplateFieldsTemplateService(crf_context_adapter=mock_adapter)

        result = service.execute(
            command=SimpleNamespace(
                actor_user_id=7,
                selected_study_id=3,
                study_id=3,
                file_name="crf_template_fields_import_template.xlsx",
                file_content=b"xlsx",
            )
        )

        self.assertEqual(result.total_rows, 1)
        self.assertEqual(result.created_count, 1)
        self.assertEqual(result.updated_count, 0)
        self.assertEqual(result.skipped_count, 0)
        mock_adapter.resolve_import_template_by_name_or_code.assert_called_once_with(
            study_id=3,
            form_name="AE",
        )
        mock_adapter.resolve_import_section_by_name_or_code.assert_called_once_with(
            crf_template_id=17,
            section_name="General",
        )
        mock_adapter.reset_import_template_fields.assert_called_once()
        self.assertEqual(
            mock_adapter.method_calls[-2:],
            [
                call.reset_import_template_fields(
                    crf_template_id=17,
                    actor_user_id=7,
                    now=mock_adapter.reset_import_template_fields.call_args.kwargs["now"],
                ),
                call.upsert_import_template_field(
                    crf_template_id=17,
                    section_template_id=23,
                    payload=mock_adapter.upsert_import_template_field.call_args.kwargs["payload"],
                    actor_user_id=7,
                    now=mock_adapter.reset_import_template_fields.call_args.kwargs["now"],
                ),
            ],
        )
        mock_adapter.upsert_import_template_field.assert_called_once()
        payload = mock_adapter.upsert_import_template_field.call_args.kwargs["payload"]
        self.assertEqual(payload["field_key"], "AETERM")
        self.assertEqual(payload["label_vi"], "Bien co bat loi")
        self.assertEqual(payload["label_en"], "Adverse Event Term")

    @patch.object(
        ImportStudyCrfTemplateFieldsTemplateService,
        "_load_rows_from_workbook",
        return_value={
            "Template Fields": [
                (
                    2,
                    {
                        "form_name": "AE",
                        "section_name": "General",
                        "field_name": "AETERM",
                        "data_type": "TEXT",
                        "display_order": "1",
                        "control_type": "TEXT",
                    },
                ),
                (
                    3,
                    {
                        "form_name": "AE",
                        "section_name": "General",
                        "field_name": "AESTDAT",
                        "data_type": "DATE",
                        "display_order": "2",
                        "control_type": "DATE",
                    },
                ),
            ],
        },
    )
    def test_execute_resets_template_fields_once_per_form_before_importing_rows(self, mock_load_rows):
        form_template = SimpleNamespace(pk=17)
        section_template = SimpleNamespace(pk=23)
        mock_adapter = MagicMock()
        mock_adapter.resolve_import_template_by_name_or_code.return_value = form_template
        mock_adapter.resolve_import_section_by_name_or_code.return_value = section_template
        mock_adapter.reset_import_template_fields.return_value = 4
        mock_adapter.upsert_import_template_field.side_effect = [
            ("updated", SimpleNamespace(pk=31)),
            ("created", SimpleNamespace(pk=32)),
        ]
        service = ImportStudyCrfTemplateFieldsTemplateService(crf_context_adapter=mock_adapter)

        result = service.execute(
            command=SimpleNamespace(
                actor_user_id=7,
                selected_study_id=3,
                study_id=3,
                file_name="crf_template_fields_import_template.xlsx",
                file_content=b"xlsx",
            )
        )

        self.assertEqual(result.total_rows, 2)
        self.assertEqual(result.created_count, 1)
        self.assertEqual(result.updated_count, 1)
        self.assertEqual(result.skipped_count, 0)
        mock_adapter.reset_import_template_fields.assert_called_once()
        self.assertEqual(mock_adapter.upsert_import_template_field.call_count, 2)


class ImportStudyCrfValidationRulesTemplateServiceTests(SimpleTestCase):
    def setUp(self):
        self.service = ImportStudyCrfValidationRulesTemplateService()

    def test_static_validation_rules_template_contains_required_sheet_and_headers(self):
        workbook = load_workbook(
            Path(__file__).resolve().parents[2]
            / "src/staticfiles/study/templates/crf_validation_rules_import_template.xlsx",
            read_only=True,
        )

        self.assertEqual(workbook.sheetnames, ["Validation Rules"])
        self.assertEqual(
            list(next(workbook["Validation Rules"].iter_rows(values_only=True))),
            list(self.service.expected_columns["Validation Rules"]),
        )

    @patch.object(
        ImportStudyCrfValidationRulesTemplateService,
        "_load_rows_from_workbook",
        return_value={
            "Validation Rules": [
                (
                    2,
                    {
                        "study": "REACT-AF",
                        "form_code": "AE",
                        "field_name": "AETERM",
                        "rule_type": "REQUIRED",
                        "expression": None,
                        "severity": "error",
                        "mode": "HARD",
                        "vi_message": "Bat buoc",
                        "en_message": "Required",
                    },
                )
            ],
        },
    )
    def test_execute_imports_validation_rule_through_crf_adapter(self, mock_load_rows):
        form_template = SimpleNamespace(pk=17)
        field_template = SimpleNamespace(pk=23)
        mock_adapter = MagicMock()
        mock_adapter.resolve_import_validation_rule_template_by_code.return_value = form_template
        mock_adapter.resolve_import_validation_rule_field_by_key.return_value = field_template
        mock_adapter.upsert_import_validation_rule.return_value = ("created", SimpleNamespace(pk=31))
        study_repository = MagicMock()
        study_repository.get_study_by_code.return_value = SimpleNamespace(pk=3, code="REACT-AF")
        service = ImportStudyCrfValidationRulesTemplateService(
            crf_context_adapter=mock_adapter,
            study_repository=study_repository,
        )

        result = service.execute(
            command=SimpleNamespace(
                actor_user_id=7,
                selected_study_id=3,
                study_id=3,
                file_name="crf_validation_rules_import_template.xlsx",
                file_content=b"xlsx",
            )
        )

        self.assertEqual(result.total_rows, 1)
        self.assertEqual(result.created_count, 1)
        self.assertEqual(result.updated_count, 0)
        self.assertEqual(result.skipped_count, 0)
        study_repository.get_study_by_code.assert_called_once_with(code="REACT-AF")
        mock_adapter.resolve_import_validation_rule_template_by_code.assert_called_once_with(
            study_id=3,
            form_code="AE",
        )
        mock_adapter.resolve_import_validation_rule_field_by_key.assert_called_once_with(
            crf_template_id=17,
            field_name="AETERM",
        )
        mock_adapter.upsert_import_validation_rule.assert_called_once_with(
            study_id=3,
            crf_template_id=17,
            field_template_id=23,
            rule_type="REQUIRED",
            expression="",
            severity="error",
            mode="HARD",
            vi_message="Bat buoc",
            en_message="Required",
            actor_user_id=7,
            now=mock_adapter.upsert_import_validation_rule.call_args.kwargs["now"],
        )

    @patch.object(
        ImportStudyCrfValidationRulesTemplateService,
        "_load_rows_from_workbook",
        return_value={
            "Validation Rules": [
                (
                    2,
                    {
                        "study": "OTHER-STUDY",
                        "form_code": "AE",
                        "field_name": "AETERM",
                        "rule_type": "REQUIRED",
                        "expression": "$val != ''",
                        "severity": "error",
                        "mode": "HARD",
                        "vi_message": "Bat buoc",
                        "en_message": "Required",
                    },
                )
            ],
        },
    )
    def test_execute_skips_row_when_study_scope_does_not_match(self, mock_load_rows):
        mock_adapter = MagicMock()
        study_repository = MagicMock()
        study_repository.get_study_by_code.return_value = SimpleNamespace(pk=4, code="OTHER-STUDY")
        service = ImportStudyCrfValidationRulesTemplateService(
            crf_context_adapter=mock_adapter,
            study_repository=study_repository,
        )

        result = service.execute(
            command=SimpleNamespace(
                actor_user_id=7,
                selected_study_id=3,
                study_id=3,
                file_name="crf_validation_rules_import_template.xlsx",
                file_content=b"xlsx",
            )
        )

        self.assertEqual(result.created_count, 0)
        self.assertEqual(result.skipped_count, 1)
        self.assertIn("study scope", result.issues[0].reason)
        mock_adapter.upsert_import_validation_rule.assert_not_called()


class ImportStudyCrfSectionLayoutConfigsTemplateServiceTests(SimpleTestCase):
    def setUp(self):
        self.service = ImportStudyCrfSectionLayoutConfigsTemplateService()

    def test_static_section_layout_config_template_contains_required_sheet_and_headers(self):
        workbook = load_workbook(
            Path(__file__).resolve().parents[2]
            / "src/staticfiles/study/templates/crf_section_layout_configs_import_template.xlsx",
            read_only=True,
        )

        self.assertEqual(workbook.sheetnames, ["Section Layout Configs"])
        self.assertEqual(
            list(next(workbook["Section Layout Configs"].iter_rows(values_only=True))),
            list(self.service.expected_columns["Section Layout Configs"]),
        )

    @patch.object(
        ImportStudyCrfSectionLayoutConfigsTemplateService,
        "_load_rows_from_workbook",
        return_value={
            "Section Layout Configs": [
                (
                    2,
                    {
                        "form_name": "AE",
                        "section_name": "General",
                        "layout_type": "repeat_table",
                        "column_count": "2",
                        "label_position": "top",
                        "density": "compact",
                        "section_style": "plain",
                        "is_collapsible": "yes",
                        "is_expanded_by_default": "no",
                        "show_section_header": "true",
                        "show_border": "false",
                        "show_background": "true",
                        "custom_css_class": "ae-layout",
                        "custom_layout_schema": '{"columns": [{"field": "AETERM"}]}',
                    },
                )
            ],
        },
    )
    def test_execute_imports_section_layout_config_through_crf_adapter(self, mock_load_rows):
        form_template = SimpleNamespace(pk=17)
        section_template = SimpleNamespace(pk=23)
        mock_adapter = MagicMock()
        mock_adapter.resolve_import_template_by_name_or_code.return_value = form_template
        mock_adapter.resolve_import_section_by_name_or_code.return_value = section_template
        mock_adapter.upsert_section_layout_config.return_value = "created"
        service = ImportStudyCrfSectionLayoutConfigsTemplateService(crf_context_adapter=mock_adapter)

        result = service.execute(
            command=SimpleNamespace(
                actor_user_id=7,
                selected_study_id=3,
                study_id=3,
                file_name="crf_section_layout_configs_import_template.xlsx",
                file_content=b"xlsx",
            )
        )

        self.assertEqual(result.total_rows, 1)
        self.assertEqual(result.created_count, 1)
        self.assertEqual(result.updated_count, 0)
        self.assertEqual(result.skipped_count, 0)
        mock_adapter.resolve_import_template_by_name_or_code.assert_called_once_with(
            study_id=3,
            form_name="AE",
        )
        mock_adapter.resolve_import_section_by_name_or_code.assert_called_once_with(
            crf_template_id=17,
            section_name="General",
        )
        mock_adapter.upsert_section_layout_config.assert_called_once()
        payload = mock_adapter.upsert_section_layout_config.call_args.kwargs
        self.assertEqual(payload["section_template_id"], 23)
        self.assertEqual(payload["layout_type"], "repeat_table")
        self.assertEqual(payload["column_count"], 2)
        self.assertIs(payload["is_collapsible"], True)
        self.assertIs(payload["is_expanded_by_default"], False)
        self.assertEqual(payload["custom_layout_schema"], {"columns": [{"field": "AETERM"}]})
