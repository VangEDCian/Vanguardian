from contextlib import nullcontext
from pathlib import Path
from types import SimpleNamespace
from unittest.mock import MagicMock, patch

from django.test import SimpleTestCase
from openpyxl import load_workbook

from apps.shared.constants import EventFormEntryModeChoices
from apps.study.application.services.import_event_form_bindings_template import (
    ImportStudyEventFormBindingsTemplateService,
)


class ImportStudyEventFormBindingsTemplateServiceTests(SimpleTestCase):
    def setUp(self):
        self.repository = MagicMock()
        self.display_label_service = MagicMock()
        self.service = ImportStudyEventFormBindingsTemplateService(
            repository=self.repository,
            display_label_service=self.display_label_service,
        )

    def test_static_template_headers_match_expected_columns(self):
        workbook = load_workbook(
            Path(__file__).resolve().parents[2] / "src/staticfiles/study/templates/event_form_bindings_import_template.xlsx",
            read_only=True,
        )
        worksheet = workbook.worksheets[0]
        headers = list(next(worksheet.iter_rows(values_only=True)))

        self.assertEqual(headers, list(self.service.expected_columns))

    def test_static_template_sample_rows_use_repo_field_keys(self):
        workbook = load_workbook(
            Path(__file__).resolve().parents[2] / "src/staticfiles/study/templates/event_form_bindings_import_template.xlsx",
            read_only=True,
        )
        worksheet = workbook.worksheets[0]
        rows = list(worksheet.iter_rows(min_row=2, max_row=4, values_only=True))

        self.assertEqual(rows[0][1], "AE")
        self.assertEqual(rows[0][10], "AE #{{repeat_index}} - {{field:AE_TERM}}")
        self.assertEqual(rows[1][1], "CM")
        self.assertEqual(rows[1][10], "CM #{{repeat_index}} - {{field:MED_NAME}}")

    @patch.object(
        ImportStudyEventFormBindingsTemplateService,
        "_load_rows_from_workbook",
        return_value=[
            (
                2,
                {
                    "event_code": "SCREENING",
                    "form_code": "AE",
                    "display_order": "1",
                    "repeatable": "false",
                    "role_scope": "INVESTIGATOR",
                    "entry_mode": "single",
                    "display_label_enabled": "true",
                    "display_label_max_length": "120",
                    "display_label_use_choice_label": "true",
                    "display_label_empty_value_policy": "fallback",
                    "display_label_template_vi": "AE #{{repeat_index}} - {{field:AE_TERM}}",
                    "display_label_fallback_template_vi": "{{form_name}} #{{repeat_index}}",
                    "display_label_empty_text_vi": "",
                    "display_label_template_en": "AE #{{repeat_index}} - {{field:AE_TERM}}",
                    "display_label_fallback_template_en": "{{form_name}} #{{repeat_index}}",
                    "display_label_empty_text_en": "",
                },
            )
        ],
    )
    @patch("apps.study.application.services.import_event_form_bindings_template.transaction.atomic", return_value=nullcontext())
    def test_execute_replaces_existing_bindings_for_target_event_before_importing(
        self,
        _mock_atomic,
        mock_load_rows,
    ):
        event_definition = SimpleNamespace(pk=17, study_version="v1")
        form_definition = SimpleNamespace(pk=23)
        binding = SimpleNamespace(pk=91)
        repository = self.repository
        repository.get_event_form_binding.return_value = None
        repository.create_event_form_binding.return_value = binding
        crf_context_adapter = MagicMock()
        crf_context_adapter.resolve_unique_template_by_code.return_value = form_definition
        service = ImportStudyEventFormBindingsTemplateService(
            crf_context_adapter=crf_context_adapter,
            repository=repository,
            display_label_service=self.display_label_service,
        )
        service._resolve_event_definition = MagicMock(return_value=event_definition)

        result = service.execute(
            command=SimpleNamespace(
                actor_user_id=7,
                selected_study_id=3,
                study_id=3,
                file_name="event_form_bindings.xlsx",
                file_content=b"xlsx",
            )
        )

        self.assertEqual(result.total_rows, 1)
        self.assertEqual(result.created_count, 1)
        self.assertEqual(result.updated_count, 0)
        self.assertEqual(result.skipped_count, 0)
        repository.soft_delete_event_form_bindings_for_import.assert_called_once_with(
            event_definition_ids=[17],
            actor_user_id=7,
            updated_at=repository.soft_delete_event_form_bindings_for_import.call_args.kwargs["updated_at"],
        )
        repository.create_event_form_binding.assert_called_once_with(
            event_definition_id=17,
            form_definition_id=23,
            created_at=repository.create_event_form_binding.call_args.kwargs["created_at"],
            created_by_id=7,
            study_id=3,
            study_version="v1",
            display_order=1,
            is_repeatable_within_event=False,
            role_scope="INVESTIGATOR",
            entry_mode=EventFormEntryModeChoices.SINGLE,
            is_required=True,
            is_enabled=True,
            deleted=False,
            updated_at=repository.create_event_form_binding.call_args.kwargs["updated_at"],
            updated_by_id=7,
        )
        self.display_label_service.save_config.assert_called_once_with(
            binding_id=91,
            actor_user_id=7,
            is_enabled=True,
            max_length=120,
            use_choice_display_label=True,
            empty_value_policy="FALLBACK",
            translations={
                "vi": {
                    "label_template": "AE #{{repeat_index}} - {{field:AE_TERM}}",
                    "fallback_template": "{{form_name}} #{{repeat_index}}",
                    "empty_value_text": "",
                },
                "en": {
                    "label_template": "AE #{{repeat_index}} - {{field:AE_TERM}}",
                    "fallback_template": "{{form_name}} #{{repeat_index}}",
                    "empty_value_text": "",
                },
            },
        )
        mock_load_rows.assert_called_once()

    @patch.object(
        ImportStudyEventFormBindingsTemplateService,
        "_load_rows_from_workbook",
        return_value=[
            (
                2,
                {
                    "event_code": "SCREENING",
                    "form_code": "AE",
                    "display_order": "2",
                    "repeatable": "true",
                    "role_scope": "",
                    "entry_mode": "double entry",
                    "display_label_enabled": "",
                    "display_label_max_length": "",
                    "display_label_use_choice_label": "",
                    "display_label_empty_value_policy": "",
                    "display_label_template_vi": "",
                    "display_label_fallback_template_vi": "",
                    "display_label_empty_text_vi": "",
                    "display_label_template_en": "",
                    "display_label_fallback_template_en": "",
                    "display_label_empty_text_en": "",
                },
            )
        ],
    )
    @patch("apps.study.application.services.import_event_form_bindings_template.transaction.atomic", return_value=nullcontext())
    def test_execute_reactivates_binding_when_reset_import_updates_existing_row(
        self,
        _mock_atomic,
        mock_load_rows,
    ):
        event_definition = SimpleNamespace(pk=17, study_version="v1")
        form_definition = SimpleNamespace(pk=23)
        binding = SimpleNamespace(pk=52)
        repository = self.repository
        repository.get_event_form_binding.return_value = binding
        crf_context_adapter = MagicMock()
        crf_context_adapter.resolve_unique_template_by_code.return_value = form_definition
        service = ImportStudyEventFormBindingsTemplateService(
            crf_context_adapter=crf_context_adapter,
            repository=repository,
            display_label_service=self.display_label_service,
        )
        service._resolve_event_definition = MagicMock(return_value=event_definition)

        result = service.execute(
            command=SimpleNamespace(
                actor_user_id=7,
                selected_study_id=3,
                study_id=3,
                file_name="event_form_bindings.xlsx",
                file_content=b"xlsx",
            )
        )

        self.assertEqual(result.created_count, 0)
        self.assertEqual(result.updated_count, 1)
        self.assertEqual(result.skipped_count, 0)
        self.assertEqual(binding.study_id, 3)
        self.assertEqual(binding.study_version, "v1")
        self.assertEqual(binding.display_order, 2)
        self.assertTrue(binding.is_repeatable_within_event)
        self.assertIsNone(binding.role_scope)
        self.assertEqual(binding.entry_mode, "double_entry")
        self.assertTrue(binding.is_required)
        self.assertTrue(binding.is_enabled)
        self.assertFalse(binding.deleted)
        repository.save_event_form_binding.assert_called_once()
        update_fields = repository.save_event_form_binding.call_args.kwargs["update_fields"]
        self.assertIn("deleted", update_fields)
        self.assertIn("is_enabled", update_fields)
        self.assertIn("is_required", update_fields)
        self.display_label_service.save_config.assert_not_called()
        mock_load_rows.assert_called_once()
